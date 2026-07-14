"""
run_A4_auto_v39.py  --  Simplified 8-step pipeline (walk-path-y classification)
===================================================

Steps
-----
1. Legend detection: detect legend panel, extract curve colours, remove from image
2. Colour discovery: legend-based (if found) or hue-histogram fallback
3. Per-colour mask extraction (HSV range from discovered colours)
4. Per-colour mask cleaning: remove text + dashed-line artefacts
5. Walk-path-y constrained pixel classification
   - 1st pass walk on full clean mask -> get walk_y per x
   - Reclassify pixels: within walk_y +/-WALK_DENSITY_R -> curve/blob; outside -> stem
   - Detect mode_xs (blob columns) from x-density spikes on walk_mask
   - 2nd pass walk on reclassified walk_mask (curve+blob only)
6. Walk (A4 orientation-aware walk on curve-body + blob mask)
7. Data-point estimation via circular-disk density along walk path (integrated)
8. Return data_points + tcaps, write detections.json

Changelog
---------
v39: Legend/colour robustness + web-app diagnostics.
     - Performance: vectorised the per-pixel colour classification. _color_at /
       _density / _seg_y now read precomputed numpy membership + box-filtered
       density maps instead of 5M+ pure-Python calls; matplotlib DEBUG dumps
       (points_combined.jpg, aligned_density_montage.jpg) are gated behind
       DEBUG_DUMPS=1. ~2x faster (13.5s -> 6.5s) with identical output.
     - Curves are now LOCKED to the legend colours: when the unified grid is a
       1:1 legend match, the extraction palette is rebuilt from the grid's filled
       cells (near-duplicates merged) and NO black sink is added if the grid has
       an achromatic cell. So N legend colours => exactly N curves; stray colours
       from open/dashed markers, LLOQ lines, error bars are no longer invented.
       (Also recovered Picture15's dropped colour -> working set 9/9.)
     - Unified legend grid now ENFORCES a regular row lattice: from the detected
       rows it estimates the pitch and inserts missing MIDDLE rows at their true
       positions, then re-samples the image at every empty (col,row) intersection
       to fill swatches the blob detector dropped (faint / open / dashed / black
       markers). Genuinely blank cells stay empty. Recovers e.g. a black marker
       whose row was found but whose swatch was missed.
     - edit_data.json for the GUI point editor: per-curve pixel points, legend
       label + colour, and a serialisable calibration (2 pixel<->value anchors +
       log flags) so the browser can recompute data values after manual edits.
     - Data-point overlay no longer draws the legend table on itself (the editor
       and the standalone legend_overlay.png cover that).
     - Multi-COLUMN legend under a user box: the row-band sampler could stop at
       exactly LEGEND_MIN_ENTRIES (e.g. 3) and skip the 2D detector, truncating a
       2-column 5-colour legend to 3 and leaving the table undrawn. Now, whenever
       a user legend box is given, the 2D unified detector is always cross-checked
       and the richer palette is kept -> full colour count + table restored.
     - User-drawn --legend-box is now AUTHORITATIVE: the unified detector is
       confined to it (restrict_box) and never overwrites it, so a legend in a
       multi-panel figure is read where the user marked it, not elsewhere.
     - Confusable colour pairs that differ mainly in brightness (red vs orange,
       blue vs purple) are resolved by hue with a priority bias, so a bright
       anti-aliased edge of one no longer bleeds into the other's mask
       (e.g. orange mask red-contamination 34% -> 0%).
     - Per-colour masks (colormask_NN_<name>.png) with data points drawn on them,
       a colormasks.json manifest incl. the matching legend swatch RGB.
     - Legend-table overlay drawn on the data-point overlay itself, plus a
       standalone legend_overlay.png.
     - legend_diagnostic.png: shows every swatch candidate inside the legend box
       and why it was kept/dropped (accepted / near-neutral / no-label / bad-size),
       so legend-detection failures are visible without reading the log.
     - Walk NMS is now tied to the MARKER SIZE measured from the legend
       swatches (vertical ink extent at each legend cell, with the y-window
       capped to half the row spacing so it can't bleed into adjacent rows),
       instead of a fixed 8px guess. NMS = 0.5x that marker diameter, so dense
       markers on large plots aren't merged and sparse ones aren't split.
       (MARKER_DIAM env var still overrides.)
     - Default extraction mode = walk.
v38: Plot-area + legend-detection accuracy (P1/P3).
     P1 (_compute_plot_area): choose axis lines by POSITION (y-axis in the left
        region, x-axis in the bottom region) instead of a blind .max(), so a
        right-hand frame or a stray top rule can't collapse the box; add a data-
        bbox sanity fallback so the plot area can never collapse to a sliver.
     P3 (_pcm_detect_legend_unified): fold in ACHROMATIC (grey/black) swatches as
        EXTENSION-ONLY entries (they fill an empty grid slot such as a grey
        'Placebo' key but never SEED a legend), recovering missed rows without
        turning a row of black data markers into a false legend; reject candidate
        groups that span most of the plot (data-marker false positives).
v19: Legend-colour <-> colour-mask 1:1 matching. When a legend is present, the
     number of colour masks is forced to equal the number of legend swatches:
     improved legend panel/swatch extraction (full vertical span, robust band
     split) and a hard cap so extra fallback greys can't inflate the palette.

Usage
-----
    python3 run_A4_auto_v19.py <image_path> [output_dir]
"""
import cv2
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import os
import re
import sys
import json

# On Windows (and other legacy-encoded consoles), printing non-ASCII characters
# like ~, ->, +/- crashes with UnicodeEncodeError (e.g. cp949 on Korean Windows).
# Force UTF-8 output, or fall back to replacing unencodable chars, so the
# pipeline never dies on a debug print.
for _stream_name in ("stdout", "stderr"):
    _stream = getattr(sys, _stream_name, None)
    try:
        if _stream is not None and hasattr(_stream, "reconfigure"):
            _stream.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
from scipy.signal import find_peaks
from scipy.ndimage import gaussian_filter1d
try:
    import pytesseract
    from pytesseract import Output as _TESS_OUT
    _HAS_OCR = True
except Exception:
    _HAS_OCR = False

# -- Configuration --------------------------------------------------------------
LAM           = 0.5    # A4 walk lambda
R             = 8      # capsule disk radius (px)
WIN_W         = 25     # walk window width (wide for dashed-line gaps)
TAN75         = np.tan(np.radians(75))
TAN85         = np.tan(np.radians(85))  # wider angle for dashed/scatter plots

# Legend detection
LEGEND_RIGHT_FRAC   = 0.40
LEGEND_MAX_SWATCH_F = 0.18
LEGEND_MIN_PX       = 3
LEGEND_LAB_MERGE    = 30
LEGEND_PEAK_DIST    = 25
LEGEND_PEAK_PROM    = 2
LEGEND_MAX_ROW_FRAC = 0.65
LEGEND_MAX_SWATCH_W = 60
LEGEND_MIN_ENTRIES  = 3
LEGEND_BAND_PX_FRAC = 0.35  # drop swatch bands with < this fraction of median px
LEGEND_INK_MIN      = 6     # min vivid/dark pixels for a band to be a real swatch
LEGEND_SWATCH_DENSITY = 0.40  # min chrom_px/width ratio to accept a peak as swatch

# Hue-histogram fallback
HUE_SIGMA     = 2
HUE_MIN_DIST  = 8
HUE_TOLERANCE = 12
CHROM_MIN_PX  = 20
GREY_SAT_MAX  = 40   # allow slight saturation for anti-aliased dark pixels
GREY_VAL_MIN  = 20   # catch very dark pixels (near-black lines)
GREY_VAL_MAX  = 215  # exclude near-white background; raised slightly for mid-grey
GREY_MIN_PX   = 80
GREY_CLUSTER_SEP = 60   # min V-value gap to treat as separate grey/dark clusters
GREY_BLACK_VMAX  = 70   # grey clusters with peak V <= this are fused into 'black'
GREY_PEAK_DIST   = 60   # base; overridden by GREY_PEAK_DIST_EFF per background
DEDUP_LAB        = 14.0  # merge palette colours closer than this in Lab space
DEDUP_HUE        = 8     # merge chromatic colours within this hue gap (same curve)

# Nearest-centroid (mutually-exclusive) colour assignment.
# Because assignment is winner-take-all in Lab space, colour masks can NEVER
# overlap regardless of how large the radius is. So NC_MAX_DIST only needs to
# reject genuine noise (pixels far from every curve colour), not to separate
# colours from each other. A generous radius keeps anti-aliased / log-compressed
# curve-body pixels that sit 20-40 dE from the clean legend swatch colour.
NC_MAX_DIST       = 55.0   # noise cutoff only (exclusivity handles separation)
NC_MARGIN         = 1.0    # no boundary rejection (1.0 = accept the nearest)
NC_REFINE_SHRINK  = 0.90   # gentle shrink if a colour intrudes on another swatch
NC_REFINE_FLOOR   = 35.0   # never shrink a colour's radius below this
NC_MIN_LEGEND_ENTRIES = 4  # only trust nearest-centroid when the legend yields
                            # at least this many swatches (guards false legends)
# Chroma-adaptive radius: saturated colours get wide radius, achromatic get tight
NC_MIN_RADIUS     = 18.0   # radius for fully achromatic colours (grey/black)
NC_CHROMA_FULL    = 60.0   # chroma at/above which the full NC_MAX_DIST is used
NC_ACHROMATIC_CHROMA = 12.0  # chroma below this = treat colour as achromatic
# Saturation-floor gating (learned from legend swatch S, chromatic colours only)
SAT_FLOOR_PCTILE  = 20    # use this percentile of swatch S as the reference
SAT_FLOOR_FRAC    = 0.45  # accept pixels with S >= frac * reference (tolerates fade)
INK_CORE_PCT      = 35    # swatch "ink" = top INK_CORE_PCT% saturation (chromatic)
                          # or darkest INK_CORE_PCT% value (achromatic)
# Achromatic-aware distance (recovers anti-aliased thin black/grey curve lines)
NC_NEUTRAL_S_MAX  = 60    # a pixel is "neutral" (grey/black) if S <= this
NC_NEUTRAL_DARK_V = 50    # ...or if V <= this (near-black: S is unreliable here)
# Confusable chromatic pairs: when two centroids are within NC_CONFUSE_DIST in
# Lab, the more-vivid one wins contested boundary pixels via a distance bias.
NC_CONFUSE_DIST   = 50.0  # Lab dE below which two chromatic colours are "confusable"
NC_CONFUSE_BIAS   = 9.0   # (legacy) retained; pairs now resolved by L down-weighting
NC_CONFUSE_L_WEIGHT = 0.30  # L-axis weight when matching chromatic px to confusable colours
NC_ACHRO_L_WEIGHT = 0.25  # weight on L-axis when matching neutral px to grey/black
NC_ACHRO_V_MAX    = 170   # neutral px brighter than this can't be curve (page/halo)
# Shape-based achromatic cleanup (separates markers from axis/grid/error-bar lines)
SHAPE_LINE_SPAN    = 0.45  # component spanning > this fraction of image = line
SHAPE_LINE_ASPECT  = 8.0   # and aspect ratio above this = drop as axis/grid line
SHAPE_ERRBAR_MIN_LEN = 25  # thin (<=2px) strokes longer than this = error bar
_LEGEND_SWATCH_INFO = []   # stash of per-entry legend swatch samples + bbox
_LEGEND_BOX_FOR_CLEAN = None  # legend panel bbox, used to clear legend text from masks

# Axis detection
AXIS_ROW_FRAC        = 0.30
AXIS_COL_FRAC        = 0.25
AXIS_PAD             = 3
DENSITY_TICK_MAX_LEN = 12
DENSITY_MIN_TICKS    = 3

# Mask cleaning
CHAR_MAX_AREA       = 120
TEXT_CLUSTER_RADIUS = 60
TICK_AREA_MAX       = 200
TICK_MARGIN         = 15

# X-distribution pixel classification
XDIST_SIGMA       = 1.5   # gaussian smooth sigma for column histogram
XDIST_MODE_PROM   = 0.15  # min prominence as fraction of max count
XDIST_MODE_DIST   = 5     # min distance between modes (px)
BLOB_HEIGHT_RATIO = 2.5   # column cluster h/w > this -> stem candidate
BLOB_MIN_AREA     = 4     # min pixels in a column cluster to classify

# Walk density (data point estimation)
WALK_DENSITY_STEP  = 2
WALK_DENSITY_R     = 12
# Noise-colour filter: a mask is treated as noise (no real curve) if seeded
# traces disagree too much (agree > MAX) or the path barely spans x (xcov < MIN).
NOISE_AGREE_MAX    = 12.0
NOISE_XCOV_MIN     = 0.05
WALK_DENSITY_SIGMA = 2.0
WALK_DENSITY_PROM  = 0.04
WALK_DENSITY_DIST  = 15

# Endpoint injection
ENDPOINT_MERGE_DIST = 20

# -- Load image -----------------------------------------------------------------
import argparse as _argparse
_ap = _argparse.ArgumentParser()
_ap.add_argument('image', help='Input chart image')
_ap.add_argument('output_dir', nargs='?', default=None, help='Output directory')
_ap.add_argument('--legend', default=None, help='External legend image path')
_ap.add_argument('--no-stem', action='store_true', help='Skip stem/pixel classification; use raw mask directly for walk')
_ap.add_argument('--legend-box', default=None,
                 help='User-drawn legend box "x0,y0,x1,y1" (overrides auto-detect)')
_ap.add_argument('--plot-area', default=None,
                 help='User-drawn plot area "x0,y0,x1,y1" (overrides auto-detect)')
_ap.add_argument('--x-min', default=None, help='Manual x-axis minimum (data value at plot-area left edge)')
_ap.add_argument('--x-max', default=None, help='Manual x-axis maximum (data value at plot-area right edge)')
_ap.add_argument('--y-min', default=None, help='Manual y-axis minimum (data value at plot-area bottom edge)')
_ap.add_argument('--y-max', default=None, help='Manual y-axis maximum (data value at plot-area top edge)')
_ap.add_argument('--x-log', action='store_true', help='Treat x-axis as log scale')
_ap.add_argument('--y-log', action='store_true', help='Treat y-axis as log scale')
_args = _ap.parse_args()

def _parse_box(s):
    if not s:
        return None
    try:
        v = [int(round(float(t))) for t in s.replace(' ', '').split(',')]
        return tuple(v) if len(v) == 4 else None
    except Exception:
        return None

def _parse_num(s):
    if s is None or str(s).strip() == '':
        return None
    try:
        return float(s)
    except Exception:
        return None

USER_LEGEND_BOX = _parse_box(_args.legend_box)   # (x0,y0,x1,y1) or None
USER_PLOT_AREA  = _parse_box(_args.plot_area)    # (x0,y0,x1,y1) or None
USER_X_MIN = _parse_num(_args.x_min)
USER_X_MAX = _parse_num(_args.x_max)
USER_Y_MIN = _parse_num(_args.y_min)
USER_Y_MAX = _parse_num(_args.y_max)
USER_X_LOG = bool(_args.x_log)
USER_Y_LOG = bool(_args.y_log)

if not _args.image:
    print("Usage: python3 run_A4_auto_v24.py <image_path> [output_dir] [--legend legend_img]")
    sys.exit(1)
IMG_PATH = _args.image
LEGEND_IMG_PATH = _args.legend  # may be None
NO_STEM = _args.no_stem        # skip pixel classification when True
if _args.output_dir:
    OUT_DIR = _args.output_dir
else:
    _stem = os.path.splitext(os.path.basename(IMG_PATH))[0]
    # Strip 'pasted_file_' prefix and '_image' suffix for cleaner dir name
    _stem = re.sub(r'^pasted_file_', '', _stem)
    _stem = re.sub(r'_image$', '', _stem)
    OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(IMG_PATH)), f'{_stem}_v17_out')
os.makedirs(OUT_DIR, exist_ok=True)

img = cv2.imread(IMG_PATH)
if img is None:
    print(f"ERROR: cannot read {IMG_PATH}", file=sys.stderr)
    sys.exit(1)
img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
img_hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV).astype(np.int32)
img_lab = cv2.cvtColor(img, cv2.COLOR_BGR2Lab).astype(np.float32)
H, W    = img.shape[:2]
print(f"Image: {W}x{H}  ({IMG_PATH})")

# -- Background + axis detection ------------------------------------------------
def _detect_dark_background():
    border = np.concatenate([
        img_hsv[:5,  :, 2].ravel(), img_hsv[-5:, :, 2].ravel(),
        img_hsv[:,  :5, 2].ravel(), img_hsv[:, -5:, 2].ravel(),
    ])
    return float(np.median(border)) < 128

DARK_BG = _detect_dark_background()
print(f"Background: {'dark' if DARK_BG else 'light'}")

def _border_brightness():
    border = np.concatenate([
        img_hsv[:5,  :, 2].ravel(), img_hsv[-5:, :, 2].ravel(),
        img_hsv[:,  :5, 2].ravel(), img_hsv[:, -5:, 2].ravel(),
    ])
    return float(np.median(border))

# A grey-tinted background (e.g. V~230 instead of 255) makes faint grey clusters
# look like real curves. On such backgrounds we separate grey brightness modes
# more conservatively (larger required V-gap) to avoid splitting background haze
# into a phantom colour. On a clean white page we can separate aggressively, so a
# genuine grey curve (e.g. a 'Placebo' line) is told apart from black axes/text.
_BORDER_V = _border_brightness()
GREY_PEAK_DIST_EFF = 40 if _BORDER_V >= 248 else 60
print(f"Border brightness: {_BORDER_V:.0f}  ->  grey peak-distance = {GREY_PEAK_DIST_EFF}")

def _pcm_background_mask(hsv_arr):
    s = hsv_arr[:, :, 1]; v = hsv_arr[:, :, 2]
    if DARK_BG:
        return (v < 40)
    return (v > 240) & (s < 15)

BG_MASK = _pcm_background_mask(img_hsv)

def _detect_filled_plot_area():
    """Some plots fill the plotting region with a solid background colour that
    differs from the white page (e.g. a light-grey panel). That fill marks the
    plot area exactly -- and in such plots there is usually no separate axis line,
    so the fill's LEFT edge is the y-axis and its BOTTOM edge is the x-axis.

    Detect the largest solid non-white, low-saturation fill rectangle, bounding
    it by where the fill coverage is high (so faint anti-aliased margins or stray
    grey text don't stretch the box). Returns (x0,y0,x1,y1) or None.
    """
    v = img_hsv[:, :, 2]; s = img_hsv[:, :, 1]
    border = np.concatenate([img_rgb[0, :], img_rgb[-1, :],
                             img_rgb[:, 0], img_rgb[:, -1]]).reshape(-1, 3)
    bg = np.median(border, axis=0)
    if bg.mean() < 200:                  # only when the page is white-ish
        return None
    diff = np.linalg.norm(img_rgb.astype(int) - bg.astype(int), axis=2)
    fill = (s < 30) & (v >= 190) & (v <= 242) & (diff > 14)
    if fill.mean() < 0.06:
        return None
    rowcov = fill.mean(axis=1)
    colcov = fill.mean(axis=0)
    cpk = colcov.max(); rpk = rowcov.max()
    if cpk < 0.4 or rpk < 0.4:
        return None
    cols = np.where(colcov >= 0.5 * cpk)[0]
    rows = np.where(rowcov >= 0.5 * rpk)[0]
    if len(cols) < 5 or len(rows) < 5:
        return None
    def _largest_run(idx):
        idx = sorted(idx)
        runs = []; start = prev = idx[0]
        for i in idx[1:]:
            if i - prev <= 5:
                prev = i
            else:
                runs.append((start, prev)); start = prev = i
        runs.append((start, prev))
        return max(runs, key=lambda r: r[1] - r[0])
    x0, x1 = _largest_run(cols)
    y0, y1 = _largest_run(rows)
    if (x1 - x0) < 0.3 * W or (y1 - y0) < 0.3 * H:
        return None
    return (int(x0), int(y0), int(x1), int(y1))


def _detect_axes():
    v = img_hsv[:, :, 2]; s = img_hsv[:, :, 1]
    # Top priority: a COLOUR-FILLED plot panel (e.g. grey on white). In these
    # plots the axes are usually not drawn as separate lines -- the fill boundary
    # IS the frame. Use its left edge as the y-axis and its bottom edge as the
    # x-axis directly, since the fill rectangle is the most reliable signal.
    _fa = _detect_filled_plot_area()
    if _fa is not None:
        fx0, fy0, fx1, fy1 = _fa
        return np.array([fy1], dtype=int), np.array([fx0], dtype=int)

    # coloured (curve/marker) pixels -- used to verify that gaps in a broken axis
    # line are caused by data crossing it, not by it being text/empty space.
    chrom = (s > 40) & (v > 40) & (v < 250)
    whitish = v > 235

    def _max_run(px):
        m = run = 0
        for p in px:
            if p:
                run += 1; m = max(m, run)
            else:
                run = 0
        return m

    def _gap_colored_frac(line_px, chrom_px, white_px):
        """Of the GAPS (non-ink positions) within the ink's span, what fraction
        are covered by coloured data vs are white/empty? A real axis hidden by
        curves has coloured gaps; a text string has white gaps between glyphs."""
        nz = np.where(line_px)[0]
        if len(nz) < 2:
            return 0.0, 0.0
        a, b = nz[0], nz[-1]
        seg_chrom = chrom_px[a:b + 1].astype(bool)
        seg_white = white_px[a:b + 1].astype(bool)
        seg_ink = line_px[a:b + 1].astype(bool)
        gap = ~seg_ink
        ngap = int(gap.sum())
        if ngap == 0:
            return 1.0, 0.0
        return float((gap & seg_chrom).sum()) / ngap, float((gap & seg_white).sum()) / ngap

    def _detect_with(black):
        col_d = black.sum(axis=0).astype(float)
        row_d = black.sum(axis=1).astype(float)

        def _candidates(count_axis):
            """A real axis line is a LONG, mostly-continuous solid stroke. Its
            primary signature is a long continuous run (>=35% of the strip). If
            the run is shorter (curve/markers cover parts of it), we accept it
            only when the breaks are filled by COLOURED data -- that verification
            rejects long text (whose gaps are white) that would otherwise look
            axis-like by raw pixel count."""
            n = W if count_axis == 'col' else H
            L = H if count_axis == 'col' else W
            margin = 2
            out = []
            for rc in range(margin, n - margin):
                if count_axis == 'col':
                    line = black[:, rc]; cseg = chrom[:, rc]; wseg = whitish[:, rc]
                else:
                    line = black[rc, :]; cseg = chrom[rc, :]; wseg = whitish[rc, :]
                if line.sum() == 0:
                    continue
                run = _max_run(line)
                if run >= L * 0.35:
                    out.append((rc, run, 2.0))        # clean solid line
                    continue
                # broken line: verify the breaks are covered by coloured data
                nz = np.where(line)[0]
                span = nz[-1] - nz[0] + 1
                if span < L * 0.55:
                    continue
                col_frac, white_frac = _gap_colored_frac(line, cseg, wseg)
                # accept only if gaps are clearly data-covered, not white text gaps
                if col_frac >= 0.30 and white_frac <= 0.55:
                    out.append((rc, run, col_frac))
            return out

        def _legend_swatches_near(yrow):
            """Count chromatic (coloured swatch) pixels just ABOVE or BELOW yrow.
            A legend box has a top rule (swatches below it) and a bottom rule
            (swatches above it); both should be rejected as the x-axis. A real
            x-axis has only achromatic scale numbers below and the plot's data
            (also chromatic) well above -- so we look only in a NARROW band right
            next to the line, where legend swatches sit but plot data does not."""
            ss = img_hsv[:, :, 1]; vv = img_hsv[:, :, 2]
            chrom = (ss > 60) & (vv > 60)
            below = chrom[min(H, yrow + 2):min(H, yrow + 40), :].sum()
            above = chrom[max(0, yrow - 40):max(0, yrow - 2), :].sum()
            return int(max(below, above))

        def _pick(cands, count_axis):
            if not cands:
                return np.array([], dtype=int)
            idxs = np.array([c[0] for c in cands])
            gaps = np.where(np.diff(idxs) > 3)[0]
            groups = np.split(np.arange(len(cands)), gaps + 1)
            dens = col_d if count_axis == 'col' else row_d
            L = H if count_axis == 'col' else W
            def grp_run(g):
                return max(cands[i][1] for i in g)
            qualifying = [g for g in groups if grp_run(g) >= L * 0.25]
            if not qualifying:
                qualifying = [max(groups, key=grp_run)]
            if count_axis == 'col':
                chosen = min(qualifying, key=lambda g: min(cands[i][0] for i in g))
            else:
                # x-axis: reject legend-box rules (chromatic swatches right next
                # to them); among the rest, take the bottom-most.
                non_legend = [g for g in qualifying
                              if _legend_swatches_near(max(cands[i][0] for i in g)) < 150]
                pool = non_legend if non_legend else qualifying
                chosen = max(pool, key=lambda g: max(cands[i][0] for i in g))
            rows = [cands[i][0] for i in chosen]
            peak = max(rows, key=lambda r: dens[r])
            core = [r for r in rows if abs(r - peak) <= 2 and dens[r] >= dens[peak] * 0.5]
            return np.array(sorted(core), dtype=int)

        yc = _pick(_candidates('col'), 'col')
        xr = _pick(_candidates('row'), 'row')
        return xr, yc

    # dark pass, then a relaxed grey pass for any axis still missing
    xr, yc = _detect_with(((v < 110) & (s < 55)).astype(np.uint8))
    if len(xr) == 0 or len(yc) == 0:
        xr2, yc2 = _detect_with(((v < 150) & (s < 55)).astype(np.uint8))
        if len(xr) == 0:
            xr = xr2
        if len(yc) == 0:
            yc = yc2

    # -- Principle: the x-axis and y-axis are drawn in the SAME achromatic colour
    # (black or grey, never a chromatic colour). If both axes were found but their
    # measured colours disagree, the odd one out is not a real axis (e.g. a grey
    # GRIDLINE mistaken for the y-axis while the x-axis is black). Drop it.
    def _axis_colour(idxs, is_col):
        if len(idxs) == 0:
            return None
        rc = int(np.median(idxs))
        if is_col:
            line = ((s[:, rc] < 60) & (v[:, rc] < 200))
            px = img_lab[np.where(line)[0], rc] if line.any() else None
        else:
            line = ((s[rc, :] < 60) & (v[rc, :] < 200))
            px = img_lab[rc, np.where(line)[0]] if line.any() else None
        if px is None or len(px) == 0:
            return None
        return np.median(px, axis=0)

    cx = _axis_colour(yc, True)    # y-axis colour
    cr = _axis_colour(xr, False)   # x-axis colour
    if cx is not None and cr is not None:
        # compare lightness (L) and chroma; axes must be a close achromatic match
        dL = abs(float(cx[0]) - float(cr[0]))
        if dL > 45:                 # clearly different greys -> not the same axis
            # keep the darker one as the real axis colour reference; the lighter
            # one is likely a gridline. The x-axis is almost always present and
            # solid, so trust it and drop a mismatched y-axis (and vice-versa).
            if cx[0] > cr[0] + 45:       # y-axis much lighter than x-axis -> drop y
                yc = np.array([], dtype=int)
            elif cr[0] > cx[0] + 45:     # x-axis much lighter -> drop x
                xr = np.array([], dtype=int)

    # -- Filled-panel fallback (highest priority): if the plot area is a solid
    # colour fill (e.g. a grey panel on white), there is usually no separate axis
    # line. The fill's LEFT edge IS the y-axis and its BOTTOM edge IS the x-axis.
    if len(xr) == 0 or len(yc) == 0:
        fa = _detect_filled_plot_area()
        if fa is not None:
            fx0, fy0, fx1, fy1 = fa
            if len(yc) == 0:
                yc = np.array([fx0], dtype=int)
            if len(xr) == 0:
                xr = np.array([fy1], dtype=int)

    # -- Plot-box fallback: many plots have no separate axis rules but a PLOT BOX
    # (a rectangle, sometimes only an L-shaped left+bottom frame, possibly drawn
    # in a colour that contrasts with the background). Use whichever box edges
    # exist to fill a missing axis: the LEFT edge is the y-axis, the BOTTOM edge
    # is the x-axis. This runs before the numeric-label fallback because a drawn
    # frame is a stronger signal than inferring position from the scale numbers.
    if len(xr) == 0 or len(yc) == 0:
        edges = _detect_plot_box_edges()
        if len(yc) == 0 and 'left' in edges:
            yc = np.array([edges['left']], dtype=int)
        if len(xr) == 0 and 'bottom' in edges:
            xr = np.array([edges['bottom']], dtype=int)

    # -- Numeric-label fallback: many plots draw NO axis line, only tick numbers.
    # The y-scale numbers sit in a band at the far left; the x-scale numbers sit
    # in a band along the bottom. The inner edge of those bands marks the axis:
    #   * y-axis ~ just RIGHT of the left number band
    #   * x-axis ~ just ABOVE the bottom number band
    # We use this only when an axis is still missing, so it never overrides a real
    # detected line.
    if len(yc) == 0 or len(xr) == 0:
        _s = img_hsv[:, :, 1]; _v = img_hsv[:, :, 2]
        dark_txt = (_s < 70) & (_v < 130)        # achromatic dark = numbers/text
        if len(yc) == 0:
            yguess = _yaxis_from_left_numbers(dark_txt, xr)
            if yguess is not None:
                yc = np.array([yguess], dtype=int)
    # -- Scale-number cross-check removed: overriding a detected axis by the
    # number-band position wrongly moved correct axes (e.g. a vertical y-label
    # text read as the number band). The number band is used only as a fallback
    # when an axis is entirely missing (above).
    return xr, yc


def _yaxis_from_left_numbers(dark_txt, xr):
    """Locate the y-axis from the y-scale number band on the left. The numbers
    form a vertical band of small dark glyphs; the y-axis is just to its right.
    Returns an x column, or None."""
    # restrict to ABOVE the x-axis if known (numbers don't go below it)
    region = dark_txt.copy()
    if len(xr) > 0:
        region[int(xr.max()):, :] = False
    # only the left third -- y-scale numbers live there
    region[:, int(W * 0.35):] = False
    col_ink = region.sum(axis=0).astype(float)
    if col_ink.sum() < 20:
        return None
    # A y-scale-number column contains stacked digits, so it has noticeably more
    # ink than a sparse data-marker column. Threshold on per-column ink to keep
    # only number columns, then take the right edge of the left-most dense band.
    thr = max(6, 0.02 * H)
    dense = col_ink >= thr
    xs = np.where(dense)[0]
    if len(xs) == 0:
        return None
    start = int(xs.min())
    band_right = start
    gap = 0
    for x in range(start, min(W, int(W * 0.35))):
        if dense[x]:
            band_right = x; gap = 0
        else:
            gap += 1
            if gap > max(10, int(0.03 * W)):
                break
    guess = min(W - 3, band_right + 5)
    return guess


def _xaxis_from_bottom_numbers(dark_txt, yc):
    """Locate the x-axis from the x-scale number band along the bottom. The
    x-axis is just above that band. Returns a row, or None."""
    region = dark_txt.copy()
    if len(yc) > 0:
        region[:, :int(yc.min())] = False        # numbers are right of the y-axis
    # only the bottom third
    region[:int(H * 0.6), :] = False
    row_ink = region.sum(axis=1).astype(float)
    if row_ink.sum() < 20:
        return None
    thr = max(6, 0.02 * W)
    dense = row_ink >= thr
    ys = np.where(dense)[0]
    if len(ys) == 0:
        return None
    end = int(ys.max())
    band_top = end
    gap = 0
    for y in range(end, max(0, int(H * 0.6)), -1):
        if dense[y]:
            band_top = y; gap = 0
        else:
            gap += 1
            if gap > max(10, int(0.03 * H)):
                break
    guess = max(2, band_top - 5)
    return guess


def _detect_plot_box_edges():
    """Find the edges of a rectangular plotting-area box. The box may be drawn in
    an achromatic colour OR a colour that contrasts with the background (e.g. a
    grey panel border on white, or a coloured frame). Unlike a strict box test,
    this returns whatever strong straight edges exist -- left/right verticals and
    top/bottom horizontals -- so a PARTIAL frame (just left+bottom, an L-shape) is
    still usable to place the y- and x-axes.

    Returns dict with optional keys 'left','right','top','bottom' (pixel index),
    or {} if nothing box-like is found.
    """
    # Build an "edge" mask: pixels that differ from the background colour. This
    # naturally captures a contrasting-colour frame as well as a black/grey one.
    border = np.concatenate([img_rgb[0, :], img_rgb[-1, :],
                             img_rgb[:, 0], img_rgb[:, -1]]).reshape(-1, 3)
    bg = np.median(border, axis=0)
    diff = np.linalg.norm(img_rgb.astype(int) - bg.astype(int), axis=2)
    # A frame line contrasts with the background; 25 catches light-grey borders.
    edge = diff > 25

    def _run(px):
        m = run = 0
        for p in px:
            if p:
                run += 1; m = max(m, run)
            else:
                run = 0
        return m

    mgn = 2
    H_, W_ = edge.shape
    # candidate vertical / horizontal frame lines: a long continuous edge run
    vcols = [(x, _run(edge[:, x])) for x in range(mgn, W_ - mgn)]
    hrows = [(y, _run(edge[y, :])) for y in range(mgn, H_ - mgn)]
    vstrong = [x for x, r in vcols if r > H_ * 0.55]
    hstrong = [y for y, r in hrows if r > W_ * 0.55]

    def _group(idx):
        if not idx:
            return []
        idx = sorted(idx); out = []; cur = [idx[0]]
        for i in idx[1:]:
            if i - cur[-1] <= 4:
                cur.append(i)
            else:
                out.append(int(np.mean(cur))); cur = [i]
        out.append(int(np.mean(cur))); return out

    vg = _group(vstrong); hg = _group(hstrong)
    res = {}
    # left/right verticals: a frame's left edge sits in the left half, right edge
    # in the right half. Take the inner-most plausible ones.
    left_cands = [x for x in vg if x < W_ * 0.5]
    right_cands = [x for x in vg if x >= W_ * 0.5]
    top_cands = [y for y in hg if y < H_ * 0.5]
    bottom_cands = [y for y in hg if y >= H_ * 0.5]
    if left_cands:
        res['left'] = min(left_cands)       # outer-most left edge
    if right_cands:
        res['right'] = max(right_cands)
    if top_cands:
        res['top'] = min(top_cands)
    if bottom_cands:
        res['bottom'] = max(bottom_cands)
    return res


def _detect_plot_box():
    """Strict box (all four edges) -> (x0,y0,x1,y1), else None. Kept for callers
    that want a full rectangle; partial frames go through _detect_plot_box_edges."""
    e = _detect_plot_box_edges()
    if all(k in e for k in ('left', 'right', 'top', 'bottom')):
        x0, y0, x1, y1 = e['left'], e['top'], e['right'], e['bottom']
        if (x1 - x0) >= 0.3 * W and (y1 - y0) >= 0.3 * H:
            return (x0, y0, x1, y1)
    return None

AXIS_ROWS, AXIS_COLS = _detect_axes()
print(f"Axis rows: {AXIS_ROWS.tolist()},  Axis cols: {AXIS_COLS.tolist()}")
try:
    _BOX_EDGES = _detect_plot_box_edges()
    print(f"PLOT_BOX: {_BOX_EDGES}")
except Exception:
    _BOX_EDGES = {}

FILLED_AREA = _detect_filled_plot_area()


def _data_bbox_area(pad_frac=0.01):
    """P1 (v38) fallback: bounding box of plot CONTENT (chromatic curve/marker
    pixels). Used when the axis-derived rectangle is missing or has collapsed to
    a sliver. Data markers are chromatic, so their bounding box can never cut real
    data. A 0.5%-tail trim keeps stray legend swatches / speckle from stretching
    it. Returns (x0, y0, x1, y1) or None."""
    s = img_hsv[:, :, 1]; v = img_hsv[:, :, 2]
    chrom = (s > 45) & (v > 40) & (v < 250)
    ys, xs = np.where(chrom)
    if len(xs) < 30:
        return None
    x0 = int(np.percentile(xs, 0.5)); x1 = int(np.percentile(xs, 99.5))
    y0 = int(np.percentile(ys, 0.5)); y1 = int(np.percentile(ys, 99.5))
    padx = int(pad_frac * W) + 3; pady = int(pad_frac * H) + 3
    return (max(0, x0 - padx), max(0, y0 - pady),
            min(W - 1, x1 + padx), min(H - 1, y1 + pady))


def _compute_plot_area():
    """Define the plotting rectangle.

    Priority:
      1. A COLOUR-FILLED plot panel (grey on white) -- the fill marks it exactly.
      2. Otherwise derive from detected axes. v38/P1 change: pick the axis lines
         by POSITION, not by a blind .max(). The y-axis is the vertical line in
         the LEFT part of the image; the x-axis the horizontal line in the BOTTOM
         part. This stops a right-hand frame or a stray top rule from being
         mistaken for an axis (which used to collapse the box to a sliver). If the
         result is still degenerate, fall back to the chromatic data bounding box
         so the plot area can never collapse.
    Returns (x0, y0, x1, y1) inclusive, or None.
    """
    if FILLED_AREA is not None:
        return FILLED_AREA
    db = _data_bbox_area()
    if len(AXIS_ROWS) == 0 and len(AXIS_COLS) == 0:
        if db is not None:
            print("  [plot-area P1: no axes -> data-bbox fallback]")
        return db                       # was None; now data-bbox fallback
    pad = max(4, int(0.01 * max(W, H)))
    x0, y0, x1, y1 = 0, 0, W - 1, H - 1
    # --- y-axis column: prefer the LEFT region; never pick a right-hand frame ---
    if len(AXIS_COLS) > 0:
        cols = [int(c) for c in AXIS_COLS]
        left = [c for c in cols if c <= 0.45 * W]
        ycol = max(left) if left else min(cols)   # innermost-left, else leftmost
        x0 = max(0, ycol - pad)
    # --- x-axis row: prefer the BOTTOM region; never pick a top rule ---
    if len(AXIS_ROWS) > 0:
        rows = [int(r) for r in AXIS_ROWS]
        bot = [r for r in rows if r >= 0.55 * H]
        xrow = max(bot) if bot else max(rows)      # bottom band, else bottom-most
        y1 = min(H - 1, xrow + pad)
    # --- P1 sanity floor: reject collapsed rectangles via data bbox ---
    if db is not None and ((x1 - x0) < 0.25 * W or (y1 - y0) < 0.25 * H):
        dx0, dy0, dx1, dy1 = db
        left_ok = (len(AXIS_COLS) > 0 and x0 <= 0.45 * W)
        bot_ok  = (len(AXIS_ROWS) > 0 and y1 >= 0.55 * H)
        if not left_ok:
            x0 = dx0
        if not bot_ok:
            y1 = dy1
        if (x1 - x0) < 0.25 * W:      # still thin horizontally -> take data span
            x0, x1 = dx0, dx1
        if (y1 - y0) < 0.25 * H:      # still thin vertically -> take data bottom
            y1 = dy1
        print(f"  [plot-area P1: degenerate box corrected via data-bbox {db}]")
    return (int(x0), int(y0), int(x1), int(y1))

PLOT_AREA = _compute_plot_area()
if USER_PLOT_AREA is not None:
    PLOT_AREA = USER_PLOT_AREA
    # Derive axes from the user's rectangle: left edge = y-axis column,
    # bottom edge = x-axis row. Calibration reads numbers relative to these.
    _ux0, _uy0, _ux1, _uy1 = USER_PLOT_AREA
    AXIS_COLS = np.array([int(_ux0)])
    AXIS_ROWS = np.array([int(_uy1)])
    print(f"Plot area (user-provided): {PLOT_AREA}")
    print(f"Axis rows: {AXIS_ROWS},  Axis cols: {AXIS_COLS}  (from user plot area)")
else:
    print(f"Plot area: {PLOT_AREA}")


# -- Axis tick calibration ----------------------------------------------------
# Reads the numbers printed along the x- and y-axes (OCR), fits a pixel->data
# mapping (linear or log, chosen by fit), and exposes x2v/y2v so detected point
# pixels can be converted to real data coordinates. Merged inline (was a separate
# module) so the pipeline is a single file.

def _cal_ocr_group(gray, x0, y0, x1, y1):
    crop = gray[max(0, y0-4):y1+4, max(0, x0-4):x1+4]
    if crop.size == 0:
        return None
    crop = cv2.resize(crop, None, fx=10, fy=10, interpolation=cv2.INTER_CUBIC)
    _, crop = cv2.threshold(crop, 150, 255, cv2.THRESH_BINARY)
    crop = cv2.copyMakeBorder(crop, 20, 20, 20, 20, cv2.BORDER_CONSTANT, value=255)
    for psm in [7, 8, 10, 13]:
        t = pytesseract.image_to_string(
            crop, config=f'--psm {psm} -c tessedit_char_whitelist=0123456789.').strip()
        t = t.replace(' ', '')
        if re.fullmatch(r'\d+\.?\d*', t) and t not in ('.', ''):
            try:
                return float(t)
            except Exception:
                pass
    return None


def _cal_read_band(img_, x0b, y0b, x1b, y1b, orient):
    """OCR the scale numbers in a band beside an axis. Rejects the axis TITLE
    (a wide word run). Returns [(center_pixel_along_axis, value_or_None)]."""
    band = img_[y0b:y1b, x0b:x1b]
    if band.size == 0:
        return []
    gray = cv2.cvtColor(band, cv2.COLOR_BGR2GRAY)
    _, bw = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY_INV)
    n, lbl, st, cen = cv2.connectedComponentsWithStats(bw, 8)
    comps = [(st[i, 0], st[i, 1], st[i, 2], st[i, 3]) for i in range(1, n)
             if st[i, cv2.CC_STAT_AREA] >= 4 and 4 <= st[i, cv2.CC_STAT_HEIGHT] <= 40]
    if not comps:
        return []
    key = 0 if orient == 'x' else 1
    size = 2 if orient == 'x' else 3
    comps.sort(key=lambda c: c[key])
    gap = 12 if orient == 'x' else 10
    groups = []
    for c in comps:
        if groups:
            last = groups[-1][-1]
            if c[key] - (last[key] + last[size]) < gap:
                groups[-1].append(c); continue
        groups.append([c])
    out = []
    for g in groups:
        gx0 = min(c[0] for c in g); gx1 = max(c[0]+c[2] for c in g)
        gy0 = min(c[1] for c in g); gy1 = max(c[1]+c[3] for c in g)
        if orient == 'x' and (gx1-gx0) > 0.12*(x1b-x0b) and (gx1-gx0) > 60:
            continue                       # too wide -> part of the axis title
        val = _cal_ocr_group(gray, gx0, gy0, gx1, gy1)
        center = (gx0+gx1)/2 + x0b if orient == 'x' else (gy0+gy1)/2 + y0b
        out.append((center, val))
    return out


def _cal_ransac_fit(pairs):
    """Fit pixel->value (linear or log) by RANSAC to reject OCR outliers. A real
    axis is monotonic and its evenly-spaced ticks form an arithmetic (linear) or
    geometric (log) progression, so flat fits and equal-value anchor pairs are
    rejected. Returns (kind, p2v, v2p, inliers) or None."""
    P = [(p, v) for p, v in pairs if v is not None]
    if len(P) < 2:
        return None
    best = []; bestmode = None; bestparam = None
    for mode in ('linear', 'log'):
        for i in range(len(P)):
            for j in range(i+1, len(P)):
                (p1, v1), (p2, v2) = P[i], P[j]
                if p1 == p2 or v1 == v2:
                    continue
                if mode == 'log':
                    if v1 <= 0 or v2 <= 0:
                        continue
                    a = (np.log10(v2)-np.log10(v1))/(p2-p1); b = np.log10(v1)-a*p1
                    if abs(a) < 1e-9:
                        continue
                    def pred(p, a=a, b=b): return 10**(a*p+b)
                    def tol(v): return max(abs(v)*0.04, 1e-9)
                else:
                    a = (v2-v1)/(p2-p1); b = v1-a*p1
                    if abs(a) < 1e-9:
                        continue
                    def pred(p, a=a, b=b): return a*p+b
                    def tol(v): return max(abs(v)*0.02, 0.4)
                inl = [(p, v) for (p, v) in P if abs(pred(p)-v) <= tol(v)]
                if len(inl) >= 2 and len(set(v for _, v in inl)) < 2:
                    continue
                if len(inl) > len(best):
                    best = inl; bestmode = mode; bestparam = (a, b)
    if len(best) < 2:
        return None
    a, b = bestparam
    if bestmode == 'log':
        return ('log', lambda p: 10**(a*p+b), lambda val: (np.log10(val)-b)/a, best)
    return ('linear', lambda p: a*p+b, lambda val: (val-b)/a, best)


def _cal_build_p2v(inliers, kind):
    if not inliers or len(inliers) < 2:
        return None
    p = np.array([a[0] for a in inliers], float)
    v = np.array([a[1] for a in inliers], float)
    A = np.vstack([p, np.ones_like(p)]).T
    if kind == 'log':
        a, b = np.linalg.lstsq(A, np.log10(v), rcond=None)[0]
        return lambda px: float(10**(a*px+b))
    a, b = np.linalg.lstsq(A, v, rcond=None)[0]
    return lambda px: float(a*px+b)


def calibrate_from_axes(img_, axr, axc, pa):
    """Given the image and detected axis rows/cols + plot area, read the axis
    numbers and build pixel->data mappings. Returns dict with coords, x2v, y2v,
    and the OCR inliers used, or None if axes/area are missing."""
    H_, W_ = img_.shape[:2]
    if not axr or not axc or not pa:
        return None
    xax = max(axr); yax = min(axc)
    hsv = cv2.cvtColor(img_, cv2.COLOR_BGR2HSV)
    dark = ((hsv[:, :, 2] < 120) & (hsv[:, :, 1] < 70))

    def _num_band_x():
        prof = [int(dark[xax+dy, :].sum()) for dy in range(1, 45) if xax+dy < H_]
        seen_ink = False; in_gap = False; start = None
        for k, c in enumerate(prof):
            dy = k+1
            if start is None:
                if not seen_ink and c > 2:
                    seen_ink = True
                elif seen_ink and c <= 2:
                    in_gap = True
                elif in_gap and c > 2:
                    start = dy
            else:
                if c <= 2:
                    return xax+start, xax+dy
        if start is not None:
            return xax+start, xax+start+max(18, int(0.05*H_))
        d = max(11, int(0.028*H_))
        return xax+d, xax+d+max(18, int(0.05*H_))

    xstart, xend = _num_band_x()
    yoff = max(6, int(0.012*W_)); ydepth = max(46, int(0.11*W_))
    xraw = _cal_read_band(img_, yax, xstart, W_, min(H_, xend), 'x')
    yraw = _cal_read_band(img_, max(0, yax-yoff-ydepth), 0, yax-yoff, xax, 'y')
    xf = _cal_ransac_fit(xraw); yf = _cal_ransac_fit(yraw)

    def _numeric_frac(raw):
        if not raw:
            return 0.0
        return sum(1 for _, v in raw if v is not None)/len(raw)
    if _numeric_frac(xraw) < 0.34:
        xf = None
    if _numeric_frac(yraw) < 0.34:
        yf = None
    coords = None; x2v = y2v = None
    if xf and yf:
        coords = {'x_min': float(xf[1](pa[0])), 'x_max': float(xf[1](pa[2])),
                  'y_min': float(yf[1](pa[3])), 'y_max': float(yf[1](pa[1])),
                  'x_kind': xf[0], 'y_kind': yf[0]}
        x2v = _cal_build_p2v([(p, v) for p, v in xf[3]], xf[0])
        y2v = _cal_build_p2v([(p, v) for p, v in yf[3]], yf[0])
    return {'pa': pa, 'xax': xax, 'yax': yax, 'coords': coords,
            'x2v': x2v, 'y2v': y2v,
            'x_inliers': [(round(p), v) for p, v in xf[3]] if xf else [],
            'y_inliers': [(round(p), v) for p, v in yf[3]] if yf else []}


def _plot_area_mask():
    """Boolean mask, True inside the plotting rectangle (all True if unknown)."""
    m = np.zeros((H, W), dtype=bool)
    if PLOT_AREA is None:
        m[:] = True
        return m
    x0, y0, x1, y1 = PLOT_AREA
    m[y0:y1+1, x0:x1+1] = True
    return m

PLOT_MASK = _plot_area_mask()

def _detect_noise_color():
    """Measure the dominant colour of the structural noise (x/y axes, tick
    labels, title text, the LLOQ reference line). These are drawn in one ink
    colour -- almost always near-black. We sample mainly the detected AXIS LINES
    (the most reliable pure-noise pixels) and dark text. Text can appear INSIDE
    the plotting area too (titles, annotations), so we do not require text to be
    outside; we only prefer outside text when axis samples already anchor the
    colour, to avoid mixing in a genuine black curve.
    The returned Lab colour lets the cleaning stage strip this noise from the
    matching COLOUR MASK (usually black) and leave other colour masks untouched.
    """
    ys_ax, xs_ax = [], []
    # sample along detected axis rows/cols
    for ay in (AXIS_ROWS if len(AXIS_ROWS) else []):
        for x in range(0, W, 4):
            ys_ax.append(ay); xs_ax.append(x)
    for ax in (AXIS_COLS if len(AXIS_COLS) else []):
        for y in range(0, H, 4):
            ys_ax.append(y); xs_ax.append(ax)
    samples = []
    have_axis = False
    if ys_ax:
        ys_ax = np.clip(np.array(ys_ax), 0, H - 1)
        xs_ax = np.clip(np.array(xs_ax), 0, W - 1)
        v = img_hsv[ys_ax, xs_ax, 2]; s = img_hsv[ys_ax, xs_ax, 1]
        ink = (v < 160) & (s < 80)               # dark, low-sat = the rule itself
        if ink.sum() > 20:
            samples.append(img_lab[ys_ax[ink], xs_ax[ink]])
            have_axis = True
    # sample dark text. If axis pixels already anchor the noise colour, restrict
    # text sampling to OUTSIDE the plot (so a black curve inside isn't mixed in);
    # if we have no axis anchor, sample dark text anywhere -- the text may sit
    # entirely inside the plot area.
    fg = ~((img_hsv[:, :, 2] > 240) & (img_hsv[:, :, 1] < 15))
    dark = fg & (img_hsv[:, :, 2] < 90) & (img_hsv[:, :, 1] < 60)
    if have_axis and PLOT_AREA is not None:
        outside = np.ones((H, W), bool)
        x0, y0, x1, y1 = PLOT_AREA
        outside[y0:y1+1, x0:x1+1] = False
        dark = dark & outside
    yd, xd = np.where(dark)
    if len(yd) > 50:
        idx = np.random.choice(len(yd), min(2000, len(yd)), replace=False)
        samples.append(img_lab[yd[idx], xd[idx]])
    if not samples:
        # default: assume black
        return cv2.cvtColor(np.array([[(20, 20, 20)]], np.uint8),
                            cv2.COLOR_RGB2Lab)[0, 0].astype(np.float32)
    alls = np.vstack(samples)
    return np.median(alls, axis=0).astype(np.float32)

NOISE_LAB = _detect_noise_color()
print(f"Noise (axis/text/LLOQ) colour Lab: {NOISE_LAB.astype(int).tolist()}")
NOISE_MATCH_DIST = 45.0   # a colour mask within this Lab distance of the noise
                          # colour is the one we strip axis/text/LLOQ from

def _detect_ticks():
    if len(AXIS_ROWS) == 0: return False, []
    v = img_hsv[:, :, 2]; s = img_hsv[:, :, 1]
    black = ((v < 80) & (s < 40)).astype(np.uint8)
    axis_y = int(AXIS_ROWS.mean())
    tz0 = min(H-1, axis_y+1); tz1 = min(H, axis_y+DENSITY_TICK_MAX_LEN+1)
    tz0b = max(0, axis_y-DENSITY_TICK_MAX_LEN); tz1b = max(0, axis_y)
    col_sum = (black[tz0:tz1, :].sum(axis=0).astype(float) +
               black[tz0b:tz1b, :].sum(axis=0).astype(float))
    if col_sum.max() == 0: return False, []
    if len(AXIS_COLS) > 0:
        ax_col = int(AXIS_COLS.mean())
        col_sum[max(0, ax_col-5):min(W, ax_col+6)] = 0
    min_tick_dist = max(5, W // 60)
    tick_peaks, _ = find_peaks(col_sum, height=1, distance=min_tick_dist, prominence=1)
    has_ticks = len(tick_peaks) >= DENSITY_MIN_TICKS
    print(f"Tick detection: {len(tick_peaks)} ticks  has_ticks={has_ticks}")
    return has_ticks, sorted(tick_peaks.tolist())

HAS_TICKS, TICK_XS = _detect_ticks()

def _remove_axes_from_mask(mask):
    """Remove axis rows/cols and everything below the x-axis (legend/label area)."""
    out = mask.copy()
    for yr in AXIS_ROWS:
        out[max(0, yr-AXIS_PAD):min(H, yr+AXIS_PAD+1), :] = 0
    for xc in AXIS_COLS:
        out[:, max(0, xc-AXIS_PAD):min(W, xc+AXIS_PAD+1)] = 0
    if len(AXIS_ROWS) > 0:
        # Only blank pixels strictly below the axis line (with a small pad).
        # Do NOT remove the entire region above; only strip the true x-axis row
        # to avoid deleting dark curves that happen to be near the baseline.
        ax_y = int(AXIS_ROWS.max()) + AXIS_PAD + 1
        out[ax_y:, :] = 0
    return out

# -- Step 1: Legend detection ---------------------------------------------------
def _ocr_understandable(txt, conf):
    """A legend label must be READABLE text, not a stray symbol or marker glyph.
    Require decent OCR confidence and at least two alphanumeric characters, which
    rejects swatch artefacts like '-@-' or '~(R)' while keeping 'Placebo',
    'mg/kg', 'Q4W', 'Cohort', etc."""
    t = txt.strip()
    if conf < 50 or len(t) < 2:
        return False
    if sum(c.isalpha() for c in t) + sum(c.isdigit() for c in t) < 2:
        return False
    return True


def _tighten_legend_with_ocr(rough_box, swatch_points=None):
    """Tighten a legend box to the TRUE swatch+text extent.

    The box is the union of:
      (a) the OCR boxes of UNDERSTANDABLE words (the labels), and
      (b) the actual swatch ink measured inside the region (coloured marker
          pixels) -- we MEASURE the swatches rather than trusting the detector's
          box edge, which is often far to the left of the real swatches and
          leaves a big empty margin.
    A small fixed padding is added. If no readable text is found, returns None
    (not a legend). This makes the box neither too large (no detector slack,
    no empty margin) nor too small (every label word + its swatch is inside).
    """
    if not _HAS_OCR or rough_box is None:
        return rough_box
    x0, y0, x1, y1 = rough_box
    pad = 35
    rx0, ry0 = max(0, x0 - pad), max(0, y0 - pad)
    rx1, ry1 = min(W, x1 + pad), min(H, y1 + pad)
    if rx1 <= rx0 or ry1 <= ry0:
        return rough_box
    crop = img[ry0:ry1, rx0:rx1]
    try:
        up = cv2.resize(crop, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
        g = cv2.cvtColor(up, cv2.COLOR_BGR2GRAY)
        d = pytesseract.image_to_data(g, output_type=_TESS_OUT.DICT)
    except Exception:
        return rough_box

    # (a) readable label words
    words = []
    for i, txt in enumerate(d['text']):
        try:
            conf = int(d['conf'][i])
        except (ValueError, TypeError):
            conf = -1
        if _ocr_understandable(txt, conf):
            wx = rx0 + d['left'][i] // 2
            wy = ry0 + d['top'][i] // 2
            ww = d['width'][i] // 2
            wh = d['height'][i] // 2
            words.append((wx, wy, wx + ww, wy + wh))
    if not words:
        return None
    tx0 = min(w[0] for w in words); ty0 = min(w[1] for w in words)
    tx1 = max(w[2] for w in words); ty1 = max(w[3] for w in words)

    # (b) measure swatch ink. Look in the band spanned by the text rows but to
    # the LEFT of (and within) the text, for coloured OR dark compact marks.
    reg_hsv = img_hsv[ry0:ry1, rx0:rx1]
    sreg = reg_hsv[:, :, 1]; vreg = reg_hsv[:, :, 2]
    swatch_ink = ((sreg > 40) & (vreg > 40) & (vreg < 250)) | \
                 ((sreg <= 40) & (vreg >= 20) & (vreg < 160))   # coloured or dark marker
    sy, sx = np.where(swatch_ink)
    sw_x0 = sw_y0 = sw_x1 = sw_y1 = None
    if len(sx) > 0:
        ax = rx0 + sx; ay = ry0 + sy
        # restrict to swatches plausibly belonging to the label rows
        band = (ay >= ty0 - 12) & (ay <= ty1 + 12)
        if band.sum() > 0:
            ax, ay = ax[band], ay[band]
            sw_x0, sw_y0 = int(ax.min()), int(ay.min())
            sw_x1, sw_y1 = int(ax.max()), int(ay.max())

    # union of text and measured swatches, then a tight 4px pad
    bx0, by0 = tx0, ty0
    bx1, by1 = tx1, ty1
    if sw_x0 is not None:
        bx0 = min(bx0, sw_x0); by0 = min(by0, sw_y0)
        bx1 = max(bx1, sw_x1); by1 = max(by1, sw_y1)
    P = 4
    return (max(0, int(bx0) - P), max(0, int(by0) - P),
            min(W - 1, int(bx1) + P), min(H - 1, int(by1) + P))


def _markers_from_user_box(box):
    """Find legend swatch MARKERS directly inside a user-drawn box, for legends
    whose thin 'line-through' or hollow markers slip past the blob/chroma gates.

    Strategy: markers sit in a narrow COLUMN at the left of each row, before the
    label text. Detect compact non-white components, cluster them into rows, and
    for each row take the marker nearest the left (the swatch, not stray text or
    the connecting line). Achromatic markers get their DARK core colour so an open
    black circle reads black, not mid-grey. Returns [(cx,cy,(r,g,b)), ...] sorted
    top-to-bottom, or None."""
    x0, y0, x1, y1 = [int(v) for v in box]
    x0 = max(0, x0); y0 = max(0, y0); x1 = min(W, x1); y1 = min(H, y1)
    if x1 - x0 < 6 or y1 - y0 < 6:
        return None
    sub = img_rgb[y0:y1, x0:x1]
    gray = cv2.cvtColor(sub, cv2.COLOR_RGB2GRAY)
    ink = (gray < 205).astype(np.uint8)
    n, lbl, st, cen = cv2.connectedComponentsWithStats(ink, 8)
    comps = []
    for i in range(1, n):
        a = st[i, cv2.CC_STAT_AREA]
        w = st[i, cv2.CC_STAT_WIDTH]; h = st[i, cv2.CC_STAT_HEIGHT]
        if a < 6 or w > 90 or h > 26:            # skip long label words / big fills
            continue
        cx, cy = cen[i]
        pix = sub[lbl == i]
        med = np.median(pix, axis=0)
        chroma = int(med.max()) - int(med.min())
        if chroma <= 24:                          # achromatic -> use dark core
            Ls = pix.sum(axis=1)
            core = pix[Ls <= np.percentile(Ls, 30)]
            rgb = tuple(int(z) for z in (np.mean(core, axis=0) if len(core) >= 3 else med))
        else:
            rgb = tuple(int(z) for z in med)
        comps.append((float(cx), float(cy), rgb, chroma))
    if len(comps) < 3:
        return None
    # cluster into rows by y
    comps.sort(key=lambda c: c[1])
    rows = [[comps[0]]]
    for c in comps[1:]:
        (rows[-1].append(c) if c[1] - rows[-1][-1][1] <= 7 else rows.append([c]))
    out = []
    for r in rows:
        # the swatch marker is the LEFT-most compact component in the row; prefer a
        # chromatic or dark one over a faint stray. Merge same-row fragments.
        r.sort(key=lambda c: c[0])
        cand = r[0]
        # if the left-most is a thin connecting-line fragment, pick the nearest
        # sizable one within ~18px to its right
        cx = cand[0]
        rgb = cand[2]
        out.append((cand[0] + x0, cand[1] + y0, rgb))
    # keep only rows whose swatch column is near the common left column (drop text)
    lefts = sorted(o[0] for o in out)
    col_x = lefts[len(lefts) // 2]
    out = [o for o in out if abs(o[0] - col_x) <= 22]
    return out if len(out) >= 3 else None


def _pcm_detect_legend_unified(restrict_box=None):
    """Orientation-agnostic legend detector. Instead of assuming swatches are in
    a vertical column (or a separate horizontal path), find the legend by what a
    legend entry actually IS: a small, compact COLOUR swatch with a TEXT LABEL
    beside it. Then look for a set of such swatches that are collinear (share an
    x for a vertical legend, or a y for a horizontal one), evenly spaced, and of
    DIFFERENT colours. This single routine handles both layouts.

    If restrict_box=(x0,y0,x1,y1) is given (e.g. a user-drawn legend box), the
    search is confined to that rectangle so the detector cannot wander off to a
    different legend elsewhere in a multi-panel figure.

    Returns ('colors', [(x,y,(r,g,b)), ...], box) or None.
    """
    s = img_hsv[:, :, 1]; v = img_hsv[:, :, 2]
    chrom_m = ((s > 40) & (v > 40) & (v < 250)).astype(np.uint8)
    # P3 (v38): also admit ACHROMATIC (grey/black) swatches -- a grey 'Placebo'
    # key or a black treatment line, which the chroma-only mask dropped and which
    # caused whole legends (Asfotase 'Treatment', Placebo rows) to be missed.
    # Achromatic candidates are gated by SOLIDITY so text glyphs and thin axis
    # lines do not flood the blob pool (a filled marker is solid; a glyph is not).
    achrom_m = ((s <= 40) & (v < 220) & (v > 25)).astype(np.uint8)
    if restrict_box is not None:
        _rx0, _ry0, _rx1, _ry1 = [int(v) for v in restrict_box]
        _keep = np.zeros_like(chrom_m)
        _keep[max(0, _ry0):_ry1 + 1, max(0, _rx0):_rx1 + 1] = 1
        chrom_m = chrom_m & _keep
        achrom_m = achrom_m & _keep

    def _blobs_from(mask, require_solid):
        mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, np.ones((3, 3), np.uint8))
        n, lbl, st, cen = cv2.connectedComponentsWithStats(mask, 8)
        out = []
        for i in range(1, n):
            a = st[i, cv2.CC_STAT_AREA]
            w = st[i, cv2.CC_STAT_WIDTH]; h = st[i, cv2.CC_STAT_HEIGHT]
            # a swatch is small & compact (not a long curve / big fill)
            if not (15 <= a <= 2500 and w <= 120 and h <= 60):
                continue
            if require_solid:
                solid = a / float(max(1, w * h))
                bar = (8 <= w <= 60 and 2 <= h <= 16)   # a legend line / dash
                if solid < 0.55 and not bar:
                    continue                             # reject text glyphs
            cx, cy = cen[i]
            # legend entries have a TEXT LABEL adjacent (usually to the right).
            tx0, tx1 = int(cx + 10), int(min(W, cx + 160))
            ty0, ty1 = int(max(0, cy - 12)), int(min(H, cy + 12))
            if tx1 <= tx0:
                continue
            txt = int(((v[ty0:ty1, tx0:tx1] < 120) & (s[ty0:ty1, tx0:tx1] < 60)).sum())
            if txt < 100:
                continue
            col = tuple(int(z) for z in np.median(img_rgb[lbl == i], axis=0))
            if require_solid is False and (max(col) - min(col)) <= 24:
                continue          # near-neutral -> not a chromatic swatch (achromatic
                                  # entries are handled by the achromatic classifier)
            out.append((float(cx), float(cy), col))
        return out

    blobs = _blobs_from(chrom_m, require_solid=False)        # chromatic swatches only
    if len(blobs) < LEGEND_MIN_ENTRIES:
        return None

    def _diverse(group):
        u = []
        for _, _, c in group:
            if all(abs(c[0]-w[0]) + abs(c[1]-w[1]) + abs(c[2]-w[2]) > 50 for w in u):
                u.append(c)
        return len(u)

    best = None; best_score = -1e9; best_axis = 0
    for axis in (0, 1):              # 0: align x (vertical legend); 1: align y (horizontal)
        other = 1 - axis
        for anchor in blobs:
            grp = [b for b in blobs if abs(b[axis] - anchor[axis]) <= 20]
            if len(grp) < LEGEND_MIN_ENTRIES:
                continue
            grp = sorted(grp, key=lambda b: b[other])
            # dedup entries at nearly the same position along the run
            dedup = [grp[0]]; last = grp[0][other]
            for b in grp[1:]:
                if b[other] - last > 10:
                    dedup.append(b); last = b[other]
            grp = dedup
            if len(grp) < LEGEND_MIN_ENTRIES:
                continue
            pos = np.array([b[other] for b in grp])
            gaps = np.diff(pos)
            if gaps.mean() == 0:
                continue
            cv_sp = gaps.std() / gaps.mean()
            # P3 (v38): a real legend is COMPACT. If the entries span most of the
            # plot along their run, they are almost certainly data markers that
            # happen to line up -- reject so the box can't swallow the data area.
            span = float(pos.max() - pos.min())
            dim = H if axis == 0 else W
            if span > 0.75 * dim:
                continue
            ndiv = _diverse(grp)
            if ndiv < LEGEND_MIN_ENTRIES:        # must be colour-diverse
                continue
            score = ndiv * 5.0 - cv_sp * 8.0 + len(grp)
            if score > best_score:
                best_score, best, best_axis = score, grp, axis
    if best is None:
        return None
    single_best = best                     # validated single-axis result (floor)

    # ---- 2-D GRID (primary, general for N-column / N-row legends) ----
    def _cluster1d(vals, tol):
        vals = sorted(vals)
        groups = [[vals[0]]]
        for v in vals[1:]:
            (groups[-1].append(v) if v - groups[-1][-1] <= tol else groups.append([v]))
        return [float(np.mean(g)) for g in groups]

    def _nearest(v, centers):
        return int(min(range(len(centers)), key=lambda i: abs(v - centers[i])))

    def _regular(vals):
        """True if the sorted 1-D positions are EVENLY spaced (a real legend row/
        column rhythm). Irregular spacing is NOT allowed."""
        vals = sorted(vals)
        if len(vals) <= 2:
            return True
        g = np.diff(vals)
        return (g.mean() > 0) and (g.std() / g.mean() <= 0.30)

    grid_cells = None
    if len(blobs) >= LEGEND_MIN_ENTRIES:
        col_x = _cluster1d([b[0] for b in blobs], 18)
        row_y = _cluster1d([b[1] for b in blobs], 7)      # tight: keep close rows apart
        # assign each swatch to a (col,row) CELL; keep the one nearest the cell
        # centre so a split/duplicate blob doesn't create a second entry.
        cell = {}
        for b in blobs:
            ci = _nearest(b[0], col_x); ri = _nearest(b[1], row_y)
            dist = abs(b[0] - col_x[ci]) + abs(b[1] - row_y[ri])
            if (ci, ri) not in cell or dist < cell[(ci, ri)][1]:
                cell[(ci, ri)] = (b, dist)
        occ = {k: v[0] for k, v in cell.items()}
        col_cnt = {}; row_cnt = {}
        for (ci, ri) in occ:
            col_cnt[ci] = col_cnt.get(ci, 0) + 1
            row_cnt[ri] = row_cnt.get(ri, 0) + 1
        single_line = (len(col_cnt) == 1 or len(row_cnt) == 1)
        _maxcol = max(col_cnt.values())
        if _maxcol >= len(occ) - 1 and not single_line:
            # one column holds (almost) everything -> a single-column legend; a lone
            # swatch in another column is a stray. Keep only populated columns.
            cand = [b for (ci, ri), b in occ.items() if col_cnt[ci] >= 2]
        else:
            # genuine multi-column / single-line legend: ragged corners allowed
            cand = [b for (ci, ri), b in occ.items()
                    if single_line or col_cnt[ci] >= 2 or row_cnt[ri] >= 2]
        # OVERLAP guard: if two DIFFERENT-colour swatches land in the same cell the
        # grid is ambiguous/irregular -> reject it (a same-colour split blob is OK).
        raw_cells = {}
        for b in blobs:
            ci = _nearest(b[0], col_x); ri = _nearest(b[1], row_y)
            raw_cells.setdefault((ci, ri), []).append(b)

        def _cell_overlap(v):
            for i in range(len(v)):
                for j in range(i + 1, len(v)):
                    c1, c2 = v[i][2], v[j][2]
                    if abs(c1[0]-c2[0]) + abs(c1[1]-c2[1]) + abs(c1[2]-c2[2]) > 60:
                        return True
            return False
        overlapped = any(_cell_overlap(v) for v in raw_cells.values() if len(v) >= 2)
        if (len(cand) >= LEGEND_MIN_ENTRIES and _diverse(cand) >= LEGEND_MIN_ENTRIES
                and not overlapped):
            occ_cols = sorted({col_x[_nearest(b[0], col_x)] for b in cand})
            occ_rows = sorted({row_y[_nearest(b[1], row_y)] for b in cand})
            cxs = [b[0] for b in cand]; cys = [b[1] for b in cand]
            if (_regular(occ_cols) and _regular(occ_rows)
                    and (max(cxs) - min(cxs)) <= 0.85 * W
                    and (max(cys) - min(cys)) <= 0.85 * H):
                grid_cells = cand

    if grid_cells is not None and _diverse(grid_cells) >= _diverse(single_best):
        best = grid_cells
    else:
        best = single_best

    # remember the table skeleton for reporting / debugging
    _gc_x = _cluster1d([b[0] for b in best], 18)
    _gc_y = _cluster1d([b[1] for b in best], 7)
    globals()['_LAST_LEGEND_GRID'] = {
        'cols': [round(c) for c in _gc_x], 'rows': [round(r) for r in _gc_y],
        'cells': [(round(b[0]), round(b[1]), tuple(int(z) for z in b[2])) for b in best]}
    print(f"  [legend table: {len(_gc_x)} cols x {len(_gc_y)} rows, {len(best)} cells | "
          f"cols_x={[round(c) for c in _gc_x]} rows_y={[round(r) for r in _gc_y]}]")

    xs = [b[0] for b in best]; ys = [b[1] for b in best]
    colors = [(b[0], b[1], b[2]) for b in best]
    # A legend entry is a swatch PLUS its adjacent label text (alphanumerics).
    # Extend the box to the right to cover that text: from each swatch, scan the
    # dark low-sat text band to its right and take how far it reaches. This
    # captures "Cohort 1, 1 mg/kg" etc. so the legend box wraps the whole entry.
    sx = img_hsv[:, :, 1]; vx = img_hsv[:, :, 2]
    txt_right = int(max(xs))
    for (cx, cy, _c) in best:
        ty0, ty1 = int(max(0, cy - 12)), int(min(H, cy + 12))
        x = int(cx) + 4
        gap = 0; reach = int(cx)
        while x < W and gap < 25:
            colband = (vx[ty0:ty1, x] < 130) & (sx[ty0:ty1, x] < 70)
            if colband.any():
                reach = x; gap = 0
            else:
                gap += 1
            x += 1
        txt_right = max(txt_right, reach)
    box = (int(min(xs)) - 6, int(min(ys)) - 10,
           int(min(W - 1, txt_right + 4)), int(max(ys)) + 10)
    return ('colors', colors, box)


def _classify_legend_swatches(legend_box, existing_rgbs=None):
    """Decide whether the legend contains ACHROMATIC (grey / black / open-marker)
    entries, using the legend's OWN table structure rather than guessing from the
    plot (which is fooled by black error bars).

    Design (layout-agnostic, incl. multi-column legends):
      1. Detect the CHROMATIC swatches -> their x-centres give the swatch COLUMN
         positions (the table skeleton).
      2. Detect compact ACHROMATIC marks (filled square, line, or open ring).
      3. Keep an achromatic mark only if it sits AT a swatch column (aligned with
         the chromatic grid) AND has a text label to its right AND does not
         overlap a chromatic swatch -- i.e. it occupies an otherwise-empty cell of
         the legend table (a genuine grey/black entry such as 'Placebo').
    Returns the achromatic swatch RGBs not already in `existing_rgbs`.
    """
    if legend_box is None:
        return []
    x0, y0, x1, y1 = [int(v) for v in legend_box]
    x0 = max(0, x0); y0 = max(0, y0); x1 = min(W - 1, x1); y1 = min(H - 1, y1)
    if (x1 - x0) < 8 or (y1 - y0) < 8:
        return []
    reg_rgb = img_rgb[y0:y1 + 1, x0:x1 + 1]
    ss = img_hsv[y0:y1 + 1, x0:x1 + 1, 1]
    vv = img_hsv[y0:y1 + 1, x0:x1 + 1, 2]
    Hh, Ww = vv.shape

    # 1) table skeleton (columns / rows). Prefer the robust grid already found by
    #    the unified detector; fall back to a local chromatic-swatch scan.
    _grid = globals().get('_LAST_LEGEND_GRID')
    if _grid and _grid.get('cells'):
        chrom_cent = [(cx - x0, cy - y0) for (cx, cy, _c) in _grid['cells']]
        col_xs = [c - x0 for c in _grid['cols']]
        row_ys = [r - y0 for r in _grid['rows']]
    else:
        chrom = ((ss > 45) & (vv > 40) & (vv < 250)).astype(np.uint8)
        chrom = cv2.morphologyEx(chrom, cv2.MORPH_CLOSE, np.ones((3, 3), np.uint8))
        ncc, _lc, stc, cenc = cv2.connectedComponentsWithStats(chrom, 8)
        chrom_cent = []
        for i in range(1, ncc):
            a = stc[i, cv2.CC_STAT_AREA]; w = stc[i, cv2.CC_STAT_WIDTH]; h = stc[i, cv2.CC_STAT_HEIGHT]
            if 15 <= a <= 2500 and w <= 120 and h <= 60:
                chrom_cent.append((float(cenc[i][0]), float(cenc[i][1])))
        col_xs = []; row_ys = []
        for cx, cy in chrom_cent:
            if not any(abs(cx - c) <= 15 for c in col_xs):
                col_xs.append(cx)
            if not any(abs(cy - r) <= 10 for r in row_ys):
                row_ys.append(cy)
    if not chrom_cent or not col_xs:
        return []               # no chromatic grid -> can't separate swatch from text
    min_col = min(col_xs)

    # 2) compact achromatic marks that are a legend SWATCH (not label text). Text
    #    always sits to the RIGHT of a swatch, so we anchor to the chromatic grid:
    #    accept an achromatic mark only if it is (A) at a chromatic COLUMN x (a
    #    grey/black entry sharing a column with colour entries, e.g. Placebo), or
    #    (B) LEFT of the leftmost colour column while aligned to a chromatic ROW
    #    (a lone achromatic column such as a black 'open circle' entry). Both rule
    #    out label text, which is neither at a column nor left of the swatches.
    ach = ((ss <= 45) & (vv < 220) & (vv > 25)).astype(np.uint8)
    ach = cv2.morphologyEx(ach, cv2.MORPH_CLOSE, np.ones((3, 3), np.uint8))
    na, lba, sta, cena = cv2.connectedComponentsWithStats(ach, 8)
    found = []
    for i in range(1, na):
        a = sta[i, cv2.CC_STAT_AREA]; w = sta[i, cv2.CC_STAT_WIDTH]; h = sta[i, cv2.CC_STAT_HEIGHT]
        if not (12 <= a <= 2500 and 3 <= w <= 60 and 2 <= h <= 40):
            continue
        cx, cy = float(cena[i][0]), float(cena[i][1])
        if any(abs(cx - c[0]) <= 12 and abs(cy - c[1]) <= 8 for c in chrom_cent):
            continue                                     # this IS a chromatic swatch
        at_col = any(abs(cx - c) <= 14 for c in col_xs)
        left_new = (cx < min_col - 6) and any(abs(cy - r) <= 10 for r in row_ys)
        if not (at_col or left_new):
            continue                                     # label text -> rejected
        # must lie within the legend's ROW rhythm (else it's axis/title text that
        # merely happens to align with a column in an over-tall legend box)
        _rs = sorted(row_ys)
        _pitch = float(np.median(np.diff(_rs))) if len(_rs) > 1 else 18.0
        if not (min(_rs) - 1.3 * _pitch <= cy <= max(_rs) + 1.3 * _pitch):
            continue
        x_r = sta[i, cv2.CC_STAT_LEFT] + w
        tx0, tx1 = int(x_r + 4), int(min(Ww, x_r + 150))
        ty0, ty1 = int(max(0, cy - 9)), int(min(Hh, cy + 9))
        if tx1 <= tx0:
            continue
        txt = int(((vv[ty0:ty1, tx0:tx1] < 120) & (ss[ty0:ty1, tx0:tx1] < 60)).sum())
        if txt < 50:
            continue
        _comp = reg_rgb[lba == i]
        _cmed = np.median(_comp, axis=0)
        if (int(_cmed.max()) - int(_cmed.min())) <= 28:
            # neutral swatch: median mixes the dark marker with white/AA and comes
            # out mid-grey. Use the dark marker core so a black marker reads black.
            _Ls = _comp.sum(axis=1)
            _core = _comp[_Ls <= np.percentile(_Ls, 30)]
            col = tuple(int(z) for z in (np.mean(_core, axis=0) if len(_core) >= 3 else _cmed))
        else:
            col = tuple(int(z) for z in _cmed)
        found.append((col, int(cx) + x0, int(cy) + y0))

    # 3) dedup (achromatic colours compared by lightness, leniently)
    def _neutral(c):
        return (max(c) - min(c)) <= 28
    def _same(c, e):
        if _neutral(c) and _neutral(e):
            return abs(np.mean(c) - np.mean(e)) < 55
        return abs(c[0]-e[0]) + abs(c[1]-e[1]) + abs(c[2]-e[2]) < 40
    out = []; out_pos = []
    for c, px, py in found:
        if existing_rgbs and any(_same(c, e) for e in existing_rgbs):
            continue
        if any(_same(c, o) for o in out):
            continue
        out.append(c); out_pos.append((px, py, tuple(int(z) for z in c)))
    globals()['_LAST_ACHRO_SWATCHES'] = out_pos
    return out


def _sample_legend_cell(cx, cy, bw=22, bh=11):
    """Sample the swatch colour at a legend-grid intersection directly from the
    image (fixed box). Returns {'type','rgb'} or None if the spot is essentially
    blank. Used to FILL cells the blob detector missed (faint / open / dashed
    markers, or a black marker that failed the solidity gate) once the regular
    row lattice tells us a swatch MUST be there."""
    x0 = max(0, int(cx) - bw // 2); x1 = min(W, int(cx) + bw // 2 + 1)
    y0 = max(0, int(cy) - bh // 2); y1 = min(H, int(cy) + bh // 2 + 1)
    win = img_rgb[y0:y1, x0:x1].reshape(-1, 3)
    if win.size == 0:
        return None
    nw = ~((win[:, 0] > 232) & (win[:, 1] > 232) & (win[:, 2] > 232))
    ink = win[nw]
    if len(ink) < 6:                      # basically white -> genuinely empty
        return None
    med = np.median(ink, axis=0).astype(int)
    chroma = int(med.max()) - int(med.min())
    if chroma <= 25:
        # Achromatic marker (e.g. a black circle): the median mixes the dark
        # marker with the surrounding white/anti-aliasing and comes out mid-grey,
        # which would make the palette slot attract every faded neutral pixel in
        # the plot. Use the DARK marker core instead (mean of the darkest ~30% of
        # ink pixels by lightness) so the slot's colour is the real marker colour.
        Ls = ink.sum(axis=1)
        thr = np.percentile(Ls, 30)
        core = ink[Ls <= thr]
        if len(core) >= 3:
            dc = np.mean(core, axis=0).astype(int)
            return {'type': 'achro', 'rgb': [int(dc[0]), int(dc[1]), int(dc[2])]}
        return {'type': 'achro', 'rgb': [int(med[0]), int(med[1]), int(med[2])]}
    return {'type': 'chrom',
            'rgb': [int(med[0]), int(med[1]), int(med[2])]}


def _build_unified_legend_grid(grid_cells, achro_swatches, align_tol=4, regularize=True):
    """Combine chromatic swatches (grid_cells) and achromatic swatches into ONE
    STRICT table: cluster columns (x) and rows (y) over ALL swatches, snap each to
    the nearest (col,row) intersection, and record which cells are filled (chrom /
    achro) vs EMPTY.

    Legend rows are EVENLY SPACED by construction. So when >=3 rows are found we
    estimate the row pitch and rebuild rows as a REGULAR LATTICE, inserting the
    positions of rows the detector missed. Every (col,row) intersection with no
    detected swatch is then re-sampled straight from the image; if ink is present
    there, the cell is filled (this recovers faint / open / dashed / black markers
    that the blob detector dropped even though their row clearly exists).
    Returns dict(cols, rows, cells{(ci,ri):{type,rgb}}, empty[list], aligned, resid)."""
    sw = ([(c[0], c[1], 'chrom', tuple(c[2])) for c in (grid_cells or [])] +
          [(a[0], a[1], 'achro', tuple(a[2])) for a in (achro_swatches or [])])
    if not sw:
        return None

    def _cl(vals, tol):
        vals = sorted(vals); g = [[vals[0]]]
        for v in vals[1:]:
            (g[-1].append(v) if v - g[-1][-1] <= tol else g.append([v]))
        return [float(np.mean(x)) for x in g]

    def _near(v, cs):
        return int(min(range(len(cs)), key=lambda i: abs(v - cs[i])))

    cols = sorted(_cl([s[0] for s in sw], 18))
    rows = sorted(_cl([s[1] for s in sw], 7))

    # -- regularise rows into an even lattice + insert missing row positions -----
    filled_from_sample = False
    if regularize and len(rows) >= 3:
        gaps = sorted(rows[i + 1] - rows[i] for i in range(len(rows) - 1))
        pitch = gaps[0]                                   # smallest gap = 1 row step
        # guard against a spuriously tiny gap: use the smallest gap that is at
        # least 60% of the median gap (rejects a merged double-detection).
        _med = gaps[len(gaps) // 2]
        for g in gaps:
            if g >= 0.6 * _med:
                pitch = g; break
        if pitch >= 6:
            n = int(round((rows[-1] - rows[0]) / pitch))
            if 0 < n <= 40:
                # Regular lattice spanning ONLY the detected row range, so missing
                # MIDDLE rows are inserted at their true positions. We do not
                # extend past the outermost detected rows (that risks grabbing
                # margin / label / plot ink), so as long as the top and bottom
                # entries were detected the whole stack is reconstructed.
                rows = [rows[0] + k * pitch for k in range(n + 1)]

    cells = {}; resid = 0.0
    for x, y, t, rgb in sw:
        ci = _near(x, cols); ri = _near(y, rows)
        resid = max(resid, abs(x - cols[ci]), abs(y - rows[ri]))
        cells[(ci, ri)] = {'type': t, 'rgb': list(rgb)}

    # -- fill any intersection the detector missed by sampling the image ---------
    if regularize:
        for ci, cx in enumerate(cols):
            for ri, ry in enumerate(rows):
                if (ci, ri) in cells:
                    continue
                got = _sample_legend_cell(cx, ry)
                if got is not None:
                    cells[(ci, ri)] = got
                    filled_from_sample = True

    empty = [(ci, ri) for ci in range(len(cols)) for ri in range(len(rows))
             if (ci, ri) not in cells]
    return {
        'cols': [round(c) for c in cols], 'rows': [round(r) for r in rows],
        'cells': {f'{ci},{ri}': v for (ci, ri), v in cells.items()},
        'empty': [f'{ci},{ri}' for (ci, ri) in empty],
        'filled_from_sample': filled_from_sample,
        'aligned': bool(resid <= align_tol), 'residual': round(float(resid), 1)}


def _pcm_detect_legend_panel_aux():
    """Auxiliary legend detector for legends the primary one misses (small or
    low-density swatch markers, e.g. a 12px o # ^ column). Used ONLY as a
    fallback, so the well-tested primary path for normal legends is untouched.

    A real legend is a column of small marker glyphs that are (a) left-aligned at
    a common x, (b) evenly spaced in y, and (c) DIFFERENT colours. Plot-interior
    curve cross-sections may align and even space out, but they are the SAME
    colour -- colour diversity is what separates the two. We find peaks of left-
    aligned compact colour runs, cluster them by x-start, and keep the cluster
    that is most evenly spaced AND most colour-diverse.
    """
    strip_x0 = int(W * (1 - LEGEND_RIGHT_FRAC))
    strip    = img_hsv[:, strip_x0:, :]
    rgb_strip = img_rgb[:, strip_x0:, :]
    bg       = _pcm_background_mask(strip)
    fg       = ~bg
    s_strip  = strip[:, :, 1]
    v_strip  = strip[:, :, 2]
    chrom    = fg & (s_strip > 20)
    dark     = fg & (s_strip <= 40) & (v_strip >= 20) & (v_strip < 200)
    swatch_px = chrom | dark
    row_sum  = swatch_px.sum(axis=1).astype(float)
    strip_w  = strip.shape[1]
    # candidate rows: some ink, but not a near-full row (which is a dense band)
    cand = (row_sum > 0) & (row_sum / strip_w < LEGEND_MAX_SWATCH_F)
    sm   = gaussian_filter1d(row_sum * cand.astype(float), sigma=1)
    peaks, _ = find_peaks(sm, distance=LEGEND_PEAK_DIST, prominence=LEGEND_PEAK_PROM)
    if len(peaks) < LEGEND_MIN_ENTRIES:
        return None

    # For each peak, take the leftmost compact run = the marker glyph; record its
    # x-start and median colour. Require the run to be reasonably solid.
    info = []   # (y, x_start, (r,g,b))
    for p in peaks:
        rm = swatch_px[p, :]
        cols = np.where(rm)[0]
        if len(cols) == 0:
            continue
        rs = int(cols[0]); re = rs
        for c in cols[1:]:
            if c - re <= 3:
                re = int(c)
            else:
                break
        run_w  = re - rs + 1
        run_px = int(rm[rs:re+1].sum())
        if run_w < 4 or run_px / run_w < 0.5:
            continue
        seg = rgb_strip[p, rs:re+1, :]
        info.append((int(p), rs, tuple(np.median(seg, axis=0).astype(int))))
    if len(info) < LEGEND_MIN_ENTRIES:
        return None

    xs0 = np.array([t[1] for t in info])
    ys0 = np.array([t[0] for t in info])

    def _diversity(idxs):
        cols = [info[i][2] for i in idxs]
        uniq = []
        for c in cols:
            if all(abs(c[0]-u[0]) + abs(c[1]-u[1]) + abs(c[2]-u[2]) > 60 for u in uniq):
                uniq.append(c)
        return len(uniq)

    best_sel, best_score = None, -1e9
    for i in range(len(info)):
        sel = np.where(np.abs(xs0 - xs0[i]) <= 20)[0]
        if len(sel) < LEGEND_MIN_ENTRIES:
            continue
        ys = np.sort(ys0[sel])
        gaps = np.diff(ys)
        if len(gaps) == 0:
            continue
        cv_sp = gaps.std() / max(1.0, gaps.mean())
        ndiv  = _diversity(sel)
        # A legend MUST be colour-diverse: at least 3 distinct colours, and most
        # of its entries distinct. Reject same-colour (curve) clusters outright.
        if ndiv < min(LEGEND_MIN_ENTRIES, len(sel)) or ndiv < 3:
            continue
        # Even spacing is the decisive legend signature -- a real legend is almost
        # perfectly regular (cv~0), whereas data cross-sections are irregular
        # (cv large). Weight it heavily so a tidy 3-entry legend beats a ragged
        # 4-entry data cluster. Diversity and count are secondary tie-breakers.
        score = ndiv * 1.0 + len(sel) * 0.5 - cv_sp * 30.0
        if score > best_score:
            best_score, best_sel = score, sel
    if best_sel is None:
        return None

    ys = np.sort(ys0[best_sel])
    gaps = np.diff(ys)
    step = int(np.median(gaps)) if len(gaps) else 16
    pad = max(12, step // 2)
    y0_leg = max(0, int(ys[0]) - pad)
    y1_leg = min(H - 1, int(ys[-1]) + pad)
    # left x: anchor on the aligned marker x-start
    x0_leg = strip_x0 + int(np.median(xs0[best_sel])) - 4
    return (max(0, x0_leg), y0_leg, W - 1, y1_leg)


def _pcm_detect_legend_panel_horizontal():
    """DEPRECATED -- superseded by _pcm_detect_legend_unified(), which handles both
    vertical and horizontal legends in one pass. Kept for reference; not called."""
    s = img_hsv[:, :, 1]; v = img_hsv[:, :, 2]
    chrom = (s > 40) & (v > 40) & (v < 250)
    if chrom.sum() == 0:
        return None
    # find the row band with the most chromatic ink in the top/bottom margins
    row_chrom = chrom.sum(axis=1).astype(float)
    # legends sit in the margins, away from the bulk of the plot; search the
    # top 20% and bottom 20% of the image for a dominant chromatic row.
    bands = list(range(0, int(H * 0.20))) + list(range(int(H * 0.80), H))
    if not bands:
        return None
    best_y = max(bands, key=lambda y: row_chrom[y])
    if row_chrom[best_y] < W * 0.04:
        return None
    # collect chromatic pixels in a thin band around best_y
    y0b = max(0, best_y - 8); y1b = min(H, best_y + 9)
    band = chrom[y0b:y1b, :]
    col_has = band.any(axis=0)
    xs_c = np.where(col_has)[0]
    if len(xs_c) < 3:
        return None
    # cluster the chromatic columns into segments (the individual swatches)
    segs = []
    run = [xs_c[0]]
    for x in xs_c[1:]:
        if x - run[-1] <= 25:
            run.append(x)
        else:
            segs.append((run[0], run[-1])); run = [x]
    segs.append((run[0], run[-1]))
    # each segment's colour
    seg_info = []
    for (xa, xb) in segs:
        sub = img_rgb[y0b:y1b, xa:xb+1]
        subm = chrom[y0b:y1b, xa:xb+1]
        if subm.sum() < 8:
            continue
        col = tuple(np.median(sub[subm], axis=0).astype(int))
        seg_info.append((xa, xb, col))
    if len(seg_info) < LEGEND_MIN_ENTRIES:
        return None
    # require colour diversity (a row of same-colour blobs is not a legend)
    uniq = []
    for _, _, c in seg_info:
        if all(abs(c[0]-u[0]) + abs(c[1]-u[1]) + abs(c[2]-u[2]) > 50 for u in uniq):
            uniq.append(c)
    if len(uniq) < LEGEND_MIN_ENTRIES:
        return None
    return ('horizontal', y0b, y1b, seg_info)


def _pcm_detect_legend_panel():
    strip_x0 = int(W * (1 - LEGEND_RIGHT_FRAC))
    # Scan most of the image height (legends can extend low), not just top 65%.
    strip    = img_hsv[:, strip_x0:, :]
    bg       = _pcm_background_mask(strip)
    fg       = ~bg
    s_strip  = strip[:, :, 1]
    v_strip  = strip[:, :, 2]
    # A swatch row contains either chromatic pixels OR a compact run of dark
    # pixels (black/grey markers). Include both so achromatic legend entries
    # (black, grey) don't break the panel mid-way.
    chrom    = fg & (s_strip > 20)
    dark     = fg & (s_strip <= 40) & (v_strip >= 20) & (v_strip < 200)
    swatch_px = chrom | dark
    row_chrom = swatch_px.sum(axis=1).astype(float)
    strip_w   = strip.shape[1]
    row_frac  = row_chrom / max(strip_w, 1)
    candidate = (row_frac > 0) & (row_frac < LEGEND_MAX_SWATCH_F)
    candidate_rows = np.where(candidate & (row_chrom >= LEGEND_MIN_PX))[0]
    if len(candidate_rows) == 0:
        return None
    smoothed = gaussian_filter1d(row_chrom * candidate.astype(float), sigma=1)
    peaks, _ = find_peaks(smoothed, distance=LEGEND_PEAK_DIST,
                           prominence=LEGEND_PEAK_PROM)
    # Filter peaks by swatch density (legend swatches are dense AND left-aligned;
    # curve cross-sections are sparse and scattered).
    valid_peaks = []
    swatch_x_starts = []
    swatch_run_rgb  = []
    rgb_strip = img_rgb[:, strip_x0:, :]
    for p in peaks:
        row_mask = swatch_px[p, :]
        if row_mask.any():
            cols = np.where(row_mask)[0]
            width = max(1, int(cols[-1]) - int(cols[0]) + 1)
            density = float(row_chrom[p]) / width
            if density >= LEGEND_SWATCH_DENSITY:
                valid_peaks.append(int(p))
                # leftmost compact run = the swatch marker (for x-align + colour)
                run_start = int(cols[0]); run_end = run_start
                for c in cols[1:]:
                    if c - run_end <= 3:
                        run_end = int(c)
                    else:
                        break
                swatch_x_starts.append(run_start)
                seg = rgb_strip[p, run_start:run_end+1, :]
                swatch_run_rgb.append(tuple(np.median(seg, axis=0).astype(int)))
    if len(valid_peaks) < LEGEND_MIN_ENTRIES:
        return None
    # Legend entries share a common left x-start (they're vertically aligned).
    # Anchor on the median start and keep aligned peaks; this rejects scattered
    # plot markers whose x-starts vary.
    if len(swatch_x_starts) >= 2:
        med_x = int(np.median(swatch_x_starts))
        aligned = [vp for vp, sx in zip(valid_peaks, swatch_x_starts)
                   if abs(sx - med_x) <= 25]
        if len(aligned) >= LEGEND_MIN_ENTRIES:
            valid_peaks = aligned
    # Legend entries are evenly spaced. The peak finder may miss faint (e.g.
    # grey) or merged entries, which truncates the panel and drops colours.
    # Extend the panel by one inter-entry gap on each end so every swatch row
    # is inside the box, then let swatch sampling recover all entries.
    vpk = sorted(valid_peaks)
    if len(vpk) >= 2:
        gaps = np.diff(vpk)
        step = int(np.median(gaps))
        pad = max(12, step // 2)
    else:
        pad = 12
    y0_leg = max(0, vpk[0] - pad)
    y1_leg = min(H - 1, vpk[-1] + pad)
    leg_strip = swatch_px[y0_leg:y1_leg+1, :]
    col_has_chrom = leg_strip.any(axis=0)
    if not col_has_chrom.any():
        return None
    x0_rel = int(np.where(col_has_chrom)[0].min())
    x0_leg = strip_x0 + x0_rel
    return (x0_leg, y0_leg, W-1, y1_leg)

def _pcm_sample_palette_from_legend(legend_box):
    """Extract curve colours from legend box using row-band analysis.
    Works for both internal and external legends.
    Handles chromatic (coloured) and achromatic (grey/black) entries.
    """
    if legend_box is None:
        return []
    x0, y0, x1, y1 = legend_box
    # Clamp to image bounds and normalise ordering. A user-drawn box can be
    # off-image, inverted, or degenerate; guard against an empty crop that would
    # make downstream .max()/.sum() reductions fail on a zero-size array.
    x0, x1 = sorted((int(x0), int(x1)))
    y0, y1 = sorted((int(y0), int(y1)))
    x0 = max(0, min(x0, W - 1)); x1 = max(0, min(x1, W - 1))
    y0 = max(0, min(y0, H - 1)); y1 = max(0, min(y1, H - 1))
    if x1 - x0 < 2 or y1 - y0 < 2:
        return []
    region_hsv = img_hsv[y0:y1+1, x0:x1+1, :]
    region_lab = img_lab[y0:y1+1, x0:x1+1, :]
    region_rgb = img_rgb[y0:y1+1, x0:x1+1, :]
    if region_hsv.size == 0:
        return []
    sat = region_hsv[:, :, 1]
    val = region_hsv[:, :, 2]
    Rw  = region_hsv.shape[1]

    # Foreground = any non-white pixel (include dark swatches like black markers).
    # We separate swatch from text by COLUMN position, not by darkness, so we
    # don't accidentally drop black/dark-coloured swatches.
    fg = ~((val > 235) & (sat < 15))   # exclude near-white background
    fg &= (val < 250)

    # -- Isolate the MARKER region (shape-independent) ---------------------
    # A legend row is [marker][gap][text]. The marker can be any shape (filled
    # square, hollow triangle, diamond, small circle, line). A narrow column
    # misses small/hollow markers because few of their pixels fall in it. Instead
    # we take the WHOLE marker block: the first run of foreground columns from
    # the left, stopping at the white gap that separates marker from text. This
    # captures every marker fully regardless of shape, so small * # o swatches
    # register as many pixels as large ones.
    col_fg = fg.sum(axis=0).astype(float)
    if col_fg.size == 0 or col_fg.max() == 0:
        return []
    col_sm = gaussian_filter1d(col_fg, sigma=2)
    left_limit = max(8, int(Rw * 0.55))
    # marker region begins at the first column whose density rises above the
    # EMPTY-background level, and ends at the first sustained gap. Baseline is a
    # low percentile (empty level), not the median: a text-heavy legend inflates
    # the median above a small marker's bump, making the detector land on the
    # text and sample its (black) colour instead of the marker's.
    base = float(np.percentile(col_sm[:left_limit], 10)) if left_limit > 0 else 0
    rise_thr = base + 2.0
    above = np.where(col_sm[:left_limit] >= rise_thr)[0]
    if len(above) > 0:
        m_start = int(above[0])
        # The marker is the FIRST bump. It ends either at a gap (density falls to
        # baseline -- chart6's small marker then text) OR at the first VALLEY
        # between the marker bump and the text bump (chart4: marker abuts text
        # with no empty gap, but there is still a local minimum between them).
        gap_thr = base + 1.0
        m_end = m_start
        run_low = 0
        rose = False
        prev = col_sm[m_start]
        peak_val = col_sm[m_start]
        for c in range(m_start + 1, left_limit):
            cv_ = col_sm[c]
            peak_val = max(peak_val, cv_)
            # gap termination
            if cv_ < gap_thr:
                run_low += 1
                if run_low >= 4:
                    break
            else:
                run_low = 0
            # valley termination: we've come down from the peak and started rising
            # again (text bump beginning). End the marker at the valley.
            if cv_ > prev and peak_val > base + 4 and prev <= peak_val * 0.55:
                m_end = c - 1
                break
            if cv_ >= gap_thr:
                m_end = c
            prev = cv_
        sc_x0 = max(0, m_start - 2)
        sc_x1 = min(Rw - 1, m_end + 2)
    else:
        sc_x0, sc_x1 = 0, min(left_limit, 30)

    swatch_fg = np.zeros_like(fg)
    swatch_fg[:, sc_x0:sc_x1+1] = fg[:, sc_x0:sc_x1+1]
    ys, xs = np.where(swatch_fg)
    if len(ys) < 10:
        swatch_fg = fg  # fallback: use full width
        sc_x0, sc_x1 = 0, Rw - 1
        ys, xs = np.where(swatch_fg)
        if len(ys) < 10:
            return []

    # Group rows into horizontal bands (gap > 8px = new legend entry)
    all_rows = sorted(set(ys.tolist()))
    bands = []
    band = [all_rows[0]]
    for r in all_rows[1:]:
        if r - band[-1] > 8:
            bands.append(band)
            band = [r]
        else:
            band.append(r)
    bands.append(band)

    # Drop tiny bands: a real swatch spans several rows. A 1-2px band is a stray
    # fragment (anti-aliasing, a text descender).
    bands = [b for b in bands if (b[-1] - b[0] + 1) >= 3]

    # Filter bands by COLOUR PRESENCE, not pixel count. A small/hollow marker
    # (* # o) has few pixels but they are vividly coloured (or solidly black);
    # a stray fragment is faint or near-background. Keep a band if it contains a
    # core of clearly-coloured OR clearly-dark pixels. This keeps small markers
    # while rejecting anti-aliasing wisps, independent of marker size/shape.
    def _band_has_ink(b):
        ps = sat[b[0]:b[-1]+1, sc_x0:sc_x1+1]
        pv = val[b[0]:b[-1]+1, sc_x0:sc_x1+1]
        # Chromatic ink: high saturation (no V cap -- bright vivid markers reach
        # V~248; saturation alone separates them from the white background).
        chroma_px = int(((ps > 55) & (pv > 40)).sum())
        # Achromatic ink: low-saturation but clearly darker than the page. This
        # covers black markers AND mid-grey swatches (e.g. a grey 'Placebo'
        # line at V~145, which is neither vivid nor near-black).
        grey_px   = int(((ps <= 55) & (pv >= 30) & (pv <= 215)).sum())
        return (chroma_px >= LEGEND_INK_MIN) or (grey_px >= LEGEND_INK_MIN)
    bands = [b for b in bands if _band_has_ink(b)]

    # Reject bands that are far larger than the typical entry: those are plot
    # data captured inside an over-extended panel, not legend swatches.
    if len(bands) >= 3:
        heights = sorted(b[-1] - b[0] + 1 for b in bands)
        med_h = heights[len(heights)//2]
        bands = [b for b in bands if (b[-1] - b[0] + 1) <= med_h * 2.5]

    # Drop EXTREME low-pixel outliers: a phantom fragment squeezed between real
    # entries (e.g. a stray arrow tip) has far fewer ink pixels than any true
    # swatch. Use a low fraction so genuinely small markers (* # o) still pass.
    if len(bands) >= 4:
        def _ink_px(b):
            ps = sat[b[0]:b[-1]+1, sc_x0:sc_x1+1]
            pv = val[b[0]:b[-1]+1, sc_x0:sc_x1+1]
            return int((((ps > 55) & (pv > 40)) |
                        ((ps <= 55) & (pv >= 30) & (pv <= 215))).sum())
        pxs = [_ink_px(b) for b in bands]
        med = float(np.median(pxs))
        kept = [b for b, p in zip(bands, pxs) if p >= med * 0.20]
        if len(kept) >= LEGEND_MIN_ENTRIES:
            bands = kept

    palette = []
    swatch_info = []   # parallel list: per-entry dict with Lab samples + bbox
    for band in bands:
        r_min, r_max = band[0], band[-1]
        patch_sat = sat[r_min:r_max+1, sc_x0:sc_x1+1]
        patch_val = val[r_min:r_max+1, sc_x0:sc_x1+1]
        patch_lab = region_lab[r_min:r_max+1, sc_x0:sc_x1+1, :]
        patch_rgb = region_rgb[r_min:r_max+1, sc_x0:sc_x1+1, :]

        # Prefer chromatic pixels; fall back to achromatic (grey/black) swatches.
        # Chromatic = saturated (no V ceiling: bright vivid markers reach V~248).
        # Allow dark swatches down to V=2 (near-black markers/lines).
        chrom_mask = (patch_sat > 40) & (patch_val > 30)
        grey_mask  = (patch_sat <= 40) & (patch_val >= 2) & (patch_val < 220)

        if chrom_mask.sum() >= LEGEND_MIN_PX:
            use_mask = chrom_mask
            is_chrom = True
        elif grey_mask.sum() >= LEGEND_MIN_PX:
            use_mask = grey_mask
            is_chrom = False
        else:
            continue

        # -- Representative INK colour, not the mean ----------------------
        # A swatch is line/marker ink surrounded by white gaps and anti-aliased
        # halo. Averaging all foreground pixels drags the colour toward the
        # halo/white (e.g. a black o becomes mid-grey). Instead take the
        # "ink core": for chromatic swatches the most-saturated pixels, for
        # achromatic swatches the darkest pixels -- then use their median.
        sel_lab = patch_lab[use_mask].reshape(-1, 3).astype(np.float32)
        sel_sat = patch_sat[use_mask].reshape(-1).astype(np.float32)
        sel_val = patch_val[use_mask].reshape(-1).astype(np.float32)
        if is_chrom:
            # ink = top saturation tier (most vivid pixels define the colour)
            thr = np.percentile(sel_sat, 100 - INK_CORE_PCT)
            core = sel_sat >= thr
        else:
            # ink = darkest tier (a black marker's core is the darkest pixels)
            thr = np.percentile(sel_val, INK_CORE_PCT)
            core = sel_val <= thr
        if core.sum() < max(5, LEGEND_MIN_PX // 4):
            core = np.ones(len(sel_lab), dtype=bool)  # fallback: use all
        lab_rep = np.median(sel_lab[core], axis=0).astype(np.float32)

        rep_rgb = cv2.cvtColor(np.array([[lab_rep]], np.uint8),
                               cv2.COLOR_Lab2RGB)[0, 0].astype(int)
        print(f"    Legend swatch y={r_min}-{r_max}: ink RGB={tuple(rep_rgb)}  "
              f"n={int(use_mask.sum())} core={int(core.sum())} "
              f"{'chrom' if is_chrom else 'achrom'}")
        palette.append(lab_rep)
        # Store the INK-core samples for S-floor learning + verification.
        swatch_info.append({
            'lab_samples': sel_lab[core],
            's_samples':   sel_sat[core],
            'bbox': (y0 + r_min, y0 + r_max, x0 + sc_x0, x0 + sc_x1),
        })

    # Merge near-duplicate entries (keep swatch_info in sync)
    merged = True
    while merged:
        merged = False
        for i in range(len(palette)):
            for j in range(i+1, len(palette)):
                if np.linalg.norm(palette[i] - palette[j]) < LEGEND_LAB_MERGE:
                    new_lab = ((palette[i]+palette[j])/2).astype(np.float32)
                    new_info = {
                        'lab_samples': np.vstack([swatch_info[i]['lab_samples'],
                                                  swatch_info[j]['lab_samples']]),
                        's_samples': np.concatenate([swatch_info[i]['s_samples'],
                                                     swatch_info[j]['s_samples']]),
                        'bbox': swatch_info[i]['bbox'],
                    }
                    keep_idx = [k for k in range(len(palette)) if k not in (i, j)]
                    palette = [new_lab] + [palette[k] for k in keep_idx]
                    swatch_info = [new_info] + [swatch_info[k] for k in keep_idx]
                    merged = True; break
            if merged: break
    # Attach swatch_info to the returned list via a module-level stash
    global _LEGEND_SWATCH_INFO
    _LEGEND_SWATCH_INFO = swatch_info
    return palette

def _is_achromatic(lab):
    return _chroma_of_lab(lab) < NC_ACHROMATIC_CHROMA


def _achromatic_mask_is_curve(mask):
    """Decide whether an achromatic (black/grey) colour mask is a REAL CURVE or
    just structural noise (the grey plot background, or scattered axis/text/tick
    fragments). In legend-less plots the black/grey "colour" is often not a
    plotted series at all, so extracting data points from it is wrong.

    A real curve is a thin, mostly-linear structure of moderate extent. We reject:
      * BACKGROUND  -- fills a large fraction of the image and is thick (survives
        erosion), e.g. a grey panel background.
      * FILLED BLOB -- high fill ratio inside its bounding box and thick.
      * SCATTERED   -- extremely sparse fragments spread over a wide bounding box
        with no component forming a meaningful connected run (stray marker bits,
        tick/grid-fragment dust), i.e. not a connected curve.
    Returns (is_curve: bool, reason: str).
    """
    m = (mask > 0).astype(np.uint8)
    px = int(m.sum())
    if px < 30:
        return False, "too few pixels"
    img_frac = px / float(H * W)
    ys, xs = np.where(m)
    bbox_area = (xs.max() - xs.min() + 1) * (ys.max() - ys.min() + 1)
    fill = px / float(max(1, bbox_area))
    bbox_frac = bbox_area / float(H * W)
    er = cv2.erode(m, np.ones((3, 3), np.uint8))
    thin_loss = 1.0 - er.sum() / float(max(1, px))
    n, lbl, st, _ = cv2.connectedComponentsWithStats(m, 8)
    comps = sorted((st[i, cv2.CC_STAT_AREA] for i in range(1, n)), reverse=True)
    largest_frac = (comps[0] / float(px)) if comps else 0.0

    # 1. background panel: huge area + thick
    if img_frac > 0.15 and thin_loss < 0.35:
        return False, f"background (img_frac={img_frac:.2f}, thin_loss={thin_loss:.2f})"
    # 2. filled solid blob (not a line)
    if fill > 0.25 and thin_loss < 0.40:
        return False, f"filled blob (fill={fill:.2f})"
    # 2b. horizontal rule: ink spans a wide x-range but is confined to a very
    #     narrow y-band (a single row). That is an axis / gridline / LLOQ rule,
    #     not a data curve (a curve rises and falls across its x-span).
    yspan = ys.max() - ys.min() + 1
    xspan_px = xs.max() - xs.min() + 1
    if xspan_px > 0.4 * W and yspan <= 4:
        return False, f"horizontal rule (yspan={yspan}, xspan_frac={xspan_px/float(W):.2f})"
    # 3. x-continuity: a real curve has ink in MOST x-columns across its span
    #    (it is a connected path left-to-right). Scattered noise -- stray marker
    #    fragments, tick dust -- leaves large horizontal gaps. We compute the
    #    fraction of x-columns (within the mask's x-extent) that contain any ink;
    #    a curve sits near 0.9-1.0, scattered noise well below. We only reject
    #    when the span is wide (so a genuinely short curve isn't penalised) and
    #    continuity is clearly low.
    cols = m.sum(axis=0)
    nzc = np.where(cols > 0)[0]
    if len(nzc) >= 2:
        xspan = nzc.max() - nzc.min() + 1
        x_continuity = len(nzc) / float(xspan)
        xspan_frac = xspan / float(W)
        if xspan_frac > 0.4 and x_continuity < 0.30:
            return False, (f"scattered noise (x_continuity={x_continuity:.2f}, "
                           f"xspan_frac={xspan_frac:.2f})")
    return True, f"curve (img_frac={img_frac:.3f}, fill={fill:.3f}, largest_frac={largest_frac:.2f})"


def _split_achromatic_by_shape(mask):
    """For an achromatic (grey/black) colour mask, separate genuine curve/marker
    pixels from linear noise (axis lines, grid lines, error bars) using component
    SHAPE -- a cue that survives when colour cannot distinguish them.

    Markers (filled circles, triangles, squares) are compact blobs with
    aspect~1 and high fill ratio. Axis/grid lines are extremely elongated
    (aspect very high or very low). Error bars are thin vertical strokes
    (width~1-2, tall). We keep blob-like and short components, drop long thin
    strokes that span a large fraction of the plot.

    Returns a cleaned boolean mask (markers + short connected curve segments).
    """
    m = mask.astype(np.uint8)
    if m.sum() == 0:
        return mask
    n, lbl, stats, _ = cv2.connectedComponentsWithStats(m, connectivity=8)
    keep = np.zeros_like(m, dtype=bool)
    for i in range(1, n):
        area = stats[i, cv2.CC_STAT_AREA]
        w    = stats[i, cv2.CC_STAT_WIDTH]
        h    = stats[i, cv2.CC_STAT_HEIGHT]
        if area < 3:
            continue
        long_side  = max(w, h)
        short_side = max(1, min(w, h))
        aspect     = long_side / short_side
        fill       = area / float(max(w * h, 1))
        spans_plot = long_side > SHAPE_LINE_SPAN * max(W, H)
        # Drop long thin strokes (axis / grid / long error bars / dashed rules)
        if spans_plot and aspect > SHAPE_LINE_ASPECT:
            continue
        # Drop thin vertical error-bar strokes: very narrow, tall, low fill width
        if short_side <= 2 and long_side > SHAPE_ERRBAR_MIN_LEN:
            continue
        keep[lbl == i] = True
    return keep


# -- Step 2: Colour discovery ---------------------------------------------------
def _chroma_of_lab(lab):
    """Perceptual chroma = sqrt(a*^2 + b*^2) in OpenCV Lab (a,b are offset by 128)."""
    a = float(lab[1]) - 128.0
    b = float(lab[2]) - 128.0
    return (a * a + b * b) ** 0.5


def _adaptive_radii(palette_lab):
    """Per-colour Lab acceptance radius based on chroma.

    Rationale: saturated colours (brown, magenta, blue) sit far from every other
    colour, so a wide radius safely captures their faded / anti-aliased curve
    pixels. Achromatic / pale colours (greys, light tan) cluster close together
    in Lab, so a wide radius would over-absorb neighbours -- they need a tight
    radius. We interpolate between NC_MIN_RADIUS and NC_MAX_DIST by chroma.
    """
    radii = []
    for lab in palette_lab:
        c = _chroma_of_lab(lab)
        # Map chroma 0..NC_CHROMA_FULL -> NC_MIN_RADIUS..NC_MAX_DIST
        t = min(1.0, c / NC_CHROMA_FULL)
        r = NC_MIN_RADIUS + t * (NC_MAX_DIST - NC_MIN_RADIUS)
        radii.append(r)
    return radii


def _build_masks_nearest_centroid(palette_lab, fg_mask, max_dists=None, s_mins=None,
                                  priority_bias=None):
    """Assign every foreground pixel to the single nearest palette colour in
    Lab space, so colour masks are MUTUALLY EXCLUSIVE by construction (no two
    ranges can overlap -- each pixel belongs to exactly one colour or none).

    A pixel is left UNASSIGNED (noise) if either:
      * its distance to the nearest centroid exceeds that colour's max_dist, or
      * the nearest centroid is not clearly closer than the 2nd-nearest
        (ambiguous pixel on a colour boundary), per NC_MARGIN, or
      * (chromatic colours only) the pixel's saturation is below that colour's
        learned S floor -- removes anti-aliased halo and grey pixels that were
        wrongly pulled to a chromatic centroid.

    Returns a list of boolean masks, one per palette entry.
    """
    K = len(palette_lab)
    if max_dists is None:
        max_dists = [NC_MAX_DIST] * K
    ys, xs = np.where(fg_mask)
    if len(ys) == 0:
        return [np.zeros((H, W), dtype=bool) for _ in range(K)]
    px_lab = img_lab[ys, xs, :].astype(np.float32)        # (N,3)
    px_s   = img_hsv[ys, xs, 1].astype(np.float32)        # (N,)  saturation
    px_v   = img_hsv[ys, xs, 2].astype(np.float32)        # (N,)  value
    cents  = np.stack(palette_lab).astype(np.float32)      # (K,3)
    achromatic = np.array([_is_achromatic(lab) for lab in palette_lab])

    # Standard Lab distance (N,K)
    diff = px_lab[:, None, :] - cents[None, :, :]
    d = np.linalg.norm(diff, axis=2)

    # -- Confusable-pair priority ------------------------------------------
    # For colours that are close in Lab (e.g. a maroon dashed line and a navy
    # dashed line), the anti-aliased halo of the clearer colour can fall nearer
    # the duller colour's centroid and leak into the wrong mask. We subtract a
    # small per-colour bias from its distance so the clearer colour wins those
    # contested pixels -- equivalent to "select the clear colour first, exclude
    # it, then take the other", but done inside one exclusive assignment.
    d_true = d.copy()  # acceptance tests use unbiased true distance
    d_assign = d.copy()

    # -- Chroma-weighted distance for confusable chromatic colours ---------
    # Root cause of cross-leak between similar colours (e.g. a navy 1SC line and
    # a maroon 3SC line): a colour's anti-aliased edge is BRIGHT (high L) while
    # its centroid is DARK (low L). Plain Lab distance is dominated by that large
    # dL, so the bright edge of colour A lands nearer colour B's mid-L centroid
    # and leaks into B. The hue (a*,b*) of the edge still belongs to A, though.
    # So for chromatic pixels we DOWN-WEIGHT L when confusable pairs exist: the
    # match is driven by hue, which is stable across the bright->dark AA gradient.
    if priority_bias is not None and np.any(priority_bias):
        # priority_bias != 0 marks colours that participate in a confusable pair.
        confusable_k = np.asarray(priority_bias, dtype=np.float32) != 0
        chromatic_px = (px_s > NC_NEUTRAL_S_MAX) & (px_v > NC_NEUTRAL_DARK_V)
        if confusable_k.any() and chromatic_px.any():
            dL = np.abs(diff[:, :, 0])
            d_ab2 = np.linalg.norm(diff[:, :, 1:3], axis=2)
            d_chroma = d_ab2 + NC_CONFUSE_L_WEIGHT * dL   # (N,K)
            rows = np.where(chromatic_px)[0]
            cols = np.where(confusable_k)[0]
            # apply only for (chromatic pixel, confusable colour) entries
            d_assign[np.ix_(rows, cols)] = d_chroma[np.ix_(rows, cols)]
            d[np.ix_(rows, cols)]        = d_chroma[np.ix_(rows, cols)]
            # PRIORITY: subtract the (signed) bias from the ASSIGNMENT distance so
            # the clearer/vivid colour of a confusable pair (positive bias) wins
            # the contested anti-aliased edge pixels instead of the duller one.
            # e.g. a brightened RED edge (a* drifting toward orange) still gets
            # claimed by red, so red doesn't bleed into the orange mask. Only the
            # assignment is biased; acceptance tests below use the true distance d.
            pb = np.asarray(priority_bias, dtype=np.float32)
            d_assign[np.ix_(rows, cols)] -= pb[cols][None, :]

    # -- Achromatic-aware distance -----------------------------------------
    # A thin black/grey line is anti-aliased: it stays NEUTRAL (a*~b*~128) but
    # its lightness L rises toward white as the line thins. Plain Lab distance
    # then pushes those pixels far from the dark centroid (large dL), so the
    # thin parts of the curve drop out. For achromatic centroids we therefore
    # down-weight the L axis: distance is dominated by chroma (a*,b*), which
    # stays small for any neutral pixel. This lets faded thin black line pixels
    # reach the black centroid while still excluding coloured pixels (whose a*,b*
    # are far). Only pixels that are themselves near-neutral get this treatment.
    if achromatic.any():
        # chroma-only distance to each achromatic centroid
        ab_diff = diff[:, :, 1:3]                     # (N,K,2)
        d_ab = np.linalg.norm(ab_diff, axis=2)        # (N,K)
        # mild L term so we still prefer the right grey shade, but don't dominate
        dL = np.abs(diff[:, :, 0])
        d_neutral = d_ab + NC_ACHRO_L_WEIGHT * dL
        # pixel is eligible for neutral treatment if it is itself low-saturation
        # OR very dark: at low V, HSV saturation is numerically unstable, so a
        # near-black pixel (e.g. RGB 5,3,8) can report a high S even though it is
        # perceptually black. Such pixels belong with the achromatic centroid.
        neutral_px = (px_s <= NC_NEUTRAL_S_MAX) | (px_v <= NC_NEUTRAL_DARK_V)  # (N,)
        for k in range(K):
            if achromatic[k]:
                # use the neutral distance for neutral pixels toward this centroid
                d[neutral_px, k] = d_neutral[neutral_px, k]
                d_assign[neutral_px, k] = d_neutral[neutral_px, k]

    # Assignment uses biased distance; acceptance tests below use TRUE distance.
    nearest      = np.argmin(d_assign, axis=1)
    nearest_dist = d[np.arange(len(ys)), nearest]
    # 2nd-nearest for margin test (true distance)
    d_copy = d.copy()
    d_copy[np.arange(len(ys)), nearest] = np.inf
    second_dist = d_copy.min(axis=1)
    # Per-pixel acceptance
    md = np.array([max_dists[k] for k in nearest], dtype=np.float32)
    accept = (nearest_dist <= md) & (second_dist >= nearest_dist * NC_MARGIN)
    # For achromatic targets also cap brightness so we don't grab the white page
    nearest_is_achro = achromatic[nearest]
    accept &= (~nearest_is_achro | (px_v <= NC_ACHRO_V_MAX))
    # Saturation floor (chromatic colours only; s_mins[k] = -1 disables)
    if s_mins is not None:
        smin_arr = np.array([s_mins[k] for k in nearest], dtype=np.float32)
        # only apply S-floor where the nearest centroid is chromatic
        chroma_floor = smin_arr.copy()
        chroma_floor[nearest_is_achro] = -1.0
        accept &= (px_s >= chroma_floor)
    masks = []
    for k in range(K):
        m = np.zeros((H, W), dtype=bool)
        sel = accept & (nearest == k)
        m[ys[sel], xs[sel]] = True
        masks.append(m)
    return masks


def _saturation_floors(palette_lab):
    """Learn a per-colour saturation floor from the legend swatch S samples.

    Chromatic colours: floor = a fraction of the swatch's low-percentile S, so
    genuine (possibly faded) curve pixels pass but grey/halo pixels are cut.
    Achromatic colours: floor disabled (-1) -- they legitimately have low S.
    Returns a list aligned with palette_lab; -1 means 'no S gating'.
    """
    info = _LEGEND_SWATCH_INFO
    floors = [-1.0] * len(palette_lab)
    if not info or len(info) != len(palette_lab):
        return floors
    for k, lab in enumerate(palette_lab):
        if _is_achromatic(lab):
            continue  # don't gate grey/black by saturation
        s_samp = info[k].get('s_samples', None)
        if s_samp is None or len(s_samp) == 0:
            continue
        # Use a low percentile of the swatch S so we don't over-cut; then relax
        # by SAT_FLOOR_FRAC to tolerate fading/anti-aliasing on the curve.
        s_lo = float(np.percentile(s_samp, SAT_FLOOR_PCTILE))
        floors[k] = max(0.0, s_lo * SAT_FLOOR_FRAC)
    return floors


def _confusable_priority(palette_lab):
    """Flag confusable chromatic colour pairs and rank them, so the mask builder
    can resolve them by HUE rather than brightness.

    Two chromatic colours are 'confusable' when their centroids are close in Lab
    (dE < NC_CONFUSE_DIST) -- e.g. a navy dashed line and a maroon dashed line,
    both dark, differing mainly in hue. For such colours a plain Lab match fails
    on anti-aliased edges: an edge pixel is bright (high L) and lands nearer the
    other colour's mid-L centroid. The builder fixes this by down-weighting L for
    pixels of confusable colours (see _build_masks_nearest_centroid); this routine
    just identifies WHICH colours need that treatment.

    Returns a per-colour array; non-zero entries mark colours in a confusable
    pair. The sign/magnitude (vivid colour positive) is kept for diagnostics.
    """
    K = len(palette_lab)
    bias = np.zeros(K, dtype=np.float32)
    _conf_pairs = []
    if K < 2:
        return bias
    cents = np.stack(palette_lab).astype(np.float32)
    achro = [_is_achromatic(lab) for lab in palette_lab]
    for i in range(K):
        if achro[i]:
            continue
        for j in range(i+1, K):
            if achro[j]:
                continue
            de = float(np.linalg.norm(cents[i] - cents[j]))
            # Direction-aware EDGE-LEAK test, generalised:
            #  (1) the "edge" of a colour is taken from its OWN legend swatch
            #      pixels (the brightest, high-L samples = anti-aliased rim),
            #      not a fixed white-blend assumption; falls back to a 50% blend
            #      only when swatch samples are unavailable.
            #  (2) leak is only credible between colours of ADJACENT HUE (a
            #      brightened edge keeps its hue; blue can bleed to purple, not
            #      to orange). So we also require the hue angles to be close.
            def _hue(c):
                return np.degrees(np.arctan2(c[2] - 128, c[1] - 128))

            def _bright_edge(idx, cent):
                s = None
                _inf = _LEGEND_SWATCH_INFO
                if _inf and idx < len(_inf):
                    s = _inf[idx].get('lab_samples')
                if s is not None and len(s) >= 5:
                    s = np.asarray(s, np.float32)
                    thr = np.percentile(s[:, 0], 80)          # brightest 20%
                    hi = s[s[:, 0] >= thr]
                    if len(hi):
                        return hi.mean(0)
                # fallback: blend the centroid halfway to white in RGB
                rgb = cv2.cvtColor(np.uint8([[cent]]), cv2.COLOR_Lab2RGB)[0, 0].astype(np.float32)
                rgb = np.clip(rgb * 0.5 + 255 * 0.5, 0, 255).astype(np.uint8)
                return cv2.cvtColor(np.uint8([[rgb]]), cv2.COLOR_RGB2Lab)[0, 0].astype(np.float32)

            hue_gap = abs((_hue(cents[i]) - _hue(cents[j]) + 180) % 360 - 180)
            ei, ej = _bright_edge(i, cents[i]), _bright_edge(j, cents[j])
            leak = ((np.linalg.norm(ei - cents[j]) < np.linalg.norm(ei - cents[i])) or
                    (np.linalg.norm(ej - cents[i]) < np.linalg.norm(ej - cents[j])))
            leak = leak and (hue_gap < 40.0)          # hue must be adjacent
            if de >= NC_CONFUSE_DIST and not leak:
                continue
            de_eff = de if de < NC_CONFUSE_DIST else NC_CONFUSE_DIST * 0.6
            # chroma (vividness) of each: distance of (a*,b*) from neutral 128
            ci = float(np.hypot(cents[i][1]-128, cents[i][2]-128))
            cj = float(np.hypot(cents[j][1]-128, cents[j][2]-128))
            # the more vivid colour is clearer -> positive bias; the duller one
            # gets a matching negative bias. Use a gentle decay so pairs that are
            # confusable but not identical still receive a meaningful bias.
            strength = NC_CONFUSE_BIAS * (1.0 - (de_eff / NC_CONFUSE_DIST) ** 2)
            strength = max(strength, NC_CONFUSE_BIAS * 0.35)
            if ci >= cj:
                bias[i] += strength; bias[j] -= strength
                clear, dull = i, j
            else:
                bias[j] += strength; bias[i] -= strength
                clear, dull = j, i
            _conf_pairs.append((clear, dull))
            print(f"    confusable pair: color{clear+2:02d} (clear) vs "
                  f"color{dull+2:02d} (dull), dE={de:.1f} -> bias+/-{strength:.1f}")
    globals()['_LAST_CONFUSABLE_PAIRS'] = _conf_pairs
    return bias


def _verify_and_refine_with_legend(palette_lab, max_dists):
    """Use the known legend swatch regions to verify/refine each colour's
    acceptance radius. For colour k, classify the pixels inside EVERY legend
    swatch by nearest centroid. Colour k should win its OWN swatch and must not
    win another colour's swatch. If colour k intrudes on another swatch, shrink
    colour k's max_dist so its range stops bleeding into the wrong colour.

    IMPORTANT -- achromatic colours are EXCLUDED from this check (both as the
    intruder and as the victim swatch). Black/grey legitimately appears all over
    a legend (entry text, axis-label colour, outlines of open markers), so it
    will naturally "win" pixels inside other colours' swatches and vice-versa.
    Penalising that would wrongly shrink either black or its chromatic neighbour.
    Verification is therefore a CHROMATIC-vs-CHROMATIC test only.

    Returns refined max_dists. Purely legend-driven, no per-image constants.
    """
    info = _LEGEND_SWATCH_INFO
    if not info or len(info) != len(palette_lab):
        return max_dists  # no usable legend info; leave unchanged
    cents = np.stack(palette_lab).astype(np.float32)
    K = len(palette_lab)
    achromatic = [_is_achromatic(lab) for lab in palette_lab]
    refined = list(max_dists)
    for _pass in range(3):       # a few refinement passes
        changed = False
        for own in range(K):
            if achromatic[own]:
                continue                      # don't use grey/black swatch as victim
            samples = info[own].get('lab_samples')
            if samples is None or len(samples) == 0:
                continue
            d = np.linalg.norm(samples[:, None, :] - cents[None, :, :], axis=2)
            winner = np.argmin(d, axis=1)
            # Fraction of this swatch claimed by each colour
            for other in range(K):
                if other == own or achromatic[other]:
                    continue                  # black/grey intruding is expected -> ignore
                frac_other = float((winner == other).mean())
                if frac_other > 0.15:
                    # 'other' (chromatic) is intruding on 'own' (chromatic) swatch
                    # -> shrink 'other'. Floor relative to its own starting radius.
                    floor = max_dists[other] * 0.55
                    new_md = max(floor, refined[other] * NC_REFINE_SHRINK)
                    if new_md < refined[other] - 1e-3:
                        refined[other] = new_md
                        changed = True
        if not changed:
            break
    for k in range(K):
        if abs(refined[k] - max_dists[k]) > 1e-3:
            print(f"    legend-verify: color{k+2:02d} max_dist "
                  f"{max_dists[k]:.1f} -> {refined[k]:.1f}")
    return refined


def _lab_to_hsv_range(lab_color, tolerance=HUE_TOLERANCE):
    """Convert a Lab colour to HSV range(s) for mask extraction.
    Uses generous S/V tolerance to capture anti-aliased and faded pixels.
    """
    arr = np.array([[lab_color]], dtype=np.float32)
    arr_uint8 = np.clip(arr, 0, 255).astype(np.uint8)
    rgb = cv2.cvtColor(arr_uint8, cv2.COLOR_Lab2RGB)
    bgr = cv2.cvtColor(rgb, cv2.COLOR_RGB2BGR)
    hsv = cv2.cvtColor(bgr, cv2.COLOR_BGR2HSV)[0, 0].astype(int)
    h, s, v = int(hsv[0]), int(hsv[1]), int(hsv[2])
    # Achromatic (black/grey/white)
    if s < 50:
        # Adaptive V range: tight around actual value to avoid bleed into
        # background (V~255) or other grey shades.
        # For dark lines (v<80): allow bigger downward extension, smaller upward.
        # For mid-grey (v in 80-180): symmetric but capped to avoid white bg.
        # For light-grey (v>180): tighter to avoid background confusion.
        if v < 80:
            v_lo = max(0,   v - 40)
            v_hi = min(200, v + 60)   # cap well below white background
        elif v < 150:
            v_lo = max(0,   v - 50)
            v_hi = min(220, v + 50)
        else:
            v_lo = max(0,   v - 40)
            v_hi = min(235, v + 40)
        s_hi = min(255, s + 60)   # allow some saturation for anti-aliasing
        return [(0, 180, 0, s_hi, v_lo, v_hi)]
    # Chromatic -- use wide S/V tolerance to catch anti-aliased edges
    h_lo = (h - tolerance) % 180
    h_hi = (h + tolerance) % 180
    s_lo = max(0, s - 80)          # generous lower bound
    s_hi = 255                     # no upper bound
    v_lo = max(0, v - 100)
    v_hi = min(255, v + 100)
    if h_lo <= h_hi:
        return [(h_lo, h_hi, s_lo, s_hi, v_lo, v_hi)]
    # Red wraps around H=0/180
    return [(0, h_hi, s_lo, s_hi, v_lo, v_hi),
            (h_lo, 180, s_lo, s_hi, v_lo, v_hi)]

def _extract_colors_from_legend_img(legend_img_path):
    """Extract palette_lab from an external legend image.
    Finds foreground pixel groups (one per legend row) and returns
    their Lab means in top-to-bottom order.
    Handles both chromatic (coloured) and achromatic (grey/black) entries."""
    leg = cv2.imread(legend_img_path)
    if leg is None:
        print(f"  WARNING: cannot read legend image: {legend_img_path}")
        return []
    leg_rgb = cv2.cvtColor(leg, cv2.COLOR_BGR2RGB)
    leg_hsv = cv2.cvtColor(leg, cv2.COLOR_BGR2HSV)
    leg_lab = cv2.cvtColor(leg, cv2.COLOR_BGR2Lab)
    sat = leg_hsv[:, :, 1]
    val = leg_hsv[:, :, 2]
    Hh = leg.shape[0]

    # A legend marker is a compact blob (chromatic swatch OR a grey Placebo dot),
    # often threaded by a thin connecting line. Splitting on foreground gaps fails
    # because the line keeps foreground continuous. Instead, build a per-row
    # "marker strength" = count of pixels that are either clearly chromatic
    # (sat>40) OR mid-grey ink (low sat, mid val) -- the line is thin so it barely
    # registers, while a marker row peaks. Then split into bands at low valleys.
    chrom_row = ((sat > 40) & (val > 25)).sum(axis=1)
    grey_row  = ((sat <= 40) & (val > 40) & (val < 210)).sum(axis=1)
    strength  = np.maximum(chrom_row, grey_row).astype(float)
    thr = max(3.0, strength.max() * 0.35)
    on = strength >= thr
    # group consecutive on-rows into marker bands
    bands = []
    r = 0
    while r < Hh:
        if on[r]:
            r0 = r
            while r < Hh and on[r]:
                r += 1
            bands.append((r0, r - 1))
        else:
            r += 1
    if not bands:
        print("  WARNING: no marker rows in legend image")
        return []

    palette_lab = []
    for (r_min, r_max) in bands:
        patch_sat = sat[r_min:r_max+1, :]
        patch_val = val[r_min:r_max+1, :]
        patch_lab = leg_lab[r_min:r_max+1, :, :]
        patch_rgb = leg_rgb[r_min:r_max+1, :, :]
        chrom_mask = (patch_sat > 40) & (patch_val > 25)
        grey_mask  = (patch_sat <= 40) & (patch_val > 40) & (patch_val < 210)
        if chrom_mask.sum() >= 3:
            use_mask = chrom_mask
        elif grey_mask.sum() >= 3:
            use_mask = grey_mask
        else:
            continue
        lab_mean = patch_lab[use_mask].mean(axis=0).astype(np.float32)
        rgb_med  = np.median(patch_rgb[use_mask], axis=0).astype(int)
        print(f"  Legend row y={r_min}-{r_max}: RGB={tuple(rgb_med)}  n={int(use_mask.sum())}")
        palette_lab.append(lab_mean)
    return palette_lab


def _discover_colors():
    print("\n[Stage 2] Colour discovery ...")

    # If external legend image provided, use it directly
    if LEGEND_IMG_PATH is not None:
        print(f"  Using external legend image: {LEGEND_IMG_PATH}")
        palette_lab = _extract_colors_from_legend_img(LEGEND_IMG_PATH)
        if len(palette_lab) > 0:
            fg_mask = ~_pcm_background_mask(img_hsv)
            h_ch = img_hsv[:, :, 0]
            s_ch = img_hsv[:, :, 1]
            v_ch = img_hsv[:, :, 2]
            claimed = np.zeros((H, W), dtype=bool)
            colors = []
            for i, lab in enumerate(palette_lab):
                name = f'color{i+2:02d}'
                hsv_ranges = _lab_to_hsv_range(lab)
                raw = np.zeros((H, W), dtype=bool)
                for (h_lo, h_hi, s_lo, s_hi, v_lo, v_hi) in hsv_ranges:
                    band = (fg_mask
                            & (h_ch >= h_lo) & (h_ch <= h_hi)
                            & (s_ch >= s_lo) & (s_ch <= s_hi)
                            & (v_ch >= v_lo) & (v_ch <= v_hi))
                    raw |= band
                raw &= ~claimed
                claimed |= raw
                raw_mask = raw.astype(np.uint8)
                px_count = int(raw_mask.sum())
                rgb_mean = (tuple(img_rgb[raw].mean(axis=0).astype(int))
                            if px_count > 0 else (128, 128, 128))
                clean_mask = _clean_mask_advanced(raw_mask, is_achromatic=_is_achromatic(lab), mask_lab=lab)
                cd = {
                    'name':        name,
                    'mean_rgb':    rgb_mean,
                    'px_count':    px_count,
                    '_raw_mask':   raw,
                    '_clean_mask': clean_mask,
                }
                colors.append(cd)
                print(f"    {name}: raw={px_count}px  clean={clean_mask.sum()}px  RGB={rgb_mean}")
            return colors, None

    if USER_LEGEND_BOX is not None:
        legend_box = USER_LEGEND_BOX
        horiz_palette = None
        print(f"  Legend panel (user-provided): {legend_box}")
    else:
        legend_box = _pcm_detect_legend_panel()
        horiz_palette = None
        if legend_box is None:
            # Primary detector found nothing; try the auxiliary (small-marker) one.
            aux = _pcm_detect_legend_panel_aux()
            if aux is not None:
                if len(_pcm_sample_palette_from_legend(aux)) >= LEGEND_MIN_ENTRIES:
                    legend_box = aux
                    print("  Legend panel (aux small-marker detector)")
    global _LEGEND_BOX_FOR_CLEAN
    _LEGEND_BOX_FOR_CLEAN = legend_box
    if legend_box and USER_LEGEND_BOX is None:
        # A real legend must contain READABLE text. Tighten the box to the
        # swatch+text extent; if OCR finds no understandable words in the
        # candidate region, this isn't a legend and we discard it.
        x0b, y0b, x1b, y1b = legend_box
        sw_pts = [(x0b + 6, (y0b + y1b) // 2)]
        tb = _tighten_legend_with_ocr(legend_box, sw_pts)
        if tb is None and _HAS_OCR:
            print("  Legend candidate rejected (no readable text)")
            legend_box = None
            _LEGEND_BOX_FOR_CLEAN = None
        elif tb is not None:
            legend_box = tb
            _LEGEND_BOX_FOR_CLEAN = legend_box
        if legend_box:
            print(f"  Legend panel: x={legend_box[0]}-{legend_box[2]}, y={legend_box[1]}-{legend_box[3]}")

    fg_mask = ~_pcm_background_mask(img_hsv)
    h_ch = img_hsv[:, :, 0]
    s_ch = img_hsv[:, :, 1]
    v_ch = img_hsv[:, :, 2]

    palette_lab = _pcm_sample_palette_from_legend(legend_box)
    # If the column-based sampler didn't yield a usable legend, try the
    # ORIENTATION-AGNOSTIC unified detector: it finds text-labelled colour
    # swatches and the collinear/even/diverse group among them, so it reads a
    # legend whether it is laid out vertically or horizontally.
    # ALSO run it whenever the user drew a legend box: the row-band sampler
    # under-reads MULTI-COLUMN legends (it merges/drops columns), so it can stop
    # at exactly LEGEND_MIN_ENTRIES and never escalate. When a user box is given
    # we therefore always cross-check with the 2D detector and keep whichever
    # finds MORE colours, so a 2-column legend isn't truncated to 3.
    _want_unified = (len(palette_lab) < LEGEND_MIN_ENTRIES) or (USER_LEGEND_BOX is not None)
    if _want_unified:
        # If the user drew a legend box, confine the unified detector to it so it
        # cannot pick a different legend elsewhere (multi-panel figures), and keep
        # the user's box as authoritative.
        uni = _pcm_detect_legend_unified(restrict_box=USER_LEGEND_BOX)
        if uni is not None:
            _, colors, box = uni
            hp = []
            for (cx, cy, rgb) in colors:
                lab = cv2.cvtColor(np.array([[list(rgb)]], np.uint8),
                                   cv2.COLOR_RGB2Lab)[0, 0].astype(np.float32)
                hp.append(lab)
            # accept when it reaches the minimum AND improves on the sampler
            if len(hp) >= LEGEND_MIN_ENTRIES and len(hp) >= len(palette_lab):
                palette_lab = hp
                horiz_palette = hp           # use the direct-colour NC path
                if USER_LEGEND_BOX is not None:
                    legend_box = USER_LEGEND_BOX          # user box is authoritative
                else:
                    legend_box = box
                    # Tighten to swatch+readable-text extent (trim white margin).
                    sw_pts = [(int(cx), int(cy)) for (cx, cy, _rgb) in colors]
                    tb = _tighten_legend_with_ocr(box, sw_pts)
                    if tb is not None:
                        legend_box = tb
                globals()['_LEGEND_BOX_FOR_CLEAN'] = legend_box
                globals()['_LEGEND_SWATCH_INFO'] = [{'lab': p} for p in hp]
                print(f"  Legend panel (unified detector): {len(hp)} colours, "
                      f"box x={legend_box[0]}-{legend_box[2]}, "
                      f"y={legend_box[1]}-{legend_box[3]}")

    # Fallback for a USER-drawn box when the unified detector under-reads the
    # legend (thin "line-through" markers, hollow/open markers): find the swatch
    # markers directly as compact connected components in the swatch COLUMN (the
    # left part of the box, before the label text) and build the grid from them.
    # This reads a legend whose markers the chromatic/blob gates miss.
    if (USER_LEGEND_BOX is not None) and (len(palette_lab) < LEGEND_MIN_ENTRIES):
        mk = _markers_from_user_box(USER_LEGEND_BOX)
        if mk is not None and len(mk) >= LEGEND_MIN_ENTRIES:
            # mk = list of (cx, cy, (r,g,b)) top-to-bottom
            hp = []
            for (cx, cy, rgb) in mk:
                lab = cv2.cvtColor(np.array([[list(rgb)]], np.uint8),
                                   cv2.COLOR_RGB2Lab)[0, 0].astype(np.float32)
                hp.append(lab)
            palette_lab = hp
            horiz_palette = hp
            legend_box = USER_LEGEND_BOX
            globals()['_LEGEND_BOX_FOR_CLEAN'] = legend_box
            globals()['_LEGEND_SWATCH_INFO'] = [{'lab': p} for p in hp]
            # populate the grid-cell + achromatic structures so the unified TABLE
            # builds (chromatic cells + achromatic swatches separated by chroma).
            _chrom_cells = [(cx, cy, rgb) for (cx, cy, rgb) in mk
                            if (max(rgb) - min(rgb)) > 24]
            _achro_sw = [(cx, cy, rgb) for (cx, cy, rgb) in mk
                         if (max(rgb) - min(rgb)) <= 24]
            def _cl1d(vals, tol):
                if not vals:
                    return []
                vs = sorted(vals); g = [[vs[0]]]
                for x in vs[1:]:
                    (g[-1].append(x) if x - g[-1][-1] <= tol else g.append([x]))
                return [float(np.mean(z)) for z in g]
            _gx = _cl1d([c[0] for c in mk], 18)
            _gy = _cl1d([c[1] for c in mk], 7)
            globals()['_LAST_LEGEND_GRID'] = {
                'cols': [round(c) for c in _gx], 'rows': [round(r) for r in _gy],
                'cells': [(round(c[0]), round(c[1]), tuple(int(z) for z in c[2]))
                          for c in _chrom_cells]}
            globals()['_LAST_ACHRO_SWATCHES'] = [
                (round(c[0]), round(c[1]), tuple(int(z) for z in c[2])) for c in _achro_sw]
            print(f"  Legend panel (user-box marker fallback): {len(mk)} swatches "
                  f"({len(_chrom_cells)} chromatic + {len(_achro_sw)} achromatic)")

    if len(palette_lab) < LEGEND_MIN_ENTRIES:
        print("  Fallback: hue-histogram colour discovery")
        global _LEGEND_SWATCH_INFO
        _LEGEND_SWATCH_INFO = []   # no reliable legend swatches in fallback
        # Keep any legend colours found so far, but remember how many so we can
        # de-duplicate them against the hue-derived palette afterwards (a legend
        # colour and its hue twin would otherwise split one curve into two).
        _n_legend_prefix = len(palette_lab)
        chrom_mask = fg_mask & (s_ch > 40)
        if legend_box:
            x0l, y0l, x1l, y1l = legend_box
            chrom_mask[y0l:y1l+1, x0l:x1l+1] = False
        hue_vals = h_ch[chrom_mask].astype(float)
        if len(hue_vals) >= CHROM_MIN_PX:
            # Build circular 360-degree histogram (H repeated) to handle red wrap-around
            hist180 = np.zeros(180, dtype=float)
            for hv in hue_vals:
                hist180[int(hv) % 180] += 1
            hist360 = np.concatenate([hist180, hist180])
            smooth360 = gaussian_filter1d(hist360, sigma=HUE_SIGMA)
            peaks360, _ = find_peaks(smooth360, distance=HUE_MIN_DIST,
                                      prominence=smooth360.max() * 0.05)
            # Map peaks back to 0-179 and deduplicate
            seen_hues = []
            for ph360 in peaks360:
                ph = int(ph360) % 180
                if any(abs(ph - sh) < HUE_MIN_DIST for sh in seen_hues): continue
                seen_hues.append(ph)
                # For red (H near 0 or near 179), merge both ends
                if ph <= HUE_TOLERANCE:
                    near = chrom_mask & ((h_ch <= ph + HUE_TOLERANCE) | (h_ch >= 180 - HUE_TOLERANCE))
                elif ph >= 180 - HUE_TOLERANCE:
                    near = chrom_mask & ((h_ch >= ph - HUE_TOLERANCE) | (h_ch <= HUE_TOLERANCE))
                else:
                    near = chrom_mask & (np.abs(h_ch.astype(int) - ph) <= HUE_TOLERANCE)
                if near.sum() < CHROM_MIN_PX: continue
                lab_mean = img_lab[near].mean(axis=0).astype(np.float32)
                palette_lab.append(lab_mean)
        grey_mask = (fg_mask & (s_ch <= GREY_SAT_MAX) &
                     (v_ch >= GREY_VAL_MIN) & (v_ch <= GREY_VAL_MAX))
        # Exclude pixels that already belong to a chromatic hue so grey
        # detection doesn't absorb anti-aliased edges of coloured lines.
        grey_mask &= ~chrom_mask
        if grey_mask.sum() >= GREY_MIN_PX:
            # Cluster grey pixels by brightness (V) to distinguish dark lines
            # (e.g. black curve, V~30) from mid-grey lines (V~150).
            # Without clustering, averaging them produces a single wrong V
            # that matches neither, causing both to be missed or blurred together.
            v_vals = v_ch[grey_mask].astype(float)
            v_hist = np.zeros(256, dtype=float)
            for vv in v_vals.astype(int):
                v_hist[vv] += 1
            v_smooth = gaussian_filter1d(v_hist, sigma=8)
            v_peaks, _ = find_peaks(v_smooth, prominence=v_smooth.max() * 0.15,
                                    distance=GREY_PEAK_DIST_EFF)
            clusters_added = 0
            dark_clusters = []
            for pk_v in v_peaks:
                # Use a wider half-window so anti-aliased and slightly
                # brighter pixels of the same line are captured.
                half = max(GREY_CLUSTER_SEP, 30)
                cluster_m = (grey_mask &
                             (v_ch >= max(0,   int(pk_v) - half)) &
                             (v_ch <= min(255, int(pk_v) + half)))
                if cluster_m.sum() >= GREY_MIN_PX // 2:
                    # A near-black curve is rendered with a dark marker core
                    # (low V) plus an anti-aliased thin line that drifts brighter.
                    # If we split those into separate V clusters the curve breaks
                    # apart. Collect all sufficiently-dark clusters and fuse them
                    # into ONE black entry; keep brighter (true grey) clusters
                    # separate.
                    if int(pk_v) <= GREY_BLACK_VMAX:
                        dark_clusters.append(cluster_m)
                    else:
                        lab_mean = img_lab[cluster_m].mean(axis=0).astype(np.float32)
                        palette_lab.append(lab_mean)
                        clusters_added += 1
                        print(f"    grey cluster V~{pk_v}: {cluster_m.sum()} px")
            if dark_clusters:
                black_m = np.zeros((H, W), dtype=bool)
                for cm in dark_clusters:
                    black_m |= cm
                # Use the darkest pixels as the representative (ink-core), so the
                # centroid is genuinely black rather than a grey average.
                bv = v_ch[black_m]
                dark_core = black_m & (v_ch <= max(GREY_VAL_MIN + 10,
                                                   int(np.percentile(bv, 40))))
                src = dark_core if dark_core.sum() >= GREY_MIN_PX // 2 else black_m
                lab_black = img_lab[src].mean(axis=0).astype(np.float32)
                palette_lab.append(lab_black)
                clusters_added += 1
                print(f"    black (fused {len(dark_clusters)} dark clusters): "
                      f"{int(black_m.sum())} px  core={int(dark_core.sum())}")
            if clusters_added == 0:
                # No distinct clusters found -- use single mean
                lab_mean = img_lab[grey_mask].mean(axis=0).astype(np.float32)
                palette_lab.append(lab_mean)
                print(f"    grey (single): {grey_mask.sum()} px")

    if len(palette_lab) == 0:
        print("  WARNING: no colours discovered")
        return [], legend_box

    # De-duplicate near-identical palette colours -- FALLBACK PATH ONLY. In the
    # fallback a legend colour and its hue-histogram twin can both be present and
    # would split one curve into two masks. The legend (NC) path is trusted and
    # must not be merged here. We detect fallback by the empty swatch stash.
    _is_fallback = (len(_LEGEND_SWATCH_INFO) == 0)
    if _is_fallback and len(palette_lab) > 1:
        def _hue_of(lab_c):
            rgb = cv2.cvtColor(np.clip(np.array([[lab_c]], np.float32), 0, 255)
                               .astype(np.uint8), cv2.COLOR_Lab2RGB)
            hsv = cv2.cvtColor(rgb, cv2.COLOR_RGB2HSV)[0, 0]
            return int(hsv[0]), int(hsv[1])   # (hue, sat)
        deduped = []
        for lab in palette_lab:
            h_new, s_new = _hue_of(lab)
            dup = False
            for kept in deduped:
                if np.linalg.norm(lab - kept) < DEDUP_LAB:
                    dup = True; break
                h_k, s_k = _hue_of(kept)
                # same-hue chromatic twins (both reasonably saturated)
                if (s_new > 40 and s_k > 40 and
                        min(abs(h_new - h_k), 180 - abs(h_new - h_k)) <= DEDUP_HUE):
                    dup = True; break
            if not dup:
                deduped.append(lab)
        if len(deduped) < len(palette_lab):
            print(f"  De-duplicated palette: {len(palette_lab)} -> {len(deduped)} colours")
        palette_lab = deduped

    use_nc = ((legend_box is not None)
              and (len(_LEGEND_SWATCH_INFO) == len(palette_lab))
              and (len(palette_lab) >= NC_MIN_LEGEND_ENTRIES))

    if legend_box is not None and len(_LEGEND_SWATCH_INFO) > 0:
        # Enforce/declare the legend<->mask 1:1 contract: every legend swatch
        # produces exactly one colour mask and vice-versa.
        n_leg = len(_LEGEND_SWATCH_INFO)
        _is_1to1 = (n_leg == len(palette_lab))
        globals()['_LEGEND_1TO1'] = bool(_is_1to1) and n_leg >= 2
        print(f"  Legend<->mask matching: {n_leg} legend swatch(es) -> "
              f"{len(palette_lab)} colour(s)"
              + ("  [1:1 OK]" if _is_1to1
                 else f"  [MISMATCH -- using {len(palette_lab)}]"))

    if use_nc:
        # -- Mutually-exclusive masks via nearest-centroid assignment ------
        # When a reliable legend is available, assign each foreground pixel to
        # its single nearest palette colour in Lab space. Overlap between colour
        # ranges is impossible by construction, and the legend swatches let us
        # verify/refine each colour's acceptance radius.
        max_dists = _adaptive_radii(palette_lab)
        max_dists = _verify_and_refine_with_legend(palette_lab, max_dists)
        s_mins    = _saturation_floors(palette_lab)
        prio_bias = _confusable_priority(palette_lab)
        for k, sm in enumerate(s_mins):
            if sm >= 0:
                print(f"    sat-floor: color{k+2:02d} S>={sm:.0f}")
        nc_masks  = _build_masks_nearest_centroid(palette_lab, fg_mask, max_dists,
                                                  s_mins, priority_bias=prio_bias)
        colors = []
        for i, (lab, raw) in enumerate(zip(palette_lab, nc_masks)):
            name       = f'color{i+2:02d}'
            # -- Plot-area clipping (ALL colours) -----------------------------
            # Once the x/y axes give us the plotting rectangle, every pixel
            # outside it is noise: legend swatches, the title, axis tick labels,
            # the "LLOQ" annotation, etc. Remove them from EVERY colour mask
            # before pixel classification so the walk only sees real curve data.
            # (When no axes were found, PLOT_MASK is all-True -> no-op.)
            #
            # Black/grey needs the most care because axis lines, tick text and
            # annotations are themselves black: clipping to the plot rectangle is
            # exactly what separates a real dark curve from that surrounding text.
            raw = raw & PLOT_MASK
            raw_mask   = raw.astype(np.uint8)
            px_count   = int(raw_mask.sum())
            rgb_mean   = (tuple(img_rgb[raw].mean(axis=0).astype(int))
                          if px_count > 0 else (128, 128, 128))
            clean_mask = _clean_mask_advanced(raw_mask, is_achromatic=_is_achromatic(lab), mask_lab=lab)
            colors.append({'name': name, 'mean_rgb': rgb_mean, 'px_count': px_count,
                           '_raw_mask': raw, '_clean_mask': clean_mask})
            tag = 'achrom' if _is_achromatic(lab) else 'chrom'
            print(f"    {name}: raw={px_count}px  clean={clean_mask.sum()}px  RGB={rgb_mean}  [nearest-centroid,{tag}]")
        return colors, legend_box

    # -- Fallback / no-legend: independent HSV-box ranges with greedy claim --
    # Exception: achromatic (black/grey) palette entries use the same
    # achromatic-aware nearest-centroid logic as the legend path, so thin
    # anti-aliased black lines are recovered here too.
    claimed = np.zeros((H, W), dtype=bool)
    colors  = []
    # Pre-compute achromatic masks (mutually exclusive among the achromatic set)
    achro_idx = [i for i, lab in enumerate(palette_lab) if _is_achromatic(lab)]
    achro_masks = {}
    if achro_idx:
        achro_palette = [palette_lab[i] for i in achro_idx]
        masks_a = _build_masks_nearest_centroid(achro_palette, fg_mask)
        for j, i in enumerate(achro_idx):
            # Restrict grey/black masks to the plotting area so axis labels,
            # title and legend text (all grey/black, all outside the axes) are
            # excluded and a real grey curve is cleanly isolated.
            achro_masks[i] = masks_a[j] & PLOT_MASK
    for i, lab in enumerate(palette_lab):
        name       = f'color{i+2:02d}'
        if i in achro_masks:
            raw = achro_masks[i] & ~claimed
        else:
            hsv_ranges = _lab_to_hsv_range(lab)
            raw = np.zeros((H, W), dtype=bool)
            for (h_lo, h_hi, s_lo, s_hi, v_lo, v_hi) in hsv_ranges:
                band = (fg_mask
                        & (h_ch >= h_lo) & (h_ch <= h_hi)
                        & (s_ch >= s_lo) & (s_ch <= s_hi)
                        & (v_ch >= v_lo) & (v_ch <= v_hi))
                raw |= band
            raw &= ~claimed
            # Plot-area clipping for chromatic masks too: drop legend swatches,
            # title and any coloured annotation outside the plotting rectangle
            # before classification. (No-op when axes weren't found.)
            raw = raw & PLOT_MASK
        claimed |= raw
        raw_mask   = raw.astype(np.uint8)
        px_count   = int(raw_mask.sum())
        rgb_mean   = (tuple(img_rgb[raw].mean(axis=0).astype(int))
                      if px_count > 0 else (128, 128, 128))
        clean_mask = _clean_mask_advanced(raw_mask, is_achromatic=_is_achromatic(lab), mask_lab=lab)
        colors.append({'name': name, 'mean_rgb': rgb_mean, 'px_count': px_count,
                       '_raw_mask': raw, '_clean_mask': clean_mask})
        tag = ' [achro]' if i in achro_masks else ''
        print(f"    {name}: raw={px_count}px  clean={clean_mask.sum()}px  RGB={rgb_mean}{tag}")
    return colors, legend_box

# -- Step 4: Mask cleaning ------------------------------------------------------
def _clean_mask_advanced(mask, is_achromatic=False, mask_lab=None):
    """
    Remove text labels, dashed-line artefacts, and axis-adjacent noise.

    Noise elements (x/y axes, tick/title text, the LLOQ line) are drawn in one
    ink colour, measured once as NOISE_LAB (usually near-black). We only strip
    that noise from the COLOUR MASK whose colour matches NOISE_LAB -- i.e. the
    black mask when the noise is black. Other colour masks are left intact, so
    we never carve axis-shaped holes out of a coloured curve. `mask_lab` is the
    palette colour of this mask; if it is close to NOISE_LAB (or the mask is
    achromatic and the noise is achromatic) we run the shape-based removal.

    Pass 0: shape-based removal of axis lines / legend glyphs / LLOQ dashes
    Pass 1: Horizontal dashed-line row removal
    Pass 2: Text / corner label removal (small clustered components)
    Pass 3: Axis-adjacent tick / noise removal
    """
    out = mask.copy().astype(np.uint8)
    if out.sum() == 0:
        return out

    # Decide whether THIS mask carries structural noise (axis rules, gridlines,
    # legend text, error-bar stems). Any ACHROMATIC mask does: grey/black data
    # curves live in the same neutral range as the axes and grey error bars, so
    # they collect that noise and must be cleaned. For chromatic masks we gate on
    # distance to the noise colour so real coloured data is never touched.
    noise_here = is_achromatic
    if not is_achromatic and mask_lab is not None:
        try:
            d_noise = float(np.linalg.norm(np.asarray(mask_lab, np.float32) - NOISE_LAB))
            noise_here = d_noise <= NOISE_MATCH_DIST
        except Exception:
            noise_here = False

    # Pass 0: achromatic shape cleanup -- remove ONLY noise components (axis
    # rules, legend text, reference-line dashes) by their shape, leaving the
    # curve fully intact. We never blank whole rows/bands (that would cut the
    # curve where it crosses them); we operate per connected-component.
    if noise_here:
        def _drop_axis_components(mask_u8):
            """Remove components that are long, thin, straight rules (the x/y
            axes and grid lines). A curve is never a single near-perfect line."""
            nlb, lab, st, _ = cv2.connectedComponentsWithStats(mask_u8, 8)
            for i in range(1, nlb):
                a = st[i, cv2.CC_STAT_AREA]
                w = st[i, cv2.CC_STAT_WIDTH]; h = st[i, cv2.CC_STAT_HEIGHT]
                long_side = max(w, h); short_side = max(1, min(w, h))
                aspect = long_side / short_side
                fill = a / float(max(1, w * h))
                # axis/grid rule: extremely elongated, thin, and spans most of
                # the plot in one direction
                spans = ((w >= (PLOT_AREA[2]-PLOT_AREA[0]) * 0.6) if PLOT_AREA else (w >= W*0.6)) or \
                        ((h >= (PLOT_AREA[3]-PLOT_AREA[1]) * 0.6) if PLOT_AREA else (h >= H*0.6))
                if aspect >= 20 and short_side <= 6 and spans:
                    mask_u8[lab == i] = 0
            return mask_u8

        out = _drop_axis_components(out)
        if out.sum() == 0:
            return out

        # Remove the legend text/swatch glyphs. These are many small components
        # confined to the legend box. We delete components whose centre lies in
        # the legend box (so a curve passing nearby is untouched -- only glyphs
        # fully inside the box go). Whole-box blanking is avoided.
        lb = globals().get('_LEGEND_BOX_FOR_CLEAN', None)
        if lb is not None:
            lx0, ly0, lx1, ly1 = lb
            box_area_frac = ((lx1 - lx0 + 1) * (ly1 - ly0 + 1)) / float(H * W)
            if box_area_frac < 0.35:
                nlb, lab, st, cen = cv2.connectedComponentsWithStats(out, 8)
                for i in range(1, nlb):
                    cx, cy = cen[i]
                    x = st[i, cv2.CC_STAT_LEFT]; y = st[i, cv2.CC_STAT_TOP]
                    w = st[i, cv2.CC_STAT_WIDTH]; h = st[i, cv2.CC_STAT_HEIGHT]
                    # component lies (mostly) inside the legend box
                    inside = (x >= lx0-2 and y >= ly0-2 and
                              x+w <= lx1+2 and y+h <= ly1+2)
                    if inside:
                        out[lab == i] = 0
                if out.sum() == 0:
                    return out

        # Remove a horizontal reference line (e.g. "LLOQ = ...") and its inline
        # text -- but only the DASH and GLYPH components, found by shape, not by
        # wiping the row. The dashes are small, wide-flat fragments aligned at a
        # single y; the label glyphs are small blobs at the same y. We identify
        # the reference y as the row with many small flat fragments spanning a
        # wide x-range, then drop only the small components centred near it.
        if PLOT_AREA is not None:
            nlb, lab, st, cen = cv2.connectedComponentsWithStats(out, 8)
            # candidate dash components: short, wide-ish, thin fragments
            dash = []
            for i in range(1, nlb):
                a = st[i, cv2.CC_STAT_AREA]
                w = st[i, cv2.CC_STAT_WIDTH]; h = st[i, cv2.CC_STAT_HEIGHT]
                if a <= 80 and h <= 6 and 2 <= w <= 40:
                    dash.append((cen[i][1], cen[i][0], i))   # (y, x, idx)
            if len(dash) >= 8:
                dys = np.array([d[0] for d in dash])
                # find a tight y-cluster of dashes spanning a wide x-range
                ymed = np.median(dys)
                near = [d for d in dash if abs(d[0] - ymed) <= 6]
                if len(near) >= 8:
                    xsd = sorted(d[1] for d in near)
                    xspan = xsd[-1] - xsd[0]
                    plot_w = PLOT_AREA[2] - PLOT_AREA[0]
                    if xspan >= plot_w * 0.5:
                        # this is the reference line. Drop all SMALL components
                        # (dashes + label glyphs) whose centre is within a thin
                        # y-band around it -- never large curve components.
                        ylo, yhi = ymed - 16, ymed + 16
                        for i in range(1, nlb):
                            a = st[i, cv2.CC_STAT_AREA]
                            if a > 400:           # protect curve/marker blobs
                                continue
                            if ylo <= cen[i][1] <= yhi:
                                out[lab == i] = 0

        out = _split_achromatic_by_shape(out.astype(bool)).astype(np.uint8)
        if out.sum() == 0:
            return out

    # Pass 1: Dashed-line row removal (dashed-aware)
    # A true horizontal dashed line has: multiple short segments of similar length,
    # similar gaps between them, and spans a significant portion of image width.
    def _row_segments(xs):
        """Split sorted x-coords into contiguous segments (gap>2 = new segment)."""
        segs = []
        if len(xs) == 0:
            return segs
        s, e = xs[0], xs[0]
        for x in xs[1:]:
            if x - e <= 2:
                e = x
            else:
                segs.append(e - s + 1)
                s, e = x, x
        segs.append(e - s + 1)
        return segs

    # The row-wipe below removes a horizontal DASHED reference line (drawn in the
    # axis/noise ink). Restrict to the noise-coloured mask: on a chromatic data
    # mask, a flat stretch of markers at a similar y reads as a dashed line and
    # would be erased as a horizontal white streak through the curve.
    if noise_here:
      for y in range(H):
          row = out[y, :]
          if row.sum() == 0:
              continue
          xs = np.where(row > 0)[0]
          x_span = int(xs.max()) - int(xs.min()) + 1
          if x_span < W * 0.30:
              continue  # doesn't span enough of the image
          seg_lens = np.array(_row_segments(xs), dtype=float)
          if len(seg_lens) < 5:
              continue  # need at least 5 dash segments to be classified as dashed line
          if seg_lens.max() > 30:
              continue  # individual segments too long (not a short dash)
          if seg_lens.min() < 2:
              continue  # segments too thin
          # Segment lengths must be similar (low CV)
          if seg_lens.mean() > 0 and seg_lens.std() / seg_lens.mean() > 0.6:
              continue
          # Gaps between segments must be similar (low CV)
          seg_starts = []
          seg_ends = []
          s, e = xs[0], xs[0]
          for x in xs[1:]:
              if x - e <= 2:
                  e = x
              else:
                  seg_ends.append(e)
                  seg_starts.append(x)
                  s, e = x, x
          seg_ends.append(e)
          if len(seg_starts) >= 2:
              gaps = np.array([seg_starts[i] - seg_ends[i-1]
                               for i in range(1, len(seg_starts))], dtype=float)
              if gaps.mean() > 0 and gaps.std() / gaps.mean() > 0.6:
                  continue
          out[y, :] = 0

    # Pass 2: Text component removal -- ONLY on the noise-coloured (black/grey)
    # mask. Data markers are ALSO small, clustered components, so running this on
    # a chromatic curve mask deletes the data points. Text is drawn in the noise
    # ink colour, so restricting to noise_here removes labels without harming data.
    if noise_here:
      # Identify components that look like text characters:
    # - small area (< CHAR_MAX_AREA)
    # - aspect ratio close to a character (not too elongated vertically)
    # - clustered near other similar components (text words have multiple chars)
      nl, lbl, st, _ = cv2.connectedComponentsWithStats(out, connectivity=8)
      small_comps = []
      largest_area = 0
      largest_idx  = -1
      for i in range(1, nl):
          area = int(st[i, cv2.CC_STAT_AREA])
          if area > largest_area:
              largest_area = area
              largest_idx  = i
          if area < CHAR_MAX_AREA:
              w_i = int(st[i, cv2.CC_STAT_WIDTH])
              h_i = int(st[i, cv2.CC_STAT_HEIGHT])
              cx  = st[i, cv2.CC_STAT_LEFT] + w_i / 2
              cy  = st[i, cv2.CC_STAT_TOP]  + h_i / 2
              # Text characters have aspect ratio roughly between 0.3 and 3.0
              # (not extremely thin vertical lines like tick marks)
              aspect = w_i / max(h_i, 1)
              if 0.2 <= aspect <= 4.0:
                  small_comps.append((i, cx, cy, area))
      # Compute x_max and y_center of largest component (main curve body)
      if largest_idx >= 0:
          lg_left   = int(st[largest_idx, cv2.CC_STAT_LEFT])
          lg_right  = lg_left + int(st[largest_idx, cv2.CC_STAT_WIDTH])
          lg_top    = int(st[largest_idx, cv2.CC_STAT_TOP])
          lg_bottom = lg_top + int(st[largest_idx, cv2.CC_STAT_HEIGHT])
          lg_cy     = (lg_top + lg_bottom) / 2.0
      else:
          lg_right, lg_cy = W, H / 2
          lg_left = 0
      if len(small_comps) >= 2:
          sc_arr = np.array([[cx, cy] for _, cx, cy, _ in small_comps], dtype=float)
          diffs  = sc_arr[:, np.newaxis, :] - sc_arr[np.newaxis, :, :]
          dists  = np.sqrt((diffs**2).sum(axis=2))
          np.fill_diagonal(dists, np.inf)
          neighbour_count = (dists < TEXT_CLUSTER_RADIUS).sum(axis=1)
          for k, (i, cx_i, cy_i, area) in enumerate(small_comps):
              if neighbour_count[k] >= 2 and area < largest_area:
                  # Don't remove if this component is a curve continuation:
                  # - its x is beyond the largest component's right edge, AND
                  # - its y is close to the largest component's y-center
                  is_curve_continuation = (
                      cx_i > lg_right and
                      abs(cy_i - lg_cy) <= (lg_bottom - lg_top) / 2 + 10
                  )
                  if is_curve_continuation:
                      continue
                  out[lbl == i] = 0


    # Pass 3: Axis-adjacent tick / noise removal -- ONLY on the noise-coloured
    # (black/grey) mask. Tick marks and axis fragments are drawn in the axis ink
    # colour, never in a data-curve colour, so a CHROMATIC pixel near the axis is
    # real data and must be kept. We still require the thin-stroke test so that
    # even on the noise mask a data blob isn't removed.
    if noise_here:
      nl2, lbl2, st2, _ = cv2.connectedComponentsWithStats(out, connectivity=8)
      for i in range(1, nl2):
        area = int(st2[i, cv2.CC_STAT_AREA])
        if area >= TICK_AREA_MAX:
            continue
        w_i = int(st2[i, cv2.CC_STAT_WIDTH]); h_i = int(st2[i, cv2.CC_STAT_HEIGHT])
        cx_i = st2[i, cv2.CC_STAT_LEFT] + w_i / 2
        cy_i = st2[i, cv2.CC_STAT_TOP]  + h_i / 2
        near = (any(abs(cy_i - ay) <= TICK_MARGIN for ay in AXIS_ROWS) or
                any(abs(cx_i - ax) <= TICK_MARGIN for ax in AXIS_COLS))
        thin = min(w_i, h_i) <= 3          # a tick/axis stroke is thin
        if near and thin:
            out[lbl2 == i] = 0

    return out

# -- Step 5: X-distribution mode-based pixel classification --------------------
def classify_by_xdist(mask):
    """
    Classify each pixel in `mask` as curve, blob (data point/tcap), or stem.

    Strategy:
      1. Compute per-column pixel count histogram.
      2. Find modes (spikes) -- these are blob/stem columns.
      3. For each column:
         - Non-mode column: all pixels -> curve
         - Mode column: split into y-clusters; tall narrow -> stem, compact -> blob
    """
    col_counts = mask.sum(axis=0).astype(float)
    if col_counts.max() == 0:
        z = np.zeros((H, W), dtype=np.uint8)
        return z, z, z, []

    smooth = gaussian_filter1d(col_counts, sigma=XDIST_SIGMA)
    min_prom = smooth.max() * XDIST_MODE_PROM
    mode_xs_arr, _ = find_peaks(smooth, distance=XDIST_MODE_DIST,
                                 prominence=min_prom)
    mode_xs = mode_xs_arr.tolist()

    mode_set = set(mode_xs)
    non_mode_counts = [col_counts[x] for x in range(W)
                       if col_counts[x] > 0 and x not in mode_set]
    curve_baseline = float(np.median(non_mode_counts)) if non_mode_counts else 1.0
    blob_thresh    = max(curve_baseline * 2.0, curve_baseline + 4)

    curve_mask = np.zeros((H, W), dtype=np.uint8)
    blob_mask  = np.zeros((H, W), dtype=np.uint8)
    stem_mask  = np.zeros((H, W), dtype=np.uint8)

    for x in range(W):
        col_px = np.where(mask[:, x] > 0)[0]
        if len(col_px) == 0:
            continue
        if col_counts[x] < blob_thresh:
            curve_mask[col_px, x] = 1
            continue
        # Blob column: split into contiguous y-clusters
        gaps     = np.where(np.diff(col_px) > 3)[0]
        clusters = np.split(col_px, gaps + 1)
        for cl in clusters:
            if len(cl) < BLOB_MIN_AREA:
                curve_mask[cl, x] = 1
                continue
            h_cl  = int(cl.max()) - int(cl.min()) + 1
            ratio = h_cl / 1.0  # single column width = 1
            if ratio > BLOB_HEIGHT_RATIO:
                stem_mask[cl, x] = 1
            else:
                blob_mask[cl, x] = 1

    return curve_mask, blob_mask, stem_mask, mode_xs

# -- Disk precomputation --------------------------------------------------------
_dy_arr = np.arange(-R, R + 1)
_dx_arr = np.arange(-R, R + 1)
_DX, _DY = np.meshgrid(_dx_arr, _dy_arr)
DISK = (_DX**2 + _DY**2 <= R**2)

def _get_disk_pixels(xc, yc, img_h, img_w):
    y0 = max(0, yc - R); y1 = min(img_h, yc + R + 1)
    x0 = max(0, xc - R); x1 = min(img_w, xc + R + 1)
    patch = DISK[y0-yc+R:y1-yc+R, x0-xc+R:x1-xc+R]
    yr, xr = np.where(patch)
    return (yr + y0).astype(np.int32), (xr + x0).astype(np.int32)

def _apply_capsule_coverage(x0, y0, x1, y1, covered, img_h, img_w):
    dist = max(abs(x1-x0), abs(y1-y0))
    if dist == 0:
        ys_d, xs_d = _get_disk_pixels(x0, y0, img_h, img_w)
        covered[ys_d, xs_d] = 1
        return
    ts = np.linspace(0, 1, dist*2+1)
    xs_p = np.round(x0 + ts*(x1-x0)).astype(int)
    ys_p = np.round(y0 + ts*(y1-y0)).astype(int)
    seen = set()
    for xc, yc in zip(xs_p.tolist(), ys_p.tolist()):
        xc = int(np.clip(xc, 0, img_w-1)); yc = int(np.clip(yc, 0, img_h-1))
        if (xc, yc) in seen: continue
        seen.add((xc, yc))
        ys_d, xs_d = _get_disk_pixels(xc, yc, img_h, img_w)
        covered[ys_d, xs_d] = 1

def _new_capsule_cov(cx, cy, xc, yc, covered, wm, img_h, img_w):
    tmp = covered.copy()
    _apply_capsule_coverage(cx, cy, xc, yc, tmp, img_h, img_w)
    return int(np.sum((wm == 1) & (tmp == 1) & (covered == 0)))

def _segment_in_mask(x0, y0, x1, y1, dm, img_h, img_w):
    dist = max(abs(x1-x0), abs(y1-y0))
    if dist == 0: return bool(dm[y0, x0])
    ts = np.linspace(0, 1, dist*2+1)
    xs = np.clip(np.round(x0 + ts*(x1-x0)).astype(int), 0, img_w-1)
    ys = np.clip(np.round(y0 + ts*(y1-y0)).astype(int), 0, img_h-1)
    return bool(np.all(dm[ys, xs] > 0))

def _segment_in_mask_relaxed(x0, y0, x1, y1, dm, img_h, img_w, end_frac=0.10):
    dist = max(abs(x1-x0), abs(y1-y0))
    if dist == 0: return bool(dm[y0, x0])
    ts_s = np.linspace(0, end_frac, max(3, int(dist*end_frac*2+1)))
    ts_e = np.linspace(1-end_frac, 1, max(3, int(dist*end_frac*2+1)))
    xs_s = np.clip(np.round(x0 + ts_s*(x1-x0)).astype(int), 0, img_w-1)
    ys_s = np.clip(np.round(y0 + ts_s*(y1-y0)).astype(int), 0, img_h-1)
    xs_e = np.clip(np.round(x0 + ts_e*(x1-x0)).astype(int), 0, img_w-1)
    ys_e = np.clip(np.round(y0 + ts_e*(y1-y0)).astype(int), 0, img_h-1)
    return (bool(np.any(dm[ys_s, xs_s] > 0)) and
            bool(np.any(dm[ys_e, xs_e] > 0)))

def _is_rightward_75(cx, cy, xc, yc):
    dx = xc - cx; dy = yc - cy
    return dx > 0 and abs(dy) <= dx * TAN75

def _get_skipped_uncov(x_new, uncov_col):
    return sum(v for x, v in uncov_col.items() if x < x_new and v > 0)

# -- Step 6: A4 walk ------------------------------------------------------------
def _seed_walk_agreement(member, axis_rows=None, axis_cols=None,
                         n_seeds=40, win=6, seed=0):
    """Measure how curve-like a mask is by seeding random pixels and tracing a
    dense path left/right from each. On a real curve every seed converges to the
    same y in each column (low per-column spread); on scattered noise the seeds
    wander to different y's (high spread). Returns (agree, xcov):
      agree = median per-column std of seed y-votes (low => real curve)
      xcov  = fraction of image columns the union path covers.
    """
    m = member.astype(np.uint8).copy()
    for ay in (axis_rows or []):
        m[max(0, ay-2):ay+3, :] = 0
    for ax in (axis_cols or []):
        m[:, max(0, ax-2):ax+3] = 0
    m = m.astype(bool)
    H, W = m.shape
    ys, xs = np.where(m)
    if len(ys) == 0:
        return 999.0, 0.0
    rng = np.random.default_rng(seed)
    idx = rng.choice(len(ys), size=min(n_seeds, len(ys)), replace=False)
    votes = {}
    for i in idx:
        sx, sy = int(xs[i]), int(ys[i])
        # trace right then left following the local densest y within +/-win
        cur = float(sy)
        for x in range(sx, W):
            col = np.where(m[:, x])[0]
            if len(col) == 0:
                break
            near = col[np.abs(col - cur) <= win]
            if len(near) == 0:
                break
            y = int(np.median(near)); votes.setdefault(x, []).append(y)
            cur = 0.7*cur + 0.3*y
        cur = float(sy)
        for x in range(sx-1, -1, -1):
            col = np.where(m[:, x])[0]
            if len(col) == 0:
                break
            near = col[np.abs(col - cur) <= win]
            if len(near) == 0:
                break
            y = int(np.median(near)); votes.setdefault(x, []).append(y)
            cur = 0.7*cur + 0.3*y
    if len(votes) < 5:
        return 999.0, 0.0
    spreads = [np.std(v) for v in votes.values() if len(v) >= 3]
    agree = float(np.median(spreads)) if spreads else 999.0
    xcov = len(votes) / max(W, 1)
    return agree, xcov


def seed_walk_nodes(member, axis_rows=None, axis_cols=None,
                    n_seeds=40, win=6, seed=0, max_gap=40, max_vjump=40,
                    plot_area=None):
    """Build a curve path by seeding random pixels and tracing left/right from
    each, following the local densest y until the curve breaks. The union of all
    traces (per-column median of votes) is the path; small x-gaps are bridged by
    straight interpolation. Returns a list of (x, y) nodes, x-sorted.

    This is the method that gave the clean prototype results: multiple seeds make
    the big curve components get swept, so a bad start point can't truncate the
    path, and error-bar stems (few seeds, off the curve) are outvoted.

    plot_area : (x0, y0, x1, y1) - if given, everything outside this rectangle is
    blanked first, so axis labels / tick numbers outside the plotting area can
    never be seeded or traced.
    """
    m = member.astype(np.uint8).copy()
    H, W = m.shape
    if plot_area is not None:
        px0, py0, px1, py1 = [int(v) for v in plot_area]
        px0, px1 = sorted((max(0, px0), min(W - 1, px1)))
        py0, py1 = sorted((max(0, py0), min(H - 1, py1)))
        clip = np.zeros_like(m)
        clip[py0:py1+1, px0:px1+1] = m[py0:py1+1, px0:px1+1]
        m = clip
    for ay in (axis_rows or []):
        m[max(0, ay-2):ay+3, :] = 0
    for ax in (axis_cols or []):
        m[:, max(0, ax-2):ax+3] = 0
    m = m.astype(bool)
    H, W = m.shape
    ys, xs = np.where(m)
    if len(ys) == 0:
        return []
    rng = np.random.default_rng(seed)
    idx = rng.choice(len(ys), size=min(n_seeds, len(ys)), replace=False)
    votes = {}
    for i in idx:
        sx, sy = int(xs[i]), int(ys[i])
        cur = float(sy)
        for x in range(sx, W):
            col = np.where(m[:, x])[0]
            if len(col) == 0:
                break
            near = col[np.abs(col - cur) <= win]
            if len(near) == 0:
                break
            y = int(np.median(near)); votes.setdefault(x, []).append(y)
            cur = 0.7*cur + 0.3*y
        cur = float(sy)
        for x in range(sx-1, -1, -1):
            col = np.where(m[:, x])[0]
            if len(col) == 0:
                break
            near = col[np.abs(col - cur) <= win]
            if len(near) == 0:
                break
            y = int(np.median(near)); votes.setdefault(x, []).append(y)
            cur = 0.7*cur + 0.3*y
    if len(votes) < 2:
        return []
    track = {x: int(np.median(v)) for x, v in votes.items()}
    # gap-connect: link right end of a segment to left end of the next
    xs_t = sorted(track.keys())
    filled = dict(track)
    for i in range(1, len(xs_t)):
        x0, x1 = xs_t[i-1], xs_t[i]
        if 1 < x1 - x0 <= max_gap and abs(track[x1] - track[x0]) <= max_vjump:
            for xx in range(x0+1, x1):
                t = (xx - x0) / (x1 - x0)
                filled[xx] = int(round(track[x0]*(1-t) + track[x1]*t))
    nodes_out = [(x, filled[x]) for x in sorted(filled.keys())]
    # Trim vertical spikes at the two ends. A leading/trailing node that sits far
    # from both its neighbour AND the path's typical level is a stem artefact or a
    # bad gap-bridge, not the curve. Peel until the ends settle.
    _SPIKE = 25
    _med_y = float(np.median([n[1] for n in nodes_out])) if nodes_out else 0.0
    while len(nodes_out) >= 2 and (
            abs(nodes_out[0][1] - nodes_out[1][1]) > _SPIKE
            and abs(nodes_out[0][1] - _med_y) > _SPIKE):
        nodes_out.pop(0)
    while len(nodes_out) >= 2 and (
            abs(nodes_out[-1][1] - nodes_out[-2][1]) > _SPIKE
            and abs(nodes_out[-1][1] - _med_y) > _SPIKE):
        nodes_out.pop()
    return nodes_out


def detect_dense_curve(member, win=6, cont=None):
    """Extract one dense curve y=f(x) from a noisy binary mask.

    A data curve is a function: exactly one dominant y per x-column, varying
    smoothly across x. Error-bar stems (many y at one x) and scattered noise are
    not functions, so a per-column density vote plus a horizontal-continuity
    walk isolates the curve and rejects the rest.

    member : HxW bool/uint8 mask.
    win    : half-height (px) of the vertical density window per column.
    cont   : max |y jump| between adjacent kept columns (auto if None).
    Returns dict {x: y} for the curve's columns (may be sparse).
    """
    member = member.astype(bool)
    H, W = member.shape
    if cont is None:
        cont = max(10, H // 20)

    # Step 1: per-column dominant y = center of the densest +/-win band.
    col_y = {}
    col_strength = {}
    for x in range(W):
        ys = np.where(member[:, x])[0]
        if len(ys) == 0:
            continue
        # vectorised window count for each candidate y (each member pixel)
        lo = np.searchsorted(ys, ys - win, side='left')
        hi = np.searchsorted(ys, ys + win, side='right')
        counts = hi - lo
        j = int(np.argmax(counts))
        col_y[x] = int(ys[j])
        col_strength[x] = int(counts[j])
    if not col_y:
        return {}

    xs = sorted(col_y.keys())
    # Step 2: anchor from the strongest columns (robust to outlier levels).
    k = max(5, len(xs) // 4)
    strong = sorted(col_strength.items(), key=lambda kv: -kv[1])[:k]
    anchor = float(np.median([col_y[x] for x, _ in strong]))

    # Refine the anchor: among strong columns, the curve level is the y with the
    # widest horizontal support. Pick the densest y-bin over strong columns.
    strong_ys = np.array([col_y[x] for x, _ in strong])
    # bin into +/-win bands, choose the band center covering the most strong cols
    best_lvl, best_n = anchor, 0
    for cand in np.unique(strong_ys):
        n = int(np.sum(np.abs(strong_ys - cand) <= win))
        if n > best_n:
            best_n, best_lvl = n, float(cand)
    anchor = best_lvl

    # Step 3: continuity walk. Track follows the running estimate; a column that
    # deviates too far is skipped (stem / noise), and the estimate relaxes back
    # toward the anchor so the walk can re-acquire the curve after a gap.
    track = {}
    cur = anchor
    for x in xs:
        y = col_y[x]
        if abs(y - cur) <= cont:
            track[x] = y
            cur = 0.7 * cur + 0.3 * y      # smooth follow
        else:
            # skip this column; ease the estimate back toward the anchor so a
            # long stem region doesn't drag the track away permanently
            cur = 0.85 * cur + 0.15 * anchor
    return track


def run_A4_walk(walk_mask_in, dilated_mask_in, is_dashed=False, stem_mask=None, full_mask=None, force_x_start=None):
    """
    Orientation-aware A4 walk on walk_mask_in.
    stem_mask: if provided, columns dominated by stem pixels are excluded from
               the skipped-uncovered penalty so the walk doesn't stop at error bars.
    full_mask: if provided, x_end is taken from full_mask extent (not walk_mask)
               so the walk target extends beyond error-bar-only columns.
    force_x_start: if provided (int), override x_start selection to use this x value.
    Returns (nodes, coverage, reached_end).
    """
    ys_all, xs_all = np.where(walk_mask_in == 1)
    if len(xs_all) == 0:
        return [], 0.0, False

    x_span = int(xs_all.max()) - int(xs_all.min())
    y_span = int(ys_all.max()) - int(ys_all.min())
    rotated = y_span > x_span
    if rotated:
        walk_mask    = np.rot90(walk_mask_in,    k=-1)
        dilated_mask = np.rot90(dilated_mask_in, k=-1)
    else:
        walk_mask    = walk_mask_in
        dilated_mask = dilated_mask_in

    img_h, img_w = walk_mask.shape
    ys_w, xs_w   = np.where(walk_mask == 1)
    if len(xs_w) == 0:
        return [], 0.0, False

    total_mask_px = int(walk_mask.sum())

    # Classify line type
    col_counts = walk_mask.sum(axis=0).astype(float)
    zero_frac  = (col_counts == 0).sum() / max(img_w, 1)
    line_type  = 'dashed' if (is_dashed or zero_frac > 0.15) else 'solid'

    # Endpoints
    if force_x_start is not None:
        # Use forced x_start (e.g., leftmost column of full mask)
        x_start = int(force_x_start)
        cands   = np.where(xs_w == x_start)[0]
        if len(cands) == 0:
            x_start = int(xs_w[np.argmin(np.abs(xs_w - x_start))])
            cands   = np.where(xs_w == x_start)[0]
    else:
        col_px_w = walk_mask.sum(axis=0).astype(float)
        col_max = float(col_px_w.max()) if col_px_w.max() > 0 else 1.0
        main_cols = np.where(col_px_w >= col_max * 0.5)[0]
        if len(main_cols) > 0:
            x_start = int(main_cols.min())
        else:
            x_start = int(xs_w.min())
        cands   = np.where(xs_w == x_start)[0]
        if len(cands) == 0:
            x_start = int(xs_w[np.argmin(np.abs(xs_w - x_start))])
            cands   = np.where(xs_w == x_start)[0]
    y_start = int(np.median(ys_w[cands]))
    # Use full_mask extent for x_end so walk target reaches beyond stem-only columns
    if full_mask is not None:
        fm = np.rot90(full_mask, k=-1) if rotated else full_mask
        fm_xs = np.where(fm > 0)[1] if fm.ndim == 2 else np.where(fm > 0)[0]
        x_end = int(fm_xs.max()) if len(fm_xs) > 0 else int(xs_w.max())
    else:
        x_end   = int(xs_w.max())

    cx, cy  = x_start, y_start
    nodes   = [(cx, cy)]
    covered = np.zeros((img_h, img_w), dtype=np.uint8)
    _apply_capsule_coverage(cx, cy, cx, cy, covered, img_h, img_w)

    # Build uncov_col: exclude columns dominated by stem pixels (error bars)
    # so the walk doesn't stop prematurely at error bar positions.
    stem_cols = set()
    if stem_mask is not None:
        sm = np.rot90(stem_mask, k=-1) if rotated else stem_mask
        for xi in range(img_w):
            stem_cnt = int(sm[:, xi].sum()) if xi < sm.shape[1] else 0
            walk_cnt = int(walk_mask[:, xi].sum())
            # If stem pixels outnumber walk pixels at this column, treat as stem column
            if stem_cnt > walk_cnt and stem_cnt > 3:
                stem_cols.add(xi)
    uncov_col = {}
    for xi in range(img_w):
        if xi in stem_cols:
            continue  # skip stem columns from penalty
        cnt = int(walk_mask[:, xi].sum())
        if cnt > 0:
            uncov_col[xi] = cnt

    seg_fn  = _segment_in_mask_relaxed if line_type == 'dashed' else _segment_in_mask
    win_w   = WIN_W * 2 if line_type == 'dashed' else WIN_W
    max_iter = img_w * 3

    for _ in range(max_iter):
        best_score = -np.inf
        best_cand  = None

        x_lo = max(0, cx - win_w)
        x_hi = min(img_w - 1, cx + win_w)
        cand_mask = walk_mask[max(0, cy-win_w):min(img_h, cy+win_w+1),
                               x_lo:x_hi+1]
        cand_ys, cand_xs = np.where(cand_mask > 0)
        cand_xs += x_lo
        cand_ys += max(0, cy - win_w)

        tan_limit = TAN85 if line_type == 'dashed' else TAN75
        for xc, yc in zip(cand_xs.tolist(), cand_ys.tolist()):
            dx_ = xc - cx; dy_ = yc - cy
            if not (dx_ > 0 and abs(dy_) <= dx_ * tan_limit):
                continue
            if not seg_fn(cx, cy, xc, yc, dilated_mask, img_h, img_w):
                continue
            g      = _new_capsule_cov(cx, cy, xc, yc, covered, walk_mask, img_h, img_w)
            d      = float(np.hypot(xc - cx, yc - cy))
            dx_    = xc - cx; dy_ = yc - cy
            is_steep = abs(dy_) > abs(dx_)
            sk     = 0 if (line_type == 'dashed' or is_steep) else _get_skipped_uncov(xc, uncov_col)
            sc     = g - LAM * d - LAM * sk
            if sc > best_score:
                best_score = sc; best_cand = (xc, yc)

        stop_thresh = -20.0 if line_type == 'dashed' else -15.0
        if best_cand is None or best_score <= stop_thresh:
            break

        xc, yc = best_cand
        _apply_capsule_coverage(cx, cy, xc, yc, covered, img_h, img_w)
        dist_ = max(abs(xc-cx), abs(yc-cy))
        if dist_ > 0:
            ts_ = np.linspace(0, 1, dist_*2+1)
            xs_p_ = np.clip(np.round(cx + ts_*(xc-cx)).astype(int), 0, img_w-1)
            ys_p_ = np.clip(np.round(cy + ts_*(yc-cy)).astype(int), 0, img_h-1)
            for xi_, yi_ in zip(xs_p_.tolist(), ys_p_.tolist()):
                ys_d_, xs_d_ = _get_disk_pixels(xi_, yi_, img_h, img_w)
                for xd, yd in zip(xs_d_.tolist(), ys_d_.tolist()):
                    if walk_mask[yd, xd] == 1 and covered[yd, xd] == 1 and xd in uncov_col:
                        uncov_col[xd] = max(0, uncov_col[xd] - 1)
        cx, cy = xc, yc
        nodes.append((cx, cy))
        if cx >= x_end - 2:
            break

    cov_px   = int(np.sum((walk_mask == 1) & (covered == 1)))
    coverage = cov_px / max(total_mask_px, 1)

    if rotated:
        nodes = [(img_h - 1 - y, x) for x, y in nodes]

    print(f"  [walk] N={len(nodes)}  Cov={coverage:.1%}  type={line_type}")
    return nodes, coverage, True

# -- Step 7: Data-point estimation (density along walk) ------------------------
def _detect_markers_independent(full_mask):
    """Find data-point markers directly from the mask, independent of the walk.
    Markers are solid blobs (the plotted symbols); error-bar lines and dashed
    connectors are thin. Eroding the mask removes thin structures and leaves the
    marker cores, which we return as (x, y) centres. This recovers points on
    sawtooth / periodically-dosed curves where the walk only covers one segment.
    """
    m = (full_mask > 0).astype(np.uint8)
    if m.sum() == 0:
        return []
    # Exclude the legend region: its swatch markers would otherwise be detected
    # as spurious data points sitting in the upper-right empty space.
    lb = globals().get('_LEGEND_BOX_FOR_CLEAN', None)
    if lb is not None:
        lx0, ly0, lx1, ly1 = lb
        box_frac = ((lx1-lx0+1)*(ly1-ly0+1)) / float(H*W)
        if box_frac < 0.35:
            m = m.copy()
            m[max(0,ly0):min(H,ly1+1), max(0,lx0):min(W,lx1+1)] = 0
    eroded = cv2.erode(m, np.ones((3, 3), np.uint8), iterations=1)
    n, lbl, st, cen = cv2.connectedComponentsWithStats(eroded, 8)
    areas = [st[i, cv2.CC_STAT_AREA] for i in range(1, n)
             if st[i, cv2.CC_STAT_AREA] >= 15]
    if not areas:
        return []
    med = float(np.median(areas))
    pts = []
    for i in range(1, n):
        a = st[i, cv2.CC_STAT_AREA]
        if a >= max(30, med * 0.4):     # real marker, not a wisp
            pts.append((int(cen[i][0]), int(cen[i][1])))
    pts.sort()
    # merge duplicates (a marker split by erosion) within a small window
    merged = []
    for x, y in pts:
        if merged and abs(x - merged[-1][0]) < 12 and abs(y - merged[-1][1]) < 20:
            continue
        merged.append((x, y))
    return merged


def estimate_data_points(full_mask, nodes, cname):
    """
    Find data points using x-column density x marker density (+/-WALK_DENSITY_R px
    around walk-path y) combined signal.  Y coordinates are always taken from
    the walk path by linear interpolation.  Walk start/end are always included.
    Returns list of dicts: {x, y, fitness, px, density, source}.
    """
    if not nodes or len(nodes) < 2:
        # Walk failed -- fall back entirely to independent marker detection.
        mk = _detect_markers_independent(full_mask)
        return [{'x': x, 'y': y, 'fitness': 0.4, 'px': 1, 'source': 'marker'}
                for x, y in mk]

    xs_n = [n[0] for n in nodes]
    ys_n = [n[1] for n in nodes]

    # Build walk-path y lookup (linear interpolation over integer x range)
    x_min_w, x_max_w = int(min(xs_n)), int(max(xs_n))
    walk_y_at = {}  # x -> y on walk path
    for x in range(x_min_w, x_max_w + 1):
        walk_y_at[x] = float(np.interp(x, xs_n, ys_n))

    # Signal 1: x-column pixel count -- only within walk_y +/-WALK_DENSITY_R band
    # (excludes error bar pixels that are far from the walk path)
    col_px = np.zeros(W, dtype=float)
    for x in range(x_min_w, x_max_w + 1):
        wy = walk_y_at.get(x)
        if wy is None:
            col_px[x] = float(full_mask[:, x].sum())
            continue
        ys_col = np.where(full_mask[:, x] > 0)[0]
        if len(ys_col) == 0:
            continue
        min_dist_to_pixels = float(np.min(np.abs(ys_col - wy)))
        if min_dist_to_pixels > WALK_DENSITY_R:
            wy = float(np.median(ys_col))
        col_px[x] = float(np.sum(np.abs(ys_col - wy) <= WALK_DENSITY_R))

    # Signal 2: marker density -- pixels within +/-WALK_DENSITY_R of walk-path y
    # If walk_y is far from actual pixels (>WALK_DENSITY_R), use column median y instead
    marker_density = np.zeros(W, dtype=float)
    for x in range(x_min_w, x_max_w + 1):
        wy = walk_y_at[x]
        ys_col = np.where(full_mask[:, x] > 0)[0]
        if len(ys_col) == 0:
            continue
        # Check if walk_y is near the actual pixels
        min_dist_to_pixels = float(np.min(np.abs(ys_col - wy)))
        if min_dist_to_pixels > WALK_DENSITY_R:
            # Walk path is far from actual pixels -- use column median y as center
            wy = float(np.median(ys_col))
        marker_density[x] = float(np.sum(np.abs(ys_col - wy) <= WALK_DENSITY_R))

    # Combined signal (product, both normalised to [0,1])
    col_norm  = col_px / (col_px.max() + 1e-9)
    mark_norm = marker_density / (marker_density.max() + 1e-9)
    combined  = col_norm * mark_norm

    # Peak detection on raw combined signal
    if combined.max() == 0:
        peaks = np.array([], dtype=int)
    else:
        min_prom = combined.max() * 0.15
        min_dist = 20  # minimum pixel distance between data points
        peaks, _ = find_peaks(combined, prominence=min_prom, distance=min_dist)

    detections = []
    for pk in peaks:
        px = int(pk)
        if px not in walk_y_at:
            continue
        py = int(round(walk_y_at[px]))
        local_px = int(full_mask[max(0, py - WALK_DENSITY_R):min(H, py + WALK_DENSITY_R + 1),
                                  max(0, px - WALK_DENSITY_R):min(W, px + WALK_DENSITY_R + 1)].sum())
        fitness = round(float(combined[pk]) * min(1.0, local_px / 50.0), 3)
        detections.append({
            'x': px, 'y': py,
            'fitness': fitness,
            'px': local_px,
            'density': float(combined[pk]),
            'source': 'density',
        })

    # Always include walk start and end as data points
    # Y coordinate is always from the walk path (not column median)
    for ep_x, ep_y in [nodes[0], nodes[-1]]:
        ep_x = int(ep_x)
        ep_y = int(round(walk_y_at.get(ep_x, float(ep_y))))  # use walk path y
        too_close = any(abs(d['x'] - ep_x) < ENDPOINT_MERGE_DIST for d in detections)
        if not too_close:
            detections.append({'x': ep_x, 'y': ep_y, 'fitness': 0.5,
                                'px': 1, 'source': 'endpoint'})

    # Augment with walk-independent markers: the density-walk only covers the
    # span the walk reached, so on sawtooth/dashed curves it misses whole
    # segments. Add any marker not already close to an existing detection.
    markers = _detect_markers_independent(full_mask)
    for mx, my in markers:
        if not any(abs(d['x'] - mx) < ENDPOINT_MERGE_DIST for d in detections):
            detections.append({'x': int(mx), 'y': int(my), 'fitness': 0.4,
                                'px': 1, 'source': 'marker'})

    # Drop any detection that falls inside the legend box. The legend swatches
    # are the same colour as the curve, so the density/walk path can pick one up
    # as a spurious data point sitting in the legend.
    lb = globals().get('_LEGEND_BOX_FOR_CLEAN', None)
    if lb is not None:
        lx0, ly0, lx1, ly1 = lb
        box_frac = ((lx1 - lx0 + 1) * (ly1 - ly0 + 1)) / float(H * W)
        if box_frac < 0.35:
            detections = [d for d in detections
                          if not (lx0 <= d['x'] <= lx1 and ly0 <= d['y'] <= ly1)]

    # Drop detections sitting ON the x/y axis lines. After noise removal a few
    # axis-tick / axis-rule pixels can survive in a (grey/black) mask and line up
    # into a row of fake points along an axis. A real data point essentially
    # never sits exactly on the axis rule, so we remove any detection within a
    # couple of px of a detected axis row (x-axis) or column (y-axis).
    AX_TOL = 3
    if len(AXIS_ROWS) > 0:
        ax_y0, ax_y1 = int(AXIS_ROWS.min()) - AX_TOL, int(AXIS_ROWS.max()) + AX_TOL
        detections = [d for d in detections if not (ax_y0 <= d['y'] <= ax_y1)]
    if len(AXIS_COLS) > 0:
        ax_x0, ax_x1 = int(AXIS_COLS.min()) - AX_TOL, int(AXIS_COLS.max()) + AX_TOL
        detections = [d for d in detections if not (ax_x0 <= d['x'] <= ax_x1)]

    detections.sort(key=lambda d: d['x'])
    # If every detected point lies on essentially ONE horizontal line (tiny y
    # spread over a wide x-range), this colour is an axis/grid/LLOQ rule that
    # leaked through, not a real curve -- drop the whole set.
    if len(detections) >= 4:
        dys = np.array([d['y'] for d in detections])
        dxs = np.array([d['x'] for d in detections])
        y_spread = dys.max() - dys.min()
        x_spread = dxs.max() - dxs.min()
        # Where does this flat line sit relative to the plotting area? An axis /
        # grid rule hugs the top or bottom edge; a flat data curve (e.g. Placebo)
        # sits inside. Only treat edge-hugging flat lines as residue.
        _pa = PLOT_AREA if 'PLOT_AREA' in dir() and PLOT_AREA else (0, 0, W, H)
        _pay0, _pay1 = _pa[1], _pa[3]
        _line_y = float(np.median(dys))
        _edge_margin = max(6, 0.06 * (_pay1 - _pay0))
        _near_edge = (abs(_line_y - _pay0) <= _edge_margin or
                      abs(_line_y - _pay1) <= _edge_margin)
        if x_spread > 0.4 * W and y_spread <= 4 and _near_edge:
            print(f"    [{cname}] collinear-horizontal at plot edge "
                  f"(y={_line_y:.0f}) -- dropping as axis/grid residue")
            detections = []

    # Remove a lone outlier stranded in a far plot CORNER -- typically a stray
    # label glyph (e.g. a panel "(a)") that survived as one point. We target a
    # point that sits very near the top or left edge of the plotting area AND is
    # horizontally isolated from every other point. Using absolute edge-proximity
    # (not median deviation) avoids misfiring on steep curves whose real points
    # span a wide y-range.
    if len(detections) >= 3:
        pa = PLOT_AREA if PLOT_AREA is not None else (0, 0, W - 1, H - 1)
        px0, py0, px1, py1 = pa
        pw = max(1, px1 - px0); ph = max(1, py1 - py0)
        ds = sorted(detections, key=lambda d: d['x'])
        dxs = np.array([d['x'] for d in ds], float)
        dys = np.array([d['y'] for d in ds], float)
        keep = []
        for i, d in enumerate(ds):
            # y-jump to the x-nearest OTHER point
            others = [j for j in range(len(ds)) if j != i]
            jn = min(others, key=lambda j: abs(dxs[j] - dxs[i]))
            yjump = abs(dys[i] - dys[jn]) / ph
            near_top = (d['y'] - py0) < 0.06 * ph
            near_left = (d['x'] - px0) < 0.06 * pw
            xgap = abs(dxs[jn] - dxs[i])
            # corner residue: sits at the top/left edge AND jumps far in y from
            # its neighbour (a real curve point near the edge tracks its
            # neighbour's y; a stray glyph does not).
            if (near_top or near_left) and yjump > 0.30 and xgap < 0.25 * W:
                print(f"    [{cname}] dropping isolated corner outlier "
                      f"({int(d['x'])},{int(d['y'])})")
                continue
            keep.append(d)
        detections = keep

    print(f"    [{cname}] xdist+density: {len(detections)} peaks")
    return detections

# -- Step 8: T-cap extraction ---------------------------------------------------
def extract_tcaps(blob_mask, stem_mask, mode_xs, walk_nodes=None):
    """
    Extract T-cap positions from blob_mask at mode_xs columns.
    """
    if blob_mask.sum() == 0:
        return []
    walk_ys = {}
    if walk_nodes and len(walk_nodes) > 1:
        xs_n = [n[0] for n in walk_nodes]
        ys_n = [n[1] for n in walk_nodes]
        for x in range(W):
            if min(xs_n) <= x <= max(xs_n):
                walk_ys[x] = int(round(np.interp(x, xs_n, ys_n)))
    tcaps = []
    for x in mode_xs:
        col_blob = np.where(blob_mask[:, x] > 0)[0]
        if len(col_blob) == 0: continue
        wy = walk_ys.get(x, int(np.median(col_blob)))
        upper = col_blob[col_blob < wy]
        lower = col_blob[col_blob > wy]
        if len(upper) > 0:
            tcaps.append({'x': x, 'y': int(upper.min()), 'type': 'upper', 'area': len(upper)})
        if len(lower) > 0:
            tcaps.append({'x': x, 'y': int(lower.max()), 'type': 'lower', 'area': len(lower)})
    return tcaps

# -- Endpoint injection ---------------------------------------------------------
def inject_endpoints(detections, nodes):
    """Inject walk start/end as data points if not already covered.
    Y coordinate is taken directly from the walk node (walk path y).
    """
    if not nodes or len(nodes) < 2:
        return detections
    xs_n = [n[0] for n in nodes]
    ys_n = [n[1] for n in nodes]
    for ep_x, ep_y in [nodes[0], nodes[-1]]:
        ep_x = int(ep_x)
        # Use walk path y (interpolated) for the endpoint
        ep_y = int(round(float(np.interp(ep_x, xs_n, ys_n))))
        too_close = any(abs(d['x'] - ep_x) < ENDPOINT_MERGE_DIST for d in detections)
        if not too_close:
            detections.append({'x': ep_x, 'y': ep_y, 'fitness': 0.5,
                                'px': 1, 'source': 'endpoint'})
    return sorted(detections, key=lambda d: d['x'])

# -- Visualisation helpers ------------------------------------------------------
def save_raw_mask(cd, mask):
    """Save the true _raw_mask (before cleaning) with colour pixels on white bg."""
    mean_rgb = cd.get('mean_rgb', (128, 128, 128))
    true_raw = cd.get('_raw_mask')  # use original raw mask before cleaning
    use_mask = true_raw if (true_raw is not None) else mask
    canvas   = np.full((H, W, 3), 255, dtype=np.uint8)
    canvas[use_mask == 1] = mean_rgb
    canvas_bgr = cv2.cvtColor(canvas, cv2.COLOR_RGB2BGR)
    # Store canvas for later walk overlay
    cd['_raw_mask_canvas'] = canvas_bgr
    path = os.path.join(OUT_DIR, f'raw_mask_{cd["name"]}.jpg')
    cv2.imwrite(path, canvas_bgr)

def generate_montage(out_dir, orig_img, all_detections, all_tcaps_dict, all_results):
    OH, OW = orig_img.shape[:2]
    panels = []
    for k, (nodes, cov, *_) in all_results.items():
        pts_list  = all_detections.get(k, [])
        tcap_list = all_tcaps_dict.get(k, [])
        panel = np.full((OH, OW, 3), 40, dtype=np.uint8)
        if len(nodes) > 1:
            pts_arr = np.array([[int(p[0]), int(p[1])] for p in nodes], dtype=np.int32)
            cv2.polylines(panel, [pts_arr], False, (160,160,160), 2)
            cv2.circle(panel, tuple(pts_arr[0]),  5, (0,200,0),   -1)
            cv2.circle(panel, tuple(pts_arr[-1]), 5, (0,200,200), -1)
        for pt in pts_list:
            cv2.circle(panel, (int(pt['x']), int(pt['y'])), 7, (0,0,220), -1)
            cv2.circle(panel, (int(pt['x']), int(pt['y'])), 7, (255,255,255), 1)
        for tc in tcap_list:
            half = 5
            cv2.rectangle(panel, (int(tc['x'])-half, int(tc['y'])-half),
                          (int(tc['x'])+half, int(tc['y'])+half), (255,255,0), -1)
        label = f'{k}  pts={len(pts_list)}  N={len(nodes)}'
        cv2.putText(panel, label, (4,18), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (255,255,255), 3)
        cv2.putText(panel, label, (4,18), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (20,20,20), 1)
        scale = 380 / OW
        panel = cv2.resize(panel, (380, int(OH*scale)), interpolation=cv2.INTER_AREA)
        panels.append((k, panel))
    if not panels: return
    ncols = 3; nrows = (len(panels)+ncols-1)//ncols
    ph, pw = panels[0][1].shape[:2]
    grid = np.full((nrows*ph+44, ncols*pw, 3), 255, dtype=np.uint8)
    cv2.putText(grid, 'Walk + detected points', (6,30),
                cv2.FONT_HERSHEY_SIMPLEX, 0.6, (30,30,30), 1)
    for idx, (_, panel) in enumerate(panels):
        r, c = divmod(idx, ncols)
        grid[44+r*ph:44+r*ph+ph, c*pw:c*pw+pw] = panel
    cv2.imwrite(os.path.join(out_dir, 'walk_points_montage.jpg'), grid,
                [cv2.IMWRITE_JPEG_QUALITY, 93])

# ==============================================================================
# MAIN PIPELINE
# ==============================================================================
print("\n--- Stage 2: Colour discovery ---")
COLORS, LEGEND_BOX = _discover_colors()
print(f"FINAL_LEGEND_BOX: {LEGEND_BOX if LEGEND_BOX else 'NONE'}")

# --- Noise-colour filter -----------------------------------------------------
# A colour mask that contains a real data curve makes randomly-seeded traces
# converge to one y per column (low spread). A mask that is only scattered dust
# (e.g. a legend colour that has no curve in this panel, or leftover speckle)
# makes the traces scatter. Drop colours whose seed-walk agreement is poor.
#
# EXCEPTION: if the legend matched the masks 1:1, every colour is a real curve
# (a periodic/spiky dose curve can look "high spread" but is genuine), so we
# trust the legend and skip filtering.
_axis_rows_list = list(AXIS_ROWS) if 'AXIS_ROWS' in dir() else None
_axis_cols_list = list(AXIS_COLS) if 'AXIS_COLS' in dir() else None
_legend_1to1 = bool(globals().get('_LEGEND_1TO1', False))
if _legend_1to1:
    print("  [noise filter] skipped (legend matched 1:1; all colours trusted)")
else:
    _kept_colors = []
    for _cd in COLORS:
        _rm = _cd.get('_raw_mask')
        if _rm is None:
            _kept_colors.append(_cd); continue
        _mem = _rm > 0 if _rm.dtype != bool else _rm
        if _mem.ndim == 3:
            _mem = _mem.any(axis=2)
        _agree, _xcov = _seed_walk_agreement(_mem, _axis_rows_list, _axis_cols_list)
        # real curve: seeds agree (low spread) AND path spans a decent x-range
        _is_noise = (_agree > NOISE_AGREE_MAX) or (_xcov < NOISE_XCOV_MIN)
        if _is_noise:
            print(f"  [noise filter] dropping {_cd['name']} "
                  f"(agree={_agree:.1f}, xcov={_xcov:.2f})")
        else:
            _kept_colors.append(_cd)
    if _kept_colors:
        COLORS = _kept_colors
    else:
        print("  [noise filter] all colours looked like noise; keeping original set")

print(f"\n--- Stage 3-8: Per-colour processing ---")
all_results    = {}
all_detections = {}
all_tcaps_out  = {}
all_wm_full    = {}  # stores wm_full (axis-removed clean mask) per colour

# ============================================================================
# Step 3-8 REPLACEMENT: data detection via plot_digitizer.PlotDigitizer
# ----------------------------------------------------------------------------
# The surrounding v24 infrastructure is untouched: axis detection (AXIS_ROWS/
# AXIS_COLS/PLOT_AREA), legend + colour discovery (COLORS), noise-colour filter,
# text/LLOQ/dashed handling that produced the clean masks, and everything AFTER
# this block (calibration, detections.json, data_points.xlsx, overlay) still run.
# Only the per-colour curve/blob/stem classification + A4 walk + data-point +
# tcap estimation is swapped for the cleaner PlotDigitizer core.
# ----------------------------------------------------------------------------
# ---- embedded plot_digitizer core (was: from plot_digitizer import ...) ----
import pickle
from collections import defaultdict, Counter
from scipy.signal import medfilt

class PlotConfig:
    """Everything plot-specific lives here."""
    def __init__(self):
        # plot area (data region) in pixels: (x0, y0, x1, y1)
        self.plot_area = (35, 67, 332, 285)
        # legend box to exclude (x0, y0, x1, y1) or None
        self.legend_box = (70, 0, 280, 67)
        # data-point columns (x pixel of each stem) and their axis values
        self.stem_x = [43, 64, 89, 137, 185, 281, 328]
        self.x_values = [12, 14, 16, 20, 24, 32, 36]      # weeks
        # y-axis calibration: (pixel_y, value) at two reference points
        # fill these once you read the ticks; used only for value conversion
        self.y_ref = None      # e.g. [(py0_pixel, 20.0), (py1_pixel, 0.0)]
        # number of curves / their names (order matches auto-clustered palette)
        self.curve_names = ['blue', 'green', 'red', 'black']
        # only upper error bars present -> <=1 cap per color per stem
        self.upper_bars_only = True


# --------------------------------------------------------------------------- #
#  Core digitizer
# --------------------------------------------------------------------------- #
class PlotDigitizer:
    def __init__(self, image_path, config=None):
        self.cfg = config or PlotConfig()
        self.img = cv2.imread(image_path)
        if self.img is None:
            raise FileNotFoundError(image_path)
        self.H, self.W = self.img.shape[:2]
        self.rgb = cv2.cvtColor(self.img, cv2.COLOR_BGR2RGB)
        self.lab = cv2.cvtColor(self.rgb, cv2.COLOR_RGB2Lab).astype(np.float32)
        hsv = cv2.cvtColor(self.rgb, cv2.COLOR_RGB2HSV)
        self.s = hsv[:, :, 1].astype(np.float32)
        self.v = hsv[:, :, 2].astype(np.float32)
        self.a_map = self.lab[:, :, 1] - 128
        self.b_map = self.lab[:, :, 2] - 128
        self.chroma_map = np.hypot(self.a_map, self.b_map)
        self.ang_map = np.degrees(np.arctan2(self.b_map, self.a_map))

        self._build_ink_mask()
        self._auto_palette()

        # results filled by run()
        self.assign = None            # final stage-2 label map (-1/0..3)
        self.stems = None             # original-detected stem mask
        self.caps = {}                # {stem_x: {color_k: (y, left, right)}}
        self.data_points = {}         # {x_value: {color_k: y_pixel}}

    def set_palette(self, rgb_list, add_black_if_missing=True, add_black_sink=False):
        """Override the auto-clustered palette with an externally supplied one
        (e.g. v24's legend-discovered COLORS). rgb_list = [(r,g,b), ...].

        If every supplied color is chromatic and `add_black_if_missing`, a black
        entry is APPENDED (not forced onto a colored swatch) -- this handles the
        common case where the legend detector drops the black curve because it
        shares its ink color with axes/text/LLOQ. The black curve is real and
        needs its own slot, otherwise the darkest colored swatch would wrongly
        absorb it and that swatch's own curve would be lost.

        `add_black_sink`: when the legend is all-chromatic (no black CURVE), the
        plot's black structural ink (error bars, axis) would otherwise be dumped
        into the darkest chromatic mask -- because with no achromatic slot the
        code below is forced to nominate the darkest colour as `black_k`. Appending
        a real black SINK gives that ink its own bin so the coloured masks stay
        clean. The sink is flagged in `self.is_sink` and is NOT a data curve."""
        pal = [tuple(int(c) for c in rgb) for rgb in rgb_list]
        lab = [cv2.cvtColor(np.uint8([[c]]), cv2.COLOR_RGB2Lab)[0, 0] for c in pal]
        is_ach = [abs(l[1] - 128) < 8 and abs(l[2] - 128) < 8 for l in lab]

        added_black = False
        if add_black_if_missing and not any(is_ach):
            # does the plot actually contain a dark-neutral curve? (not just text)
            if self._has_dark_neutral_curve():
                pal = list(pal) + [(60, 60, 60)]
                added_black = True

        added_sink = False
        if add_black_sink and not added_black:
            has_dark_ach = any(is_ach[i] and (sum(pal[i]) / 3.0) < 80
                               for i in range(len(pal)))
            if not has_dark_ach:
                pal = list(pal) + [(40, 40, 40)]
                added_sink = True

        self.palette = pal
        self.pal_lab = np.array(
            [cv2.cvtColor(np.uint8([[c]]), cv2.COLOR_RGB2Lab)[0, 0] for c in pal],
            dtype=np.float32)
        self.achro = np.array(
            [abs(l[1] - 128) < 8 and abs(l[2] - 128) < 8 for l in self.pal_lab])
        if not self.achro.any():
            dark = int(np.argmin([l[0] for l in self.pal_lab]))
            self.achro = np.zeros(len(pal), bool)
            self.achro[dark] = True
        self.chrom_idx = [k for k in range(len(pal)) if not self.achro[k]]
        # black_k = the DARKEST achromatic slot (the sink when present), so black
        # stem/error-bar ink is assigned there and a lighter grey CURVE slot is
        # not contaminated by it.
        _achk = [k for k in range(len(pal)) if self.achro[k]]
        self.black_k = min(_achk, key=lambda k: float(self.pal_lab[k, 0]))
        # mark the sink slot (if any) so downstream excludes it from curves
        self.is_sink = np.zeros(len(pal), bool)
        if added_sink:
            self.is_sink[-1] = True
        self.pal_ang = {j: np.degrees(np.arctan2(self.pal_lab[j, 2] - 128,
                                                 self.pal_lab[j, 1] - 128))
                        for j in self.chrom_idx}
        if len(self.cfg.curve_names) != len(pal):
            base = list(self.cfg.curve_names)[:len(pal)]
            while len(base) < len(pal):
                if added_sink and len(base) == len(pal) - 1:
                    base.append('black_sink')
                elif added_black and len(base) == len(pal) - 1:
                    base.append('black')
                else:
                    base.append(f'color{len(base)}')
            self.cfg.curve_names = base
        return self

    def _has_dark_neutral_curve(self):
        """True only if the plot area contains a real dark-neutral DATA CURVE,
        not just axis lines or scattered text. We remove long axis-like straight
        runs and require the remaining dark ink to (a) cover many x-columns and
        (b) have a non-flat vertical profile (a curve moves up/down; an axis line
        or a single text band does not)."""
        x0, y0, x1, y1 = self.cfg.plot_area
        m = self.ink.copy()
        dark = m & (self.lab[:, :, 0] < 150) & (self.chroma_map < 8)
        if dark.sum() < 60:
            return False
        du = dark.astype(np.uint8)
        # remove long horizontal axis lines (full-width thin runs) and long
        # vertical axis lines via morphological opening, then subtract
        hk = cv2.getStructuringElement(cv2.MORPH_RECT, (int((x1 - x0) * 0.5), 1))
        vk = cv2.getStructuringElement(cv2.MORPH_RECT, (1, int((y1 - y0) * 0.5)))
        axis_h = cv2.morphologyEx(du, cv2.MORPH_OPEN, hk)
        axis_v = cv2.morphologyEx(du, cv2.MORPH_OPEN, vk)
        curve = du & ~axis_h.astype(bool) & ~axis_v.astype(bool)
        cy, cx = np.where(curve)
        if len(cx) < 50:
            return False
        # per-column median y of remaining dark ink -> a curve varies in y across
        # x; text sits in a narrow y-band; leftover axis fragments are flat.
        from collections import defaultdict
        col = defaultdict(list)
        for y, x in zip(cy, cx):
            col[x].append(y)
        xs_u = sorted(col)
        col_cover = len(xs_u) / max(1, (x1 - x0))
        if col_cover < 0.25:
            return False
        med_y = np.array([np.median(col[x]) for x in xs_u])
        yspread = med_y.max() - med_y.min()
        # a real curve spans wide in x AND moves in y (or covers the plot broadly)
        xspan = xs_u[-1] - xs_u[0]
        return xspan > (x1 - x0) * 0.4 and (yspread > (y1 - y0) * 0.08 or col_cover > 0.5)

    # ---- masks & palette --------------------------------------------------- #
    def _build_ink_mask(self):
        x0, y0, x1, y1 = self.cfg.plot_area
        inplot = np.zeros((self.H, self.W), bool)
        inplot[y0:y1 + 1, x0:x1 + 1] = True
        if self.cfg.legend_box:
            lx0, ly0, lx1, ly1 = self.cfg.legend_box
            inplot[ly0:ly1 + 1, lx0:lx1 + 1] = False
        self.inplot = inplot
        # ink = inside plot, not near-white
        self.ink = inplot & ~((self.v > 235) & (self.s < 25))

    def _auto_palette(self):
        """Cluster the plot's ink pixels by hue to recover the 4 legend colors.
        Robust to mild JPEG color shifts because it samples the actual image."""
        ys, xs = np.where(self.ink)
        cols = self.rgb[ys, xs].astype(np.float32)
        labc = self.lab[ys, xs]
        chroma = np.hypot(labc[:, 1] - 128, labc[:, 2] - 128)
        hue = np.degrees(np.arctan2(labc[:, 2] - 128, labc[:, 1] - 128))
        # black = dark neutral
        black_mask = (labc[:, 0] < 110) & (chroma < 15)
        black_col = np.median(cols[black_mask], axis=0) if black_mask.any() else np.array([60, 60, 60.])
        # chromatic pixels
        chrom = chroma > 20
        ch_cols, ch_hue = cols[chrom], hue[chrom]
        # cluster the chromatic hues into 3 (blue / green / red-orange).
        # Use k-means on hue angle (circular) -> here the 3 groups are well separated.
        def med(mask):
            return np.median(ch_cols[mask], axis=0) if mask.any() else None
        # heuristics for common PK palette; generalize by histogram peaks if needed
        blue_m = (ch_hue < -30) | (ch_hue > 200)
        green_m = (ch_hue >= 100) & (ch_hue < 170)
        red_m = (ch_hue > 20) & (ch_hue < 75)
        pal = [med(blue_m), med(green_m), med(red_m), black_col]
        # fallbacks
        defaults = [(150, 157, 209), (148, 189, 132), (210, 138, 117), (83, 76, 73)]
        pal = [tuple(int(c) for c in (p if p is not None else d))
               for p, d in zip(pal, defaults)]
        self.palette = pal
        self.pal_lab = np.array(
            [cv2.cvtColor(np.uint8([[c]]), cv2.COLOR_RGB2Lab)[0, 0] for c in pal],
            dtype=np.float32)
        self.achro = np.array(
            [abs(l[1] - 128) < 8 and abs(l[2] - 128) < 8 for l in self.pal_lab])
        self.chrom_idx = [k for k in range(4) if not self.achro[k]]
        self.black_k = [k for k in range(4) if self.achro[k]][0]
        self.pal_ang = {j: np.degrees(np.arctan2(self.pal_lab[j, 2] - 128,
                                                 self.pal_lab[j, 1] - 128))
                        for j in self.chrom_idx}

    # ---- STAGE 1 ----------------------------------------------------------- #
    def stage1(self):
        """1a pure color seeds + 1b antialiasing absorption."""
        ys, xs = np.where(self.ink)
        pl = self.lab[ys, xs]
        L = pl[:, 0]
        a = pl[:, 1] - 128
        b = pl[:, 2] - 128
        chroma = np.hypot(a, b)
        ang = np.degrees(np.arctan2(b, a))
        assign = np.full((self.H, self.W), -1, np.int32)

        # Data-driven grey/black split: when there are two achromatic slots (a grey
        # CURVE + a black sink), threshold the plot's own neutral-ink lightness
        # histogram (Otsu) instead of trusting the palette swatch values, which are
        # often mis-measured. The lighter cluster is the grey curve, the darker the
        # black (axis / error bars).
        self._neu_thr = None
        _achk0 = [k for k in range(len(self.palette)) if self.achro[k]]
        self._neu_lighter = self.black_k
        self._neu_darker = self.black_k
        if len(_achk0) >= 2:
            self._neu_lighter = max(_achk0, key=lambda k: float(self.pal_lab[k, 0]))
            self._neu_darker = min(_achk0, key=lambda k: float(self.pal_lab[k, 0]))
            _neuL = L[chroma < 8]
            if len(_neuL) > 40:
                _t, _ = cv2.threshold(_neuL.astype(np.uint8), 0, 255,
                                      cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                if 55 <= _t <= 190:
                    self._neu_thr = float(_t)

        def _neu_target_vec(Lvals):
            """slot index per neutral pixel: grey (lighter) vs black (darker)."""
            if len(_achk0) <= 1:
                return np.full(len(Lvals), self.black_k, np.int32)
            if self._neu_thr is not None:
                return np.where(np.asarray(Lvals) >= self._neu_thr,
                                self._neu_lighter, self._neu_darker).astype(np.int32)
            _aL = np.array([float(self.pal_lab[k, 0]) for k in _achk0])
            return np.array(_achk0)[np.argmin(
                np.abs(np.asarray(Lvals)[:, None] - _aL[None, :]), axis=1)].astype(np.int32)
        self._neu_target_vec = _neu_target_vec

        # 1a-i chromatic seeds (only if there are chromatic palette colors)
        if self.chrom_idx:
            d = np.linalg.norm(pl[:, None, :] - self.pal_lab[None, self.chrom_idx, :], axis=2)
            ci = np.argmin(d, axis=1)
            mind = d[np.arange(len(ys)), ci]
            for i, k in enumerate(self.chrom_idx):
                ad = np.abs((ang - self.pal_ang[k] + 180) % 360 - 180)
                strong = (mind <= 32) & (chroma >= 15) & (ci == i)
                pale = (chroma >= 6) & (ad < 22) & (L > 130)  # pale but clearly hued
                sel = strong | pale
                assign[ys[sel], xs[sel]] = k
        else:
            mind = np.full(len(ys), 1e9, np.float32)

        # 1a-ii black seeds: neutral only (L<=215 & chroma<8), never a hued pixel
        already = assign[ys, xs] >= 0
        is_hue = np.zeros(len(ys), bool)
        for j in self.chrom_idx:
            adj = np.abs((ang - self.pal_ang[j] + 180) % 360 - 180)
            is_hue |= (adj < 25) & (chroma >= 10)
        blk = (L <= 215) & (chroma < 8) & (~already) & (~is_hue)
        _achk = [k for k in range(len(self.palette)) if self.achro[k]]
        if len(_achk) <= 1:
            assign[ys[blk], xs[blk]] = self.black_k
        else:
            assign[ys[blk], xs[blk]] = self._neu_target_vec(L[blk])
        self.seeds = assign.copy()

        # 1b antialiasing absorption (location-based, 3 dilations of radius 1)
        k8 = np.ones((3, 3), np.uint8)
        for _ in range(3):
            un = self.ink & (assign < 0)
            best_c = np.full((self.H, self.W), -1, np.int32)
            best_s = np.full((self.H, self.W), 1e9, np.float32)
            for k in range(len(self.palette)):
                cm = (assign == k).astype(np.uint8)
                if cm.sum() == 0:
                    continue
                fr = un & cv2.dilate(cm, k8).astype(bool)
                fy, fx = np.where(fr)
                if len(fy) == 0:
                    continue
                pchr = self.chroma_map[fy, fx]
                pang = self.ang_map[fy, fx]
                pL = self.lab[fy, fx, 0]
                if self.achro[k]:
                    valid = (pchr < 5) & (pL < 235)
                    score = pchr
                else:
                    adk = np.abs((pang - self.pal_ang[k] + 180) % 360 - 180)
                    valid = (pchr >= 3) & (adk < 30)
                    score = adk
                iv = np.where(valid)[0]
                better = score[iv] < best_s[fy[iv], fx[iv]]
                sel = iv[np.where(better)[0]]
                best_s[fy[sel], fx[sel]] = score[sel]
                best_c[fy[sel], fx[sel]] = k
            newly = best_c >= 0
            if not newly.any():
                break
            assign[newly] = best_c[newly]
        self.assign = assign
        return assign

    # ---- STAGE 2 ----------------------------------------------------------- #
    def _local_dir(self, mask, cx, cy, r=5):
        y0, y1 = max(0, cy - r), min(self.H, cy + r + 1)
        x0, x1 = max(0, cx - r), min(self.W, cx + r + 1)
        yy, xx = np.where(mask[y0:y1, x0:x1])
        if len(yy) < 4:
            return None, None
        pts = np.stack([xx.astype(float), yy.astype(float)], 1)
        pts -= pts.mean(0)
        w, vv = np.linalg.eigh(pts.T @ pts)
        aniso = 999 if w.min() < 1e-6 else w.max() / w.min()
        d = vv[:, np.argmax(w)]
        ang = np.degrees(np.arctan2(d[1], d[0])) % 180
        return ang, aniso

    def refine_confusable(self, pal_lab, prio_bias, pairs):
        """Re-resolve confusable colour PAIRS on the FINAL label map. stage1/stage2
        classify pixels by plain nearest-centroid, which mixes colours that differ
        mainly in brightness (red vs orange, a navy vs a blue-grey). For each
        flagged pair we take every pixel currently assigned to either colour and
        re-decide between the two using a HUE-weighted distance (L down-weighted)
        minus the priority bias, so the brightened anti-aliased edge of the vivid
        colour stays with it instead of bleeding into the duller one."""
        if not pairs or self.assign is None:
            return
        lab = self.lab
        for (i, j) in pairs:
            if i >= len(pal_lab) or j >= len(pal_lab):
                continue
            both = (self.assign == i) | (self.assign == j)
            ys, xs = np.where(both)
            if len(ys) == 0:
                continue
            P = lab[ys, xs].astype(np.float32)
            ci = pal_lab[i]; cj = pal_lab[j]
            Lw = NC_CONFUSE_L_WEIGHT
            di = (np.abs(P[:, 0]-ci[0])*Lw + np.abs(P[:, 1]-ci[1])*1.3
                  + np.abs(P[:, 2]-ci[2])*1.3) - float(prio_bias[i])
            dj = (np.abs(P[:, 0]-cj[0])*Lw + np.abs(P[:, 1]-cj[1])*1.3
                  + np.abs(P[:, 2]-cj[2])*1.3) - float(prio_bias[j])
            to_i = di <= dj
            self.assign[ys[to_i], xs[to_i]] = i
            self.assign[ys[~to_i], xs[~to_i]] = j

    def stage2(self):
        assign = self.assign
        x0p, y0p, x1p, y1p = self.cfg.plot_area
        k8 = np.ones((3, 3), np.uint8)

        # (A) region growing: favor horizontal (curve), penalize vertical (stem)
        for _ in range(400):
            un = self.ink & (assign < 0)
            if not un.any():
                break
            best_c = np.full((self.H, self.W), -1, np.int32)
            best_s = np.full((self.H, self.W), 1e9, np.float32)
            for k in range(len(self.palette)):
                cm = (assign == k).astype(np.uint8)
                if cm.sum() == 0:
                    continue
                fr = un & cv2.dilate(cm, k8).astype(bool)
                fy, fx = np.where(fr)
                if len(fy) == 0:
                    continue
                pa = self.a_map[fy, fx]
                pb = self.b_map[fy, fx]
                pang = np.degrees(np.arctan2(pb, pa))
                pchr = np.hypot(pa, pb)
                dab = np.sqrt((pa - (self.pal_lab[k, 1] - 128)) ** 2 +
                              (pb - (self.pal_lab[k, 2] - 128)) ** 2)
                if self.achro[k]:
                    near = np.zeros(len(fy), bool)
                    for j in self.chrom_idx:
                        near |= (np.abs((pang - self.pal_ang[j] + 180) % 360 - 180) < 35) & (pchr >= 5)
                    valid = (pchr < 8) & (~near)
                else:
                    ad = np.abs((pang - self.pal_ang[k] + 180) % 360 - 180)
                    valid = ((pchr >= 6) & (ad < 35)) | ((pchr >= 4) & (ad < 30))
                score = dab.copy()
                cur = (assign == k)
                for ii in range(len(fy)):
                    ang, aniso = self._local_dir(cur, fx[ii], fy[ii])
                    if ang is None:
                        continue
                    if aniso > 4:
                        if ang < 30 or ang > 150:
                            score[ii] *= 0.4      # horizontal curve -> prefer
                        elif 60 < ang < 120:
                            score[ii] *= 3.0      # vertical stem -> avoid
                iv = np.where(valid)[0]
                better = score[iv] < best_s[fy[iv], fx[iv]]
                sel = iv[np.where(better)[0]]
                best_s[fy[sel], fx[sel]] = score[sel]
                best_c[fy[sel], fx[sel]] = k
            newly = best_c >= 0
            if not newly.any():
                break
            assign[newly] = best_c[newly]

        # achromatic slots (grey curve vs black sink); neutral ink is split by the
        # data-driven lightness threshold computed in stage1 (self._neu_thr).
        _achk = [k for k in range(len(self.palette)) if self.achro[k]]

        def _nearest_achro_scalar(Lval):
            if len(_achk) <= 1:
                return self.black_k
            if getattr(self, '_neu_thr', None) is not None:
                return self._neu_lighter if Lval >= self._neu_thr else self._neu_darker
            _aL = [float(self.pal_lab[k, 0]) for k in _achk]
            return _achk[int(np.argmin([abs(Lval - l) for l in _aL]))]

        # (B) grey vertical stems -> nearest achromatic slot by lightness
        un = self.ink & (assign < 0)
        n, lbl, st, _ = cv2.connectedComponentsWithStats(un.astype(np.uint8), 8)
        for i in range(1, n):
            w = st[i, cv2.CC_STAT_WIDTH]
            h = st[i, cv2.CC_STAT_HEIGHT]
            if h >= w * 1.3 and h >= 6:
                comp = (lbl == i)
                cy, cx = np.where(comp)
                med_ch = np.median(np.hypot(self.lab[cy, cx, 1] - 128,
                                            self.lab[cy, cx, 2] - 128))
                if med_ch < 4:
                    assign[comp] = _nearest_achro_scalar(np.median(self.lab[cy, cx, 0]))

        # (C) leftover grey -> nearest achromatic slot by lightness (per pixel)
        un = self.ink & (assign < 0)
        uy, ux = np.where(un)
        uchr = np.hypot(self.lab[uy, ux, 1] - 128, self.lab[uy, ux, 2] - 128)
        neut = uchr < 8
        ny, nx = uy[neut], ux[neut]
        if len(ny):
            if len(_achk) <= 1:
                assign[ny, nx] = self.black_k
            else:
                assign[ny, nx] = self._neu_target_vec(self.lab[ny, nx, 0])

        # (D) dust removal (must come right after grey->black)
        for k in range(len(self.palette)):
            m = (assign == k).astype(np.uint8)
            nn, ll, ss, _ = cv2.connectedComponentsWithStats(m, 8)
            for i in range(1, nn):
                if ss[i, cv2.CC_STAT_AREA] < 3:
                    assign[ll == i] = -1

        self.assign = assign
        return assign

    # ---- STAGE 2E : original stem/cap detection + graft --------------------- #
    def detect_stems(self):
        """Detect error-bar stems strictly INSIDE the axes, and DECOUPLE two uses:
          * self.stems          — PERMISSIVE (any tall vertical inside the axes):
                                  used to find data COLUMNS, so completeness does
                                  not collapse on plots with short/few error bars.
          * self.stems_confirmed— STRICT (thin + isolated + tall, not on the curve
                                  line): only these are subtracted from the curve
                                  in _seg_y, so real curve pixels are not eaten.
        Excludes the y-axis / x-axis lines themselves (they are not data)."""
        x0p, y0p, x1p, y1p = self.cfg.plot_area
        y0p = max(y0p, getattr(self.cfg, 'data_top', y0p))    # below a top legend band
        ax_x = getattr(self.cfg, 'axis_x', x0p)      # y-axis column
        ax_y = getattr(self.cfg, 'axis_y', y1p)      # x-axis row
        interior = np.zeros_like(self.ink)
        interior[max(0, y0p):max(0, ax_y - 2),
                 min(self.W, ax_x + 2):min(self.W, x1p + 1)] = True
        ink = (self.ink & interior).astype(np.uint8)

        # PERMISSIVE stems (for columns)
        self.stems = cv2.morphologyEx(
            ink, cv2.MORPH_OPEN,
            cv2.getStructuringElement(cv2.MORPH_RECT, (1, 9)))

        # curve region = wide HORIZONTAL ink (curve line + marker body)
        self.curve_region = cv2.morphologyEx(
            ink, cv2.MORPH_OPEN,
            cv2.getStructuringElement(cv2.MORPH_RECT, (4, 1))).astype(bool)

        # STRICT confirmed error bars (for subtracting from the curve)
        st = self.stems.astype(bool) & ~self.curve_region            # thickness
        left = np.zeros_like(self.ink); left[:, 3:] = self.ink[:, :-3]
        right = np.zeros_like(self.ink); right[:, :-3] = self.ink[:, 3:]
        st = st & ~(left & right)                                    # isolation
        n, lbl, stt, _ = cv2.connectedComponentsWithStats(st.astype(np.uint8), 8)
        conf = np.zeros_like(self.stems)
        for i in range(1, n):
            if stt[i, cv2.CC_STAT_HEIGHT] >= 12 and stt[i, cv2.CC_STAT_WIDTH] <= 3:
                conf[lbl == i] = 1                                    # tall & thin
        self.stems_confirmed = conf
        return self.stems

    def _stem_groups(self):
        sy, sx = np.where(self.stems)
        cols = {}
        for x in np.unique(sx):
            ys = sy[sx == x]
            if len(ys) >= 8:
                cols[x] = (ys.min(), ys.max())
        groups = []
        for x in sorted(cols):
            if groups and x - groups[-1][-1] <= 3:
                groups[-1].append(x)
            else:
                groups.append([x])
        return groups, cols

    def detect_caps(self):
        """Detect error-bar caps on the ORIGINAL, classify by color, validate.
        A valid cap: short (4-13px), horizontal, vertically isolated, symmetric
        about the stem, crosses the stem as ONE contiguous run, cap-color ==
        stem-below-color, <=1 per color (topmost)."""
        x0p, y0p, x1p, y1p = self.cfg.plot_area
        groups, cols = self._stem_groups()

        def classify(y, left, right, xc):
            cc = np.array([self.lab[y, x] for x in range(left, right + 1)])
            a = np.median(cc[:, 1]) - 128
            b = np.median(cc[:, 2]) - 128
            ch = np.hypot(a, b)
            ang = np.degrees(np.arctan2(b, a))
            if ch < 6 or not self.chrom_idx:
                own = self.black_k
            else:
                _diffs = [abs((ang - self.pal_ang[j] + 180) % 360 - 180)
                          for j in self.chrom_idx]
                own = self.chrom_idx[int(np.argmin(_diffs))]
            below = []
            for dy in range(3, 15):
                yy = y + dy
                if yy > y1p:
                    break
                for x in range(xc - 1, xc + 2):
                    if self.assign[yy, x] >= 0:
                        below.append(self.assign[yy, x])
            bl = Counter(below).most_common(1)[0][0] if below else None
            return own, bl

        final = {}
        for g in groups:
            xc = int(np.mean(g))
            y0 = min(cols[x][0] for x in g)
            y1 = max(cols[x][1] for x in g)
            found = {}
            for y in range(max(y0p + 2, y0 - 3), min(y1p - 2, y1 + 3)):
                centers = [x for x in range(xc - 1, xc + 2) if self.ink[y, x]]
                if not centers:                       # cap must cross the stem
                    continue
                cx = centers[len(centers) // 2]
                left = cx
                while left - 1 >= x0p and self.ink[y, left - 1]:
                    left -= 1
                right = cx
                while right + 1 <= x1p and self.ink[y, right + 1]:
                    right += 1
                w = right - left + 1
                if not (4 <= w <= 13):
                    continue
                if abs((cx - left) - (right - cx)) > 2:      # symmetric
                    continue
                above = sum(1 for x in range(left, right + 1) if self.ink[y - 2, x])
                below = sum(1 for x in range(left, right + 1) if self.ink[y + 2, x])
                _gapmax = 1 if getattr(self, '_stem_level', 5) >= 5 else 3
                if above > _gapmax or below > 3:             # vertically isolated / topmost
                    continue
                if y > y1p - 4:                              # skip x-axis
                    continue
                own, bl = classify(y, left, right, cx)
                if bl is not None and own == bl:             # color-consistent
                    if own not in found or y < found[own][0]:   # topmost per color
                        found[own] = (y, left, right)
            final[xc] = found
        self.caps = final
        return final

    def graft_caps(self):
        """Graft each validated cap's pixels into its color mask, removing the
        fragment from whatever wrong mask it was in."""
        x0p, y0p, x1p, y1p = self.cfg.plot_area
        for xc, cd in self.caps.items():
            for k, (yc, left, right) in cd.items():
                for yy in range(yc - 1, yc + 2):
                    if not (0 <= yy < self.H):
                        continue
                    centers = [x for x in range(xc - 1, xc + 2) if self.ink[yy, x]]
                    if not centers:
                        continue
                    cx = centers[len(centers) // 2]
                    l = cx
                    while l - 1 >= x0p and self.ink[yy, l - 1] and cx - (l - 1) <= 8:
                        l -= 1
                    r = cx
                    while r + 1 <= x1p and self.ink[yy, r + 1] and (r + 1) - cx <= 8:
                        r += 1
                    for x in range(l, r + 1):
                        self.assign[yy, x] = k
        return self.assign

    # ---- DATA POINTS ------------------------------------------------------- #
    def _build_colormaps(self):
        """Vectorised precompute of per-colour membership maps (H×W bool) and their
        5×5 density maps, so _color_at / _density become O(1) array lookups instead
        of 5M+ pure-Python calls. Mirrors _color_at_slow's logic exactly."""
        H, W = self.H, self.W
        ink = self.ink
        ch = self.chroma_map
        ang = self.ang_map
        L = self.lab[:, :, 0]
        achro_ks = [j for j in range(len(self.palette)) if self.achro[j]]
        nearest_idx = None
        if len(achro_ks) > 1:
            palL = np.array([float(self.pal_lab[j, 0]) for j in achro_ks], np.float32)
            d = np.abs(palL[:, None, None] - L[None, :, :].astype(np.float32))
            nearest_idx = np.argmin(d, axis=0)          # index into achro_ks
        st = getattr(self, 'stems_confirmed', None)
        st = st.astype(bool) if st is not None else None
        merge = getattr(self, '_merge_achro', False)
        self._colormap = {}
        self._density_map = {}
        for k in range(len(self.palette)):
            if self.achro[k]:
                m = ink & (ch < 9)
                if merge and not self.is_sink[k]:
                    m = m & (L < 195)
                    if st is not None:
                        m = m & (~st)
                elif len(achro_ks) <= 1:
                    m = m & (L < 170)
                else:
                    m = m & (L < 195) & (nearest_idx == achro_ks.index(k))
            else:
                angdiff = np.abs(((ang - self.pal_ang[k] + 180) % 360) - 180)
                m = ink & (ch >= 10) & (angdiff < 38)
            self._colormap[k] = m
            self._density_map[k] = cv2.boxFilter(m.astype(np.float32), -1, (5, 5),
                                                 normalize=False)
        return self._colormap

    def _color_at(self, y, x, k):
        if y < 0 or y >= self.H or x < 0 or x >= self.W:
            return False
        cm = getattr(self, '_colormap', None)
        if cm is not None and k in cm:
            return bool(cm[k][y, x])
        return self._color_at_slow(y, x, k)

    def _color_at_slow(self, y, x, k):
        if y < 0 or y >= self.H or x < 0 or x >= self.W:
            return False
        if not self.ink[y, x]:
            return False
        ch = self.chroma_map[y, x]
        ang = self.ang_map[y, x]
        L = self.lab[y, x, 0]
        if self.achro[k]:
            if ch >= 9:
                return False
            # approach 2: when there is exactly ONE achromatic CURVE (+ a sink),
            # that curve is often FRAGMENTED (its dark open-marker falls in the
            # sink, the grey slot keeps only the error bar). Search the MERGED
            # neutral ink (grey + sink) minus confirmed error-bar stems, and let
            # curve continuity pick the marker -- so the curve is not lost.
            if getattr(self, '_merge_achro', False) and not self.is_sink[k]:
                if L >= 195:
                    return False
                _st = getattr(self, 'stems_confirmed', None)
                if _st is not None and _st[y, x]:
                    return False
                return True
            achro_ks = [j for j in range(len(self.palette)) if self.achro[j]]
            if len(achro_ks) <= 1:
                return L < 170
            # multiple achromatic slots (e.g. a grey curve + a black sink): a
            # neutral pixel belongs ONLY to the slot whose palette lightness is
            # nearest -- so black axis/error-bar ink is not counted as the grey
            # curve, and vice versa.
            if L >= 195:
                return False
            nearest = min(achro_ks, key=lambda j: abs(float(self.pal_lab[j, 0]) - L))
            return nearest == k
        return ch >= 10 and abs((ang - self.pal_ang[k] + 180) % 360 - 180) < 38

    def _density(self, y, xc, k):
        dm = getattr(self, '_density_map', None)
        if dm is not None and k in dm and 0 <= y < self.H and 0 <= xc < self.W:
            return int(dm[k][y, xc])
        return sum(1 for yy in range(y - 2, y + 3) for xx in range(xc - 2, xc + 3)
                   if 0 <= yy < self.H and 0 <= xx < self.W and self._color_at(yy, xx, k))

    def _cap_mask(self):
        m = np.zeros((self.H, self.W), bool)
        for xc, cd in self.caps.items():
            for k, (yc, l, r) in cd.items():
                for yy in range(yc - 2, yc + 3):
                    for x in range(l - 2, r + 3):
                        if 0 <= yy < self.H and 0 <= x < self.W:
                            m[yy, x] = True
        return m

    def _detect_reference_line(self):
        """Find a horizontal DASHED neutral reference line (LLOQ/BLQ/baseline)
        inside the plot: a thin y where many SHORT neutral segments (dashes) span
        a wide x-range. Returns its y (or None). Requiring the DASHED pattern
        avoids flagging a solid, near-flat grey data curve as a reference line."""
        x0, y0, x1, y1 = self.cfg.plot_area
        x0 = max(0, x0); x1 = min(self.W - 1, x1)
        y0 = max(0, y0); y1 = min(self.H - 1, y1)
        if x1 - x0 < 40:
            return None
        neutral = self.ink & (self.chroma_map < 9) & (self.lab[:, :, 0] < 170)
        plot_w = x1 - x0
        best = None; bestscore = 0
        for y in range(y0 + 4, y1 - 13):
            xs = np.where(neutral[y, x0:x1 + 1])[0]
            if len(xs) < plot_w * 0.3:
                continue
            segs = 1
            for i in range(1, len(xs)):
                if xs[i] - xs[i - 1] > 2:
                    segs += 1
            xspan = xs[-1] - xs[0]
            above = int(neutral[max(0, y - 3), x0:x1 + 1].sum())
            below = int(neutral[min(self.H - 1, y + 3), x0:x1 + 1].sum())
            thin = (above < len(xs) * 0.4) and (below < len(xs) * 0.4)
            if xspan >= plot_w * 0.6 and segs >= 6 and thin and len(xs) > bestscore:
                bestscore = len(xs); best = y
        return best

    def _vrun_at(self, y, xc, k):
        """Contiguous vertical run of colour k through (xc, round(y)). A compact
        MARKER has a short run; a tall ERROR-BAR column has a long run. Used to tell
        a real (strong) marker from a strong error-bar blob in the outlier test."""
        y = int(round(y))
        if not (0 <= y < self.H) or not self._color_at(y, xc, k):
            return 0
        run = 1; yy = y - 1
        while yy >= 0 and self._color_at(yy, xc, k):
            run += 1; yy -= 1
        yy = y + 1
        while yy < self.H and self._color_at(yy, xc, k):
            run += 1; yy += 1
        return run

    def _marker_centroid(self, xc, y, k, capmask, rx=6, ry=6):
        """Refine an integer density peak (xc, y) to a sub-pixel marker centre by
        taking the colour-k pixels in a small box around it, weighted by how close
        each pixel is to the palette colour (Lab distance). Returns cy (float) or
        None. This centres open markers (rings) on their middle and damps error-bar
        / overlap contamination via the colour weighting and the bounded box."""
        x0 = max(0, xc - rx); x1 = min(self.W, xc + rx + 1)
        y0 = max(0, int(y) - ry); y1 = min(self.H, int(y) + ry + 1)
        palk = self.pal_lab[k].astype(np.float32)
        ys = []; ws = []
        for yy in range(y0, y1):
            for xx in range(x0, x1):
                if capmask[yy, xx]:
                    continue
                if self._color_at(yy, xx, k):
                    d = float(np.linalg.norm(self.lab[yy, xx].astype(np.float32) - palk))
                    ys.append(yy); ws.append(1.0 / (1.0 + d))
        if not ws:
            return None
        return float(np.average(ys, weights=np.array(ws)))

    def _seg_y(self, k, x, ref, capmask):
        """curve y at column x (excl stem/cap), cluster nearest to ref."""
        x0p, y0p, x1p, y1p = self.cfg.plot_area
        y0p = max(y0p, getattr(self.cfg, 'data_top', y0p))
        if x < 0 or x >= self.W:          # guard image boundary
            return None
        _lq = getattr(self, 'lloq_y', None)
        _ay = int(getattr(self.cfg, 'axis_y', y1p))
        _st = getattr(self, 'stems_confirmed', self.stems)
        yend = min(_ay - 2, self.H)
        cm = getattr(self, '_colormap', None)
        if cm is not None and k in cm and yend > y0p:
            col = cm[k][y0p:yend, x].copy()                 # vectorised column
            col &= ~_st[y0p:yend, x].astype(bool)
            col &= ~capmask[y0p:yend, x].astype(bool)
            if _lq is not None:
                lo = max(y0p, _lq - 3); hi = min(yend, _lq + 4)
                if hi > lo:
                    col[lo - y0p:hi - y0p] = False
            ys = (np.where(col)[0] + y0p).tolist()
        else:
            ys = [y for y in range(y0p, yend)
                  if self._color_at(y, x, k) and not (_st[y, x] or capmask[y, x])
                  and (_lq is None or abs(y - _lq) > 3)]
        if not ys:
            return None
        ys = sorted(ys)
        cl = [[ys[0]]]
        for yy in ys[1:]:
            if yy - cl[-1][-1] <= 4:
                cl[-1].append(yy)
            else:
                cl.append([yy])
        meds = [np.median(c) for c in cl]
        return min(meds, key=lambda m: abs(m - ref)) if ref is not None else np.median(meds)

    def extract_data_points(self):
        """Per data column & color: 2-D density blob for visible markers, and
        left/right segment extrapolation to recover z-order-hidden markers."""
        x0p, y0p, x1p, y1p = self.cfg.plot_area
        y0p = max(y0p, getattr(self.cfg, 'data_top', y0p))    # below a top legend band
        # single achromatic CURVE (+ sink): merge grey+sink ink so its fragmented
        # marker is recoverable via continuity (see _color_at). Determine this
        # BEFORE building the vectorised colour maps so they reflect it.
        _achc = [k for k in range(len(self.palette))
                 if self.achro[k] and not getattr(self, 'is_sink', np.zeros(len(self.palette), bool))[k]]
        self._merge_achro = (len(_achc) == 1)
        if self._merge_achro:
            print("  [single achromatic curve: merging grey+sink ink for continuity]")
        self._build_colormaps()                               # vectorised membership + density
        capmask = self._cap_mask()
        # neutral dashed reference line (LLOQ/BLQ) to exclude from achromatic search
        self.lloq_y = self._detect_reference_line()
        if self.lloq_y is not None:
            print(f"  [reference line (LLOQ/BLQ) at y={self.lloq_y} excluded from data]")

        # 1) density estimate (visible markers)
        density = {}
        strength = {}
        for xc, xv in zip(self.cfg.stem_x, self.cfg.x_values):
            density[xv] = {}
            strength[xv] = {}
            for k in range(len(self.palette)):
                if self.achro[k]:
                    # dark/neutral curve: search below its cap (avoid error bar).
                    # Exclude only the actual axis line (up to 2 px above the
                    # detected axis row) instead of cutting a wide 13 px band, so
                    # real low data points near the axis are preserved.
                    cap_y = self.caps.get(xc, {}).get(k, (None,))[0]
                    lo = (cap_y + 5) if cap_y else y0p
                    _ay = getattr(self.cfg, 'axis_y', y1p)
                    cands = [y for y in range(lo, min(_ay - 2, self.H))
                             if self.lloq_y is None or abs(y - self.lloq_y) > 3]
                    best = (max(cands, key=lambda y: self._density(y, xc, k))
                            if cands else None)
                    strg = self._density(best, xc, k) if best is not None else 0
                    if (best is not None and strg >= 5):
                        cen = self._marker_centroid(xc, best, k, capmask)
                        density[xv][k] = cen if cen is not None else float(best)
                        strength[xv][k] = strg
                    else:
                        density[xv][k] = None
                        strength[xv][k] = 0
                else:
                    bi = max(range(y0p, y1p), key=lambda y: self._density(y, xc, k))
                    sc = self._density(bi, xc, k)
                    if sc >= 5:
                        cen = self._marker_centroid(xc, bi, k, capmask)
                        density[xv][k] = cen if cen is not None else float(bi)
                        strength[xv][k] = sc
                    else:
                        density[xv][k] = None      # colour absent in this column
                        strength[xv][k] = 0

        # 2) segment extrapolation, using neighbor data points as continuity ref.
        #    Also OVERRIDE a density estimate that is an outlier vs its neighbors
        #    (that's a z-order-hidden marker whose blob got captured by an error
        #    bar / another curve) -- recover it purely from left/right segments.
        def neighbor_ref(idx, k):
            neigh = [density[self.cfg.x_values[j]][k]
                     for j in (idx - 1, idx + 1)
                     if 0 <= j < len(self.cfg.x_values)
                     and density[self.cfg.x_values[j]][k] is not None]
            return int(np.mean(neigh)) if neigh else None

        # continuity-based points for the single merged achromatic curve
        _cont = {}
        _achc_k = None
        if getattr(self, '_merge_achro', False):
            _achc = [kk for kk in range(len(self.palette))
                     if self.achro[kk] and not self.is_sink[kk]]
            if len(_achc) == 1:
                _achc_k = _achc[0]
                _cont = self._achro_continuity(_achc_k, capmask)

        # per-colour span of columns where the colour is actually present. Outside
        # this span the curve has not started / has ended (e.g. dropped below LLOQ)
        # -> emit NO point instead of fabricating one at y=0 or by extrapolation.
        _present = {}
        for k in range(len(self.palette)):
            idxs = [i for i, xv in enumerate(self.cfg.x_values)
                    if density[xv][k] is not None]
            _present[k] = (min(idxs), max(idxs)) if idxs else None

        data = {}
        for idx, (xc, xv) in enumerate(zip(self.cfg.stem_x, self.cfg.x_values)):
            data[xv] = {}
            for k in range(len(self.palette)):
                if k == _achc_k and xc in _cont:
                    # achromatic curve: use CONTINUITY selection (handles fragmented
                    # open markers + large error bars) instead of density argmax
                    data[xv][k] = int(round(_cont[xc]))
                    continue
                dens = density[xv][k]
                nref = neighbor_ref(idx, k)
                # A weak/ambiguous peak that jumps far, OR a STRONG peak that sits on
                # a tall vertical ERROR BAR (long vertical run), is an outlier -> recover
                # from segments. A strong COMPACT marker that jumps (e.g. a placebo curve
                # starting at the bottom) is respected.
                _strong = strength[xv].get(k, 0) >= 8
                _errbar = (dens is not None and self._vrun_at(dens, xc, k) > 16)
                is_outlier = (dens is not None and nref is not None
                              and abs(dens - nref) > 50
                              and (not _strong or _errbar))
                if dens is None or is_outlier:
                    span = _present.get(k)
                    inside = span is not None and span[0] <= idx <= span[1]
                    if dens is None and not inside:
                        data[xv][k] = None        # colour absent here -> no point
                        continue
                    # occluded (curve exists on both sides): recover from segments
                    est = self._extrapolate(k, xc, nref, capmask)
                    data[xv][k] = est if est is not None else nref
                else:
                    # trust the density marker (visible); nudge toward segment fit
                    est = self._extrapolate(k, xc, dens, capmask)
                    if est is not None and abs(est - dens) < 20:
                        data[xv][k] = int(round((est + dens) / 2))
                    else:
                        data[xv][k] = int(round(dens))
        self.data_points = data
        return data

    def extract_curve_walk(self, nms_px=4, prom_turn=15, R=None):
        """Alternative extractor: per COLOUR, trace the curve with a multi-seed
        walk (seed_walk_nodes), then take data points as (a) the path's TURNING
        POINTS -- local y-extrema, so a vertical re-dose jump yields BOTH its
        bottom (trough) and top (peak) -- plus (b) density peaks along the path
        for markers on flat/sloped parts. Peaks closer than nms_px (= 0.5x the
        marker diameter) are merged. Populates self.walk_points = {k:[(x,y)...]}.
        Chromatic curves only; achromatic keeps the continuity extractor."""
        if R is None:
            R = WALK_DENSITY_R
        if getattr(self, '_colormap', None) is None:
            self._build_colormaps()          # ensure vectorised maps exist
        x0p, y0p, x1p, y1p = self.cfg.plot_area
        y0p = max(y0p, getattr(self.cfg, 'data_top', y0p))
        ax_rows = [int(getattr(self.cfg, 'axis_y', y1p))]
        ax_cols = [int(getattr(self.cfg, 'axis_x', x0p))]
        self.walk_points = {}
        for k in range(len(self.palette)):
            if getattr(self, 'is_sink', np.zeros(len(self.palette), bool))[k]:
                continue
            mask = (self.assign == k)
            # Exclude the LLOQ/BLQ reference-line band: it is neutral/dark so it
            # lands in the black/grey mask, and the walk would otherwise trace it
            # into a row of fake points along that horizontal line.
            _lq = getattr(self, 'lloq_y', None)
            if _lq is not None:
                mask = mask.copy()
                mask[max(0, int(_lq) - 3):min(self.H, int(_lq) + 4), :] = False
            if mask.sum() < 8:
                self.walk_points[k] = []
                continue
            nodes = seed_walk_nodes(mask.astype(np.uint8), axis_rows=ax_rows,
                                    axis_cols=ax_cols, plot_area=self.cfg.plot_area)
            if len(nodes) < 2:
                self.walk_points[k] = []
                continue
            xs = np.array([n[0] for n in nodes]); ys = np.array([n[1] for n in nodes])
            xr = np.arange(xs.min(), xs.max() + 1)
            yp = np.interp(xr, xs, ys)
            # turning points: troughs (max y) and peaks (min y) of the path
            tmax, _ = find_peaks(yp, prominence=prom_turn)
            tmin, _ = find_peaks(-yp, prominence=prom_turn)
            # density along path (pixels within +/-R of the path y)
            colpx = np.zeros(self.W)
            for i, x in enumerate(xr):
                col = np.where(mask[:, x])[0]
                if len(col):
                    c = yp[i]
                    if np.min(np.abs(col - c)) > R:
                        c = float(np.median(col))
                    colpx[x] = np.sum(np.abs(col - c) <= R)
            comb = (colpx / (colpx.max() + 1e-9)) ** 2
            dpk, _ = find_peaks(comb, prominence=0.15, distance=max(1, nms_px))
            xset = (set(int(xr[i]) for i in set(list(tmax) + list(tmin)))
                    | set(int(x) for x in dpk)
                    | {int(xs.min()), int(xs.max())})
            pts = sorted((x, int(round(np.interp(x, xr, yp)))) for x in xset)
            # NMS merge across the combined set (turning + density)
            merged = []
            for x, y in pts:
                if merged and abs(x - merged[-1][0]) < nms_px:
                    continue
                merged.append((x, y))
            self.walk_points[k] = merged
        return self.walk_points

    def _achro_clusters(self, xc, k, capmask):
        """y-clusters of the (merged) achromatic mask at column xc, excluding
        stems/caps/lloq/axis. Returns [(median_y, size), ...]."""
        x0p, y0p, x1p, y1p = self.cfg.plot_area
        y0p = max(y0p, getattr(self.cfg, 'data_top', y0p))
        ay = getattr(self.cfg, 'axis_y', y1p)
        _lq = getattr(self, 'lloq_y', None)
        ys = [y for y in range(y0p, min(ay - 2, self.H))
              if self._color_at(y, xc, k) and not capmask[y, xc]
              and (_lq is None or abs(y - _lq) > 3)]
        if not ys:
            return []
        ys.sort()
        cl = [[ys[0]]]
        for yy in ys[1:]:
            (cl[-1].append(yy) if yy - cl[-1][-1] <= 4 else cl.append([yy]))
        return [(float(np.median(c)), len(c)) for c in cl]

    def _achro_continuity(self, k, capmask):
        """Pick the achromatic curve point per column by CURVE CONTINUITY: the
        y-cluster that best connects the curve (nearest to a robust local trend),
        NOT the densest blob -- so an error-bar spike or a stray black cluster that
        breaks continuity is rejected in favour of the marker that continues it."""
        cols = list(self.cfg.stem_x)
        clusters = {xc: self._achro_clusters(xc, k, capmask) for xc in cols}
        allmed = [m for xc in cols for (m, s) in clusters[xc]]
        if not allmed:
            return {}
        M = float(np.median(allmed))

        def _neutral_ink(y, xc):
            if not (0 <= y < self.H):
                return False
            for xx in range(max(0, xc - 2), min(self.W, xc + 3)):
                if (self.ink[y, xx] and self.chroma_map[y, xx] < 9
                        and self.lab[y, xx, 0] < 195):
                    return True
            return False

        # seed: nearest cluster to the global median
        p = {}
        for xc in cols:
            cs = clusters[xc]
            if cs:
                p[xc] = min(cs, key=lambda ms: abs(ms[0] - M))[0]
        # iterate: each column snaps to the cluster nearest to its NEIGHBOURS'
        # current estimate (self excluded). If no stem-removed cluster is near that
        # continuity reference, the marker is probably EMBEDDED in the error-bar
        # column (stem removal deleted it) -> snap to neutral ink at y~ref instead.
        for _ in range(5):
            newp = {}
            for i, xc in enumerate(cols):
                neigh = [p[cols[j]] for j in (i - 1, i + 1)
                         if 0 <= j < len(cols) and cols[j] in p]
                ref = float(np.median(neigh)) if neigh else M
                cs = clusters[xc]
                best = min(cs, key=lambda ms: abs(ms[0] - ref))[0] if cs else None
                if best is None or abs(best - ref) > 15:
                    band = [y for y in range(int(ref) - 6, int(ref) + 7)
                            if _neutral_ink(y, xc)]
                    if band:
                        best = float(np.median(band))
                if best is not None:
                    newp[xc] = best
            p = newp
        return p

    def _extrapolate(self, k, xc, ref, capmask, span=14):
        left, right = [], []
        for dx in range(3, span):
            yl = self._seg_y(k, xc - dx, ref, capmask)
            if yl is not None:
                left.append((xc - dx, yl))
            yr = self._seg_y(k, xc + dx, ref, capmask)
            if yr is not None:
                right.append((xc + dx, yr))
        # reject segment points that sit far from the continuity ref -- those are
        # error-bar / overlap remnants, not the curve (e.g. a stem stub at dx=3).
        if ref is not None:
            left = [p for p in left if abs(p[1] - ref) <= 20]
            right = [p for p in right if abs(p[1] - ref) <= 20]
        preds = []
        for seg in (left, right):
            if len(seg) >= 3:
                xs = np.array([p[0] for p in seg])
                ys = np.array([p[1] for p in seg])
                A = np.polyfit(xs, ys, 1)
                preds.append(A[0] * xc + A[1])
        return int(np.mean(preds)) if preds else None

    # ---- value conversion -------------------------------------------------- #
    def to_values(self):
        """Convert data-point y pixels to axis values using cfg.y_ref."""
        if not self.cfg.y_ref:
            return None
        (p0, v0), (p1, v1) = self.cfg.y_ref
        def conv(py):
            return v0 + (py - p0) * (v1 - v0) / (p1 - p0)
        out = {}
        for xv, d in self.data_points.items():
            out[xv] = {self.cfg.curve_names[k]: (round(conv(y), 2) if y is not None else None)
                       for k, y in d.items()}
        return out

    # ---- orchestration ----------------------------------------------------- #
    def run(self):
        self.stage1()
        self.stage2()
        self.detect_stems()
        self.detect_caps()
        self.graft_caps()
        self.extract_data_points()
        return self

    # ---- visualization ----------------------------------------------------- #
    def render_masks(self, path, scale=1.7):
        names = self.cfg.curve_names
        pal = self.palette
        def grid(vis):
            o = vis.copy()
            for pct in range(0, 101, 10):
                x = int(self.W * pct / 100); y = int(self.H * pct / 100)
                c = (150, 150, 255) if pct % 50 == 0 else (215, 215, 215)
                cv2.line(o, (x, 0), (x, self.H - 1), c, 1)
                cv2.line(o, (0, y), (self.W - 1, y), c, 1)
            return cv2.addWeighted(o, 0.6, vis, 0.4, 0)
        def panel(mask, label, fc):
            vis = np.full((self.H, self.W, 3), 255, np.uint8)
            vis[mask] = fc
            vis = grid(vis)
            big = cv2.resize(vis, (int(self.W * scale), int(self.H * scale)),
                             interpolation=cv2.INTER_NEAREST)
            ban = np.full((22, big.shape[1], 3), 255, np.uint8)
            cv2.putText(ban, label, (4, 15), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 0, 0), 1)
            return np.vstack([ban, big])
        un = self.ink & (self.assign < 0)
        ps = [panel(self.assign == k, f'{names[k]}: {int((self.assign==k).sum())}px',
                    (pal[k][2], pal[k][1], pal[k][0])) for k in range(len(self.palette))]
        ps.append(panel(un, f'UNASSIGNED: {int(un.sum())}px', (0, 140, 255)))
        blank = np.full_like(ps[0], 255)
        grid_img = np.vstack([np.hstack([ps[0], ps[1]]),
                              np.hstack([ps[2], ps[3]]),
                              np.hstack([ps[4], blank])])
        cv2.imwrite(path, grid_img)

    def render_datapoints(self, path, scale=2.5):
        draw = {0: (230, 120, 50), 1: (40, 160, 40), 2: (30, 30, 220), 3: (0, 0, 0)}
        vis = self.img.copy()
        for xv, xc in zip(self.cfg.x_values, self.cfg.stem_x):
            for k in range(len(self.palette)):
                y = self.data_points[xv][k]
                if y is not None:
                    cv2.circle(vis, (xc, y), 3, draw[k], -1)
        big = cv2.resize(vis, None, fx=scale, fy=scale, interpolation=cv2.INTER_NEAREST)
        for k in range(len(self.palette)):
            pts = [(int(self.cfg.stem_x[i] * scale),
                    int(self.data_points[self.cfg.x_values[i]][k] * scale))
                   for i in range(len(self.cfg.x_values))
                   if self.data_points[self.cfg.x_values[i]][k] is not None]
            for i in range(1, len(pts)):
                cv2.line(big, pts[i - 1], pts[i], draw[k], 1)
        cv2.imwrite(path, big)


# --------------------------------------------------------------------------- #
# ---- end embedded plot_digitizer core ----


# --- configure a digitizer that reuses v24's plot-area, legend and palette ---
_pa = PLOT_AREA  # (x0, y0, x1, y1)
_cfg = PlotConfig()
_lb = tuple(int(v) for v in LEGEND_BOX) if LEGEND_BOX else None

# --- FIX: keep the plot-area from swallowing the legend --------------------
# v24's axis detector sometimes returns y0=0 (or an edge that overlaps the
# legend panel). If the legend sits along the top/bottom of the plot area and
# spans most of its width, trim the plot area so the legend's own swatch/marker
# ink can't be mistaken for data stems / points.
_pax0, _pay0, _pax1, _pay1 = int(_pa[0]), int(_pa[1]), int(_pa[2]), int(_pa[3])
_extra_exclude = None
if _lb:
    _lx0, _ly0, _lx1, _ly1 = _lb
    _pa_w = max(1, _pax1 - _pax0)
    _pa_h = max(1, _pay1 - _pay0)
    _overlap_w = max(0, min(_pax1, _lx1) - max(_pax0, _lx0))
    _wide = _overlap_w > 0.5 * _pa_w
    _lb_cy = 0.5 * (_ly0 + _ly1)
    # Legend inside the plot region: _build_ink_mask already drops the legend box
    # itself. But when the axis detector set the plot-area top to y=0, the TITLE
    # text ABOVE the legend is still inside the plot area and its vertical strokes
    # get picked up as fake stems. If the legend sits in the top third, also
    # exclude everything from the plot top down to the legend box (the title band).
    if _wide and _lb_cy < _pay0 + _pa_h * 0.33 and _ly0 > _pay0 + 2:
        _extra_exclude = (_pax0, _pay0, _pax1, _ly0 - 1)   # title band above legend
        print(f"  [excluding title band above legend: "
              f"y {_pay0}-{_ly0 - 1}]")
    # symmetric case: legend in bottom third -> exclude the band below it
    elif _wide and _lb_cy > _pay1 - _pa_h * 0.33 and _ly1 < _pay1 - 2:
        _extra_exclude = (_pax0, _ly1 + 1, _pax1, _pay1)
        print(f"  [excluding band below legend: y {_ly1 + 1}-{_pay1}]")
_cfg.plot_area = (_pax0, _pay0, _pax1, _pay1)
_cfg.legend_box = _lb
if (_pay0, _pay1) != (int(_pa[1]), int(_pa[3])):
    print(f"  [plot-area trimmed to exclude legend: y {int(_pa[1])}-{int(_pa[3])} "
          f"-> {_pay0}-{_pay1}]")

_dig = PlotDigitizer.__new__(PlotDigitizer)   # build without re-reading file
_dig.cfg = _cfg
_dig.img = img
_dig.H, _dig.W = img.shape[:2]
_dig.rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
_dig.lab = cv2.cvtColor(_dig.rgb, cv2.COLOR_RGB2Lab).astype(np.float32)
_hsv = cv2.cvtColor(_dig.rgb, cv2.COLOR_RGB2HSV)
_dig.s = _hsv[:, :, 1].astype(np.float32)
_dig.v = _hsv[:, :, 2].astype(np.float32)
_dig.a_map = _dig.lab[:, :, 1] - 128
_dig.b_map = _dig.lab[:, :, 2] - 128
_dig.chroma_map = np.hypot(_dig.a_map, _dig.b_map)
_dig.ang_map = np.degrees(np.arctan2(_dig.b_map, _dig.a_map))
_dig.assign = None
_dig.stems = None
_dig.caps = {}
_dig.data_points = {}
_dig._build_ink_mask()
# apply the extra title/legend-band exclusion computed above, if any
if _extra_exclude is not None:
    _ex0, _ey0, _ex1, _ey1 = _extra_exclude
    _dig.ink[max(0, _ey0):_ey1 + 1, max(0, _ex0):_ex1 + 1] = False
    _dig.inplot[max(0, _ey0):_ey1 + 1, max(0, _ex0):_ex1 + 1] = False

# palette + curve names straight from v24's discovered COLORS (legend order)
_names = [cd['name'] for cd in COLORS]
_rgbs = [tuple(int(c) for c in cd['mean_rgb']) for cd in COLORS]
# v38: decide ACHROMATIC (grey/black) entries from the LEGEND STRUCTURE, not from
# a plot-based guess (which was fooled by black error bars -> phantom black curve).
# Any achromatic swatch actually present in the legend grid is appended here; then
# add_black_if_missing is disabled so set_palette never invents a black curve.
# Dedup against the LEGEND's own chromatic swatch colours (from the grid), not the
# digitizer's re-measured palette -- a curve that the digitizer mis-measured as
# grey must not swallow a genuine black legend entry.
_grid_now = globals().get('_LAST_LEGEND_GRID')
_chrom_legend = [tuple(c[2]) for c in _grid_now['cells']] if _grid_now and _grid_now.get('cells') else _rgbs
_ach_from_legend = _classify_legend_swatches(_lb, existing_rgbs=_chrom_legend)
for _ac in _ach_from_legend:
    _rgbs.append(_ac)
    _names.append('black' if (sum(_ac) / 3.0) < 90 else 'grey')
if _ach_from_legend:
    print(f"  [legend achromatic swatches: {_ach_from_legend}]")
# unified STRICT table: chromatic + achromatic swatches snapped to (col,row)
# intersections, empty cells kept as empty, column/row alignment verified.
_unified = _build_unified_legend_grid(
    (_grid_now or {}).get('cells'), globals().get('_LAST_ACHRO_SWATCHES', []))
globals()['_LAST_UNIFIED_GRID'] = _unified
if _unified is not None:
    print(f"  [unified legend grid: {len(_unified['cols'])} cols x {len(_unified['rows'])} rows, "
          f"{len(_unified['cells'])} filled / {len(_unified['empty'])} empty, "
          f"aligned={_unified['aligned']} (residual={_unified['residual']}px)]")
try:
    with open(os.path.join(OUT_DIR, 'legend_grid.json'), 'w') as _gf:
        json.dump({'legend_box': _lb, 'grid': globals().get('_LAST_LEGEND_GRID'),
                   'achro': globals().get('_LAST_ACHRO_SWATCHES', []),
                   'unified': _unified}, _gf)
except Exception:
    pass
_cfg.curve_names = list(_names)

# v39: When the unified legend grid is reliable, the number of CURVES must equal
# the number of legend colours. Rebuild the extraction palette straight from the
# grid's filled cells (deduping any near-duplicate the raw swatch pass produced),
# and DON'T add a black sink if the grid already contains an achromatic cell (the
# black marker IS a real curve, not a sink). This stops the extractor from
# re-discovering extra colours (open/dashed markers, LLOQ) beyond the legend.
_add_sink = True
if _unified is not None and _unified.get('cells') and globals().get('_LEGEND_1TO1'):
    _cell_items = sorted(_unified['cells'].items(),
                         key=lambda kv: (int(kv[0].split(',')[1]), int(kv[0].split(',')[0])))
    _grid_rgbs = []
    _grid_names = []
    _has_achro = False
    def _close(a, b, t=18):
        return (abs(a[0]-b[0]) + abs(a[1]-b[1]) + abs(a[2]-b[2])) <= t
    for _key, _cell in _cell_items:
        _rgb = tuple(int(v) for v in _cell['rgb'])
        if any(_close(_rgb, g) for g in _grid_rgbs):    # merge near-duplicates
            continue
        _grid_rgbs.append(_rgb)
        if _cell.get('type') == 'achro':
            _has_achro = True
            _grid_names.append('black' if (sum(_rgb) / 3.0) < 90 else 'grey')
        else:
            _grid_names.append(f'color{len(_grid_rgbs)+1:02d}')
    if len(_grid_rgbs) >= LEGEND_MIN_ENTRIES:
        _rgbs = _grid_rgbs
        _names = _grid_names
        _cfg.curve_names = list(_names)
        _add_sink = not _has_achro
        print(f"  [extraction palette locked to legend grid: {len(_rgbs)} colours"
              f"{' (achro present -> no sink)' if _has_achro else ''}]")

_dig.set_palette(_rgbs, add_black_if_missing=False, add_black_sink=_add_sink)
# resync the working name list to whatever the digitizer ended up with.
_names = list(_dig.cfg.curve_names)
# Keep COLORS consistent with the (possibly re-locked) palette so downstream code
# that iterates COLORS and looks up _names.index(cd['name']) never desyncs. Build
# one COLORS entry per digitizer palette slot, carrying over any prior metadata.
_prev_by_name = {cd.get('name'): cd for cd in COLORS}
_pal_rgbs = getattr(_dig, "palette", None)
_new_colors = []
for _i, _nm in enumerate(_names):
    _cd = _prev_by_name.get(_nm, {})
    if _pal_rgbs is not None and _i < len(_pal_rgbs):
        _mrgb = [int(v) for v in _pal_rgbs[_i]]
    else:
        _mrgb = _cd.get('mean_rgb', [0, 0, 0])
    _new_colors.append({'name': _nm, 'mean_rgb': _mrgb,
                        'px_count': _cd.get('px_count', 0)})
COLORS = _new_colors

# data-point columns (stems) auto-detected from the ORIGINAL by the digitizer.
# Give the digitizer the ACTUAL axis positions first, so stem/cap detection stays
# strictly INSIDE the axes (the y-axis line itself must not become a false stem).
_cfg.axis_y = int(AXIS_ROWS.max()) if len(AXIS_ROWS) else _cfg.plot_area[3]
_cfg.axis_x = int(AXIS_COLS.max()) if len(AXIS_COLS) else _cfg.plot_area[0]
# If a legend sits as a BAND across the TOP (no top axis to bound the plot, so the
# auto plot-area swallowed the legend), push the data-region top below it, so the
# legend swatches are not mistaken for markers / curve.
_cfg.data_top = _cfg.plot_area[1]
if _lb is not None:
    _lx0, _ly0, _lx1, _ly1 = _lb
    if _ly0 <= _cfg.plot_area[1] + 4 and _ly1 < 0.5 * H:      # a top band
        _cfg.data_top = int(_ly1) + 2
        print(f"  [data-region top pushed below top legend -> y0={_cfg.data_top}]")
_dig._stem_level = int(os.environ.get('STEM_LEVEL', '5'))
_dig.detect_stems()
_groups, _stemcols = _dig._stem_groups()
_stem_x = sorted(int(np.mean(g)) for g in _groups)
_cfg.stem_x = _stem_x
_cfg.x_values = list(range(len(_stem_x)))     # pixel-index; calibration maps later

# --- run the core pipeline ---------------------------------------------------
_dig.stage1()
_dig.stage2()
# Apply confusable-pair resolution to the FINAL label map (stage1/stage2 classify
# by plain nearest-centroid and would otherwise mix brightness-similar colours in
# the masks the walk / colour-mask output use). Compute the confusable pairs on
# the ACTUAL extraction palette, then re-split those pairs by hue on self.assign.
try:
    _final_lab = [cv2.cvtColor(np.uint8([[list(rgb)]]), cv2.COLOR_RGB2Lab)[0, 0].astype(np.float32)
                  for rgb in _rgbs]
    globals()['_LEGEND_SWATCH_INFO'] = globals().get('_LEGEND_SWATCH_INFO', [])
    _final_bias = _confusable_priority(_final_lab)
    _final_pairs = globals().get('_LAST_CONFUSABLE_PAIRS', [])
    # Only re-split pairs that are genuinely CLOSE in Lab on the final masks. The
    # priority routine also flags hue-adjacent-but-far pairs (useful as a bias in
    # discovery) but re-splitting those on muddy JPEG palettes can shuffle pixels
    # the wrong way, so gate the FINAL re-split tightly.
    def _pair_de(a, b):
        return float(np.linalg.norm(np.array(_final_lab[a]) - np.array(_final_lab[b])))
    _final_pairs = [(i, j) for (i, j) in _final_pairs if _pair_de(i, j) < 40.0]
    if _final_pairs:
        _dig.refine_confusable(_final_lab, _final_bias, _final_pairs)
        print(f"  [confusable resolved on final masks: {len(_final_pairs)} close pair(s)]")
except Exception as _e:
    print(f"  [confusable refine skipped: {_e}]")

# ---- DEBUG: dump per-colour Stage-1 (seeds) and Stage-2 (assign) masks -------
try:
    _sdir = os.path.join(OUT_DIR, 'stage_masks')
    os.makedirs(_sdir, exist_ok=True)
    for _old in os.listdir(_sdir):                       # clear stale masks first
        if _old.startswith(('stage1_', 'stage2_')) and _old.endswith('.png'):
            try:
                os.remove(os.path.join(_sdir, _old))
            except OSError:
                pass
    _pal = _dig.palette
    _nm = list(_dig.cfg.curve_names) if getattr(_dig.cfg, 'curve_names', None) else []
    print(f"  [stage-dump palette size={len(_pal)} names={_nm}]")
    for _stage_name, _lblmap in (('stage1', _dig.seeds), ('stage2', _dig.assign)):
        if _lblmap is None:
            continue
        for _k in range(len(_pal)):
            _m = (_lblmap == _k)
            _canvas = np.full((_dig.H, _dig.W, 3), 255, np.uint8)
            _col = (int(_pal[_k][2]), int(_pal[_k][1]), int(_pal[_k][0]))   # BGR
            _canvas[_m] = _col
            _lab = _nm[_k] if _k < len(_nm) else f'color{_k:02d}'
            cv2.putText(_canvas, f'{_stage_name} {_lab} ({int(_m.sum())}px)',
                        (5, 18), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 1)
            cv2.imwrite(os.path.join(_sdir, f'{_stage_name}_{_k:02d}.png'), _canvas)
    print(f"  [stage masks dumped -> {_sdir}]")
    try:
        np.save(os.path.join(OUT_DIR, 'assign.npy'), _dig.assign)
    except Exception:
        pass
except Exception as _e:
    print(f"  [stage mask dump skipped: {_e}]")
_dig.detect_stems()
_dig.detect_caps()
_dig.graft_caps()
# ---- DEBUG: stem + Tcap estimation overlay ---------------------------------
try:
    _sc = cv2.imread(IMG_PATH).copy()
    # axis lines (green) for reference
    if len(AXIS_COLS):
        cv2.line(_sc, (int(AXIS_COLS.max()), 0), (int(AXIS_COLS.max()), H), (0, 200, 0), 1)
    if len(AXIS_ROWS):
        cv2.line(_sc, (0, int(AXIS_ROWS.max())), (W, int(AXIS_ROWS.max())), (0, 200, 0), 1)
    _sy, _sx = np.where(_dig.stems.astype(bool))
    _sc[_sy, _sx] = (255, 255, 0)                      # permissive stems (columns) -> cyan
    _cf = getattr(_dig, 'stems_confirmed', None)
    if _cf is not None:
        _cy, _cx = np.where(_cf.astype(bool))
        _sc[_cy, _cx] = (255, 0, 0)                    # confirmed error bars -> blue
    for _xc, _cd in _dig.caps.items():
        for _k, (_yy, _l, _r) in _cd.items():
            cv2.line(_sc, (int(_l), int(_yy)), (int(_r), int(_yy)), (0, 0, 255), 1)  # cap -> red
            cv2.circle(_sc, (int(_xc), int(_yy)), 2, (255, 0, 255), -1)              # stem_x -> magenta
    cv2.imwrite(os.path.join(OUT_DIR, 'stem_cap_overlay.png'), _sc)
    _ncap = sum(len(v) for v in _dig.caps.values())
    print(f"  [stem/cap overlay: {int(_dig.stems.sum())} stem px, {_ncap} Tcaps -> stem_cap_overlay.png]")
except Exception as _e:
    print(f"  [stem/cap overlay skipped: {_e}]")
_dig.extract_data_points()
# Optional alternative extractor (per-colour curve walk: turning points + density
# peaks, NMS = 0.5x marker diameter). Enable with EXTRACT_MODE=walk.
_EXTRACT_MODE = os.environ.get('EXTRACT_MODE', 'walk')
if _EXTRACT_MODE == 'walk':
    # Marker diameter, measured from the LEGEND swatches: at each legend cell
    # centre, the vertical extent of the non-white swatch ink is the marker's
    # height. This ties the walk NMS window to the plot's actual marker size
    # instead of a fixed guess, so densely-spaced markers on big plots aren't
    # merged and sparse markers on small plots aren't split.
    def _measure_marker_height():
        _grid = globals().get('_LAST_UNIFIED_GRID')
        _cells = []
        _rows_y = []
        if _grid and _grid.get('cols') and _grid.get('rows'):
            _rows_y = sorted(int(r) for r in _grid['rows'])
            for _cx in _grid['cols']:
                for _cy in _grid['rows']:
                    _cells.append((int(_cx), int(_cy)))
        else:
            for _c in (globals().get('_LAST_LEGEND_GRID') or {}).get('cells', []):
                _cells.append((int(_c[0]), int(_c[1])))
        if not _cells:
            return None
        # cap the vertical half-window to half the row spacing minus 1, so the
        # measurement of one swatch never bleeds into the row above/below.
        _yhalf = 12
        if len(_rows_y) > 1:
            _gap = min(_rows_y[i+1] - _rows_y[i] for i in range(len(_rows_y)-1))
            _yhalf = max(3, min(12, _gap // 2 - 1))
        _rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        _hs = []
        for (_cx, _cy) in _cells:
            _x0 = max(0, _cx - 8); _x1 = min(W, _cx + 9)      # narrow: marker, not label
            _y0 = max(0, _cy - _yhalf); _y1 = min(H, _cy + _yhalf + 1)
            _win = _rgb[_y0:_y1, _x0:_x1]
            if _win.size == 0:
                continue
            _nw = ~((_win[:, :, 0] > 235) & (_win[:, :, 1] > 235) & (_win[:, :, 2] > 235))
            _rows = np.where(_nw.any(axis=1))[0]
            if len(_rows):
                _hs.append(int(_rows.max() - _rows.min() + 1))
        if not _hs:
            return None
        return int(np.median(_hs))

    _measured = None
    if 'MARKER_DIAM' in os.environ:
        _mdiam = int(os.environ['MARKER_DIAM'])          # explicit override wins
    else:
        _measured = _measure_marker_height()
        _mdiam = _measured if (_measured and _measured >= 3) else 8
    _nms = max(1, round(_mdiam * 0.5))
    _dig.extract_curve_walk(nms_px=_nms)
    _src_tag = ("measured from legend swatches" if _measured
                else ("env override" if 'MARKER_DIAM' in os.environ else "default"))
    print(f"  [extract mode = WALK | marker diam = {_mdiam}px ({_src_tag}) "
          f"| NMS = {_nms}px (0.5x marker diam)]")

# --- emit v24-format structures ---------------------------------------------
all_detections = {}
all_tcaps_out = {}
all_results = {}

for _k, _cname in enumerate(_names):
    # skip the black SINK slot (absorbs error-bar/axis ink; not a data curve)
    if getattr(_dig, 'is_sink', None) is not None and _k < len(_dig.is_sink) and _dig.is_sink[_k]:
        continue
    _dets = []
    if _EXTRACT_MODE == 'walk' and getattr(_dig, 'walk_points', None) is not None:
        # curve-walk points: a list of (x, y) per colour (multiple per column ok)
        for _wx, _wy in _dig.walk_points.get(_k, []):
            _dets.append({'x': int(_wx), 'y': int(_wy),
                          'fitness': 1.0, 'px': 1, 'source': 'walk'})
        _dets.sort(key=lambda d: d['x'])
        all_detections[_cname] = _dets
        continue
    # default: shared stem_x columns x density (one point per column)
    for _xc, _xv in zip(_cfg.stem_x, _cfg.x_values):
        _py = _dig.data_points.get(_xv, {}).get(_k)
        if _py is None:
            continue
        # --- FIX: ignore points that landed on the x-axis or y-axis lines ----
        # (axis ink can survive cleaning and get picked up as a spurious marker)
        _on_axis = False
        try:
            for _ar in (list(AXIS_ROWS) if 'AXIS_ROWS' in dir() and AXIS_ROWS else []):
                if abs(int(_py) - int(_ar)) <= 2:
                    _on_axis = True
                    break
            if not _on_axis:
                for _ac in (list(AXIS_COLS) if 'AXIS_COLS' in dir() and AXIS_COLS else []):
                    if abs(int(_xc) - int(_ac)) <= 2:
                        _on_axis = True
                        break
        except Exception:
            _on_axis = False
        if _on_axis:
            continue
        _dets.append({'x': int(_xc), 'y': int(_py),
                      'fitness': 1.0, 'px': 1, 'source': 'digitizer'})
    _dets.sort(key=lambda d: d['x'])
    all_detections[_cname] = _dets

    # tcaps -> list of {'x','y','type'} (upper caps only in these plots)
    _tc = []
    for _xc, _cd in _dig.caps.items():
        if _k in _cd:
            _y, _l, _r = _cd[_k]
            _tc.append({'x': int(_xc), 'y': int(_y), 'type': 'upper',
                        'area': int(_r - _l + 1)})
    all_tcaps_out[_cname] = _tc

    # walk_nodes -> the traced curve as [[x,y],...]; use the per-column mask centres
    _mask_k = (_dig.assign == _k)
    _nodes = []
    for _x in range(_cfg.plot_area[0], _cfg.plot_area[2] + 1):
        _col = np.where(_mask_k[:, _x])[0]
        if len(_col):
            _nodes.append((int(_x), int(np.median(_col))))
    all_results[_cname] = (_nodes, 1.0)

    print(f"  {_cname}: {len(_dets)} data_points  {len(_tc)} tcaps")

# also expose the label->clean-mask so downstream overlays that read
# cd['_raw_mask_canvas'] keep working (attach current stage-2 mask)
for cd in COLORS:
    _k = _names.index(cd['name'])
    cd['_raw_mask_canvas'] = (_dig.assign == _k).astype(np.uint8) * 255
    cd['_clean_mask'] = (_dig.assign == _k)

# Post-filter: remove colours with 0 detections
all_detections_out = {k: v for k, v in all_detections.items() if len(v) > 0}
removed = [k for k in all_detections if k not in all_detections_out]
if removed:
    print(f"\n  Post-filter: removed {removed}")

# Collect walk_nodes
all_walk_nodes = {}
for cname, (nodes, cov, *_) in all_results.items():
    if cname in all_detections_out:
        all_walk_nodes[cname] = [[int(p[0]), int(p[1])] for p in nodes]

# Write JSON
json_path = os.path.join(OUT_DIR, 'detections.json')
with open(json_path, 'w') as f:
    json.dump({
        'detections': all_detections_out,
        'tcaps':      {k: all_tcaps_out.get(k, []) for k in all_detections_out},
        'walk_nodes': all_walk_nodes,
    }, f, indent=2)
print(f"\nJSON: {json_path}")

# -- Default outputs: Excel of data points + data-point overlay image ----------
# Uses the tick-calibration module with the axes/plot-area already computed here
# (no pipeline re-run), then maps every detected point pixel->data coordinate,
# records which curve (colour) it belongs to, and renders an overlay.
try:
    _cal = calibrate_from_axes(img, list(AXIS_ROWS), list(AXIS_COLS), PLOT_AREA)
except Exception as _e:
    _cal = None
    print(f"  [calibration skipped: {_e}]")

_x2v = _cal['x2v'] if _cal else None
_y2v = _cal['y2v'] if _cal else None
_coords = _cal['coords'] if _cal else None

# Manual axis-range override. If the user gave any of x_min/x_max/y_min/y_max
# (and/or log flags), build the pixel->value mapping directly from the plot-area
# edges. Any endpoint the user did NOT give is filled in from the OCR result so
# partial manual entry still works. Left edge->x_min, right->x_max,
# bottom->y_min, top->y_max.
def _make_p2v(px_lo, px_hi, v_lo, v_hi, is_log):
    if px_hi == px_lo or v_lo is None or v_hi is None:
        return None, None
    if is_log:
        if v_lo <= 0 or v_hi <= 0:
            return None, None
        import math as _m
        a = (_m.log10(v_hi) - _m.log10(v_lo)) / (px_hi - px_lo)
        b = _m.log10(v_lo) - a * px_lo
        return (lambda px: float(10 ** (a * px + b))), 'log'
    a = (v_hi - v_lo) / (px_hi - px_lo)
    b = v_lo - a * px_lo
    return (lambda px: float(a * px + b)), 'linear'

_has_manual = (any(v is not None for v in
                    (USER_X_MIN, USER_X_MAX, USER_Y_MIN, USER_Y_MAX))
               or USER_X_LOG or USER_Y_LOG)
if PLOT_AREA is not None and _has_manual:
    _pa_x0, _pa_y0, _pa_x1, _pa_y1 = PLOT_AREA
    _oc = _coords or {}
    # x: left edge = min, right edge = max
    _xlo = USER_X_MIN if USER_X_MIN is not None else _oc.get('x_min')
    _xhi = USER_X_MAX if USER_X_MAX is not None else _oc.get('x_max')
    _xlog = USER_X_LOG or (_oc.get('x_kind') == 'log')
    # y: bottom edge (larger pixel y) = min, top edge (smaller pixel y) = max
    _ylo = USER_Y_MIN if USER_Y_MIN is not None else _oc.get('y_min')
    _yhi = USER_Y_MAX if USER_Y_MAX is not None else _oc.get('y_max')
    _ylog = USER_Y_LOG or (_oc.get('y_kind') == 'log')

    _nx2v, _xkind = _make_p2v(_pa_x0, _pa_x1, _xlo, _xhi, _xlog)
    _ny2v, _ykind = _make_p2v(_pa_y1, _pa_y0, _ylo, _yhi, _ylog)
    # Apply each axis independently: only replace an axis's mapping if we could
    # build it; otherwise keep whatever OCR gave for that axis.
    _user_x = (USER_X_MIN is not None or USER_X_MAX is not None or USER_X_LOG)
    _user_y = (USER_Y_MIN is not None or USER_Y_MAX is not None or USER_Y_LOG)
    if _user_x and _nx2v is not None:
        _x2v = _nx2v
    if _user_y and _ny2v is not None:
        _y2v = _ny2v
    # Rebuild coords summary from whatever mappings we now have.
    if _x2v is not None and _y2v is not None:
        _fx_lo = _xlo if (_user_x and _nx2v is not None) else (_oc.get('x_min'))
        _fx_hi = _xhi if (_user_x and _nx2v is not None) else (_oc.get('x_max'))
        _fy_lo = _ylo if (_user_y and _ny2v is not None) else (_oc.get('y_min'))
        _fy_hi = _yhi if (_user_y and _ny2v is not None) else (_oc.get('y_max'))
        _fx_kind = _xkind if (_user_x and _nx2v is not None) else _oc.get('x_kind', 'linear')
        _fy_kind = _ykind if (_user_y and _ny2v is not None) else _oc.get('y_kind', 'linear')
        _coords = {'x_min': _fx_lo, 'x_max': _fx_hi, 'y_min': _fy_lo, 'y_max': _fy_hi,
                   'x_kind': _fx_kind, 'y_kind': _fy_kind}
        print(f"  Calibration: X[{_fx_lo},{_fx_hi}] {_fx_kind}  "
              f"Y[{_fy_lo},{_fy_hi}] {_fy_kind}  (manual where given)")
    else:
        print("  [manual calibration incomplete; using OCR values where available]")

# median colour per curve (sampled at its own points) -> identifies the curve
_curve_rgb = {}
for _cname, _pts in all_detections_out.items():
    _samp = []
    for _p in _pts:
        _x, _y = int(_p['x']), int(_p['y'])
        _patch = img[max(0, _y-2):_y+3, max(0, _x-2):_x+3].reshape(-1, 3)
        _samp.extend(_patch.tolist())
    if _samp:
        _m = np.median(_samp, axis=0)
        _curve_rgb[_cname] = (int(_m[2]), int(_m[1]), int(_m[0]))  # BGR->RGB
    else:
        _curve_rgb[_cname] = (0, 0, 0)

def _hex(rgb):
    return '#%02X%02X%02X' % (int(rgb[0]), int(rgb[1]), int(rgb[2]))

# ---- Curve labels from the legend text (swatch -> label to the right) ---------
# _LEGEND_SWATCH_INFO is index-aligned with the palette: colorNN corresponds to
# swatch_info[N-2] (color02 -> index 0). For each colour we OCR the strip of the
# legend just to the RIGHT of its swatch to get the printed label. Falls back to
# the colorNN id when no legend text is available or OCR is empty.
def _curve_labels():
    labels = {_c: _c for _c in all_detections_out.keys()}   # default colorNN
    if not _HAS_OCR:
        return labels
    # Prefer the user-drawn legend box; else the detected one.
    lb = USER_LEGEND_BOX if USER_LEGEND_BOX is not None else \
         globals().get('_LEGEND_BOX_FOR_CLEAN', None)
    if lb is None:
        return labels
    lx0, ly0, lx1, ly1 = [int(v) for v in lb]
    lx0 = max(0, min(lx0, W - 1)); lx1 = max(0, min(lx1, W))
    ly0 = max(0, min(ly0, H - 1)); ly1 = max(0, min(ly1, H))
    if lx1 - lx0 < 8 or ly1 - ly0 < 6:
        return labels
    panel = img[ly0:ly1, lx0:lx1]

    # Find legend rows: horizontal bands that contain ink (swatch + text).
    g = cv2.cvtColor(panel, cv2.COLOR_BGR2GRAY)
    ink = (g < 200).astype(np.uint8)
    rowsum = ink.sum(axis=1)
    rows = np.where(rowsum > 2)[0]
    if len(rows) == 0:
        return labels
    bands = []
    cur = [rows[0]]
    for r in rows[1:]:
        if r - cur[-1] <= 3:
            cur.append(r)
        else:
            bands.append((cur[0], cur[-1])); cur = [r]
    bands.append((cur[0], cur[-1]))

    # Detected curve colours in RGB, to match swatches against.
    curve_names = list(all_detections_out.keys())
    curve_rgbs = {c: _curve_rgb[c] for c in curve_names}

    for (ry0, ry1) in bands:
        if ry1 - ry0 < 4:
            continue
        band = panel[ry0:ry1+1, :]
        bh, bw = band.shape[:2]
        # The swatch is the left-most coloured (chromatic or dark) block. Find the
        # left region's dominant colour, then the label is the text to its right.
        bhsv = cv2.cvtColor(band, cv2.COLOR_BGR2HSV)
        colmask = (bhsv[:, :, 1] > 60) | (cv2.cvtColor(band, cv2.COLOR_BGR2GRAY) < 110)
        cols_with_ink = np.where(colmask.any(axis=0))[0]
        if len(cols_with_ink) == 0:
            continue
        swatch_x0 = cols_with_ink[0]
        # swatch assumed to end where a white gap of >=3px begins
        swatch_x1 = swatch_x0
        run_gap = 0
        for x in range(swatch_x0, bw):
            if colmask[:, x].any():
                swatch_x1 = x; run_gap = 0
            else:
                run_gap += 1
                if run_gap >= 3 and x - swatch_x0 > 3:
                    break
        # swatch colour = median of coloured pixels in that block
        block = band[:, swatch_x0:swatch_x1+1].reshape(-1, 3)
        bm = block[np.any(block < 240, axis=1)]
        if bm.size == 0:
            continue
        med = np.median(bm, axis=0)
        srgb = (int(med[2]), int(med[1]), int(med[0]))
        # nearest curve colour
        best_c, best_d = None, 1e9
        for c in curve_names:
            r_, g_, b_ = curve_rgbs[c]
            d = (r_-srgb[0])**2 + (g_-srgb[1])**2 + (b_-srgb[2])**2
            if d < best_d:
                best_d, best_c = d, c
        if best_c is None or best_d > 90**2:      # too far -> no confident match
            continue
        # OCR the text to the right of the swatch
        tx0 = swatch_x1 + 2
        if bw - tx0 < 6:
            continue
        strip = band[:, tx0:]
        sg = cv2.cvtColor(strip, cv2.COLOR_BGR2GRAY)
        sg = cv2.resize(sg, None, fx=3, fy=3, interpolation=cv2.INTER_CUBIC)
        _, sg = cv2.threshold(sg, 150, 255, cv2.THRESH_BINARY)
        try:
            txt = ' '.join(pytesseract.image_to_string(sg, config='--psm 7').split())
        except Exception:
            txt = ''
        if len(txt) >= 2:
            labels[best_c] = txt
    return labels

_CURVE_LABEL = _curve_labels()

# ---- Coordinate fallback: normalized [0,1] over the plot area ----------------
# If neither manual ranges nor OCR gave a calibration, express every point as a
# fraction of the plot area (x: left->0, right->1; y: bottom->0, top->1) so the
# data file always has usable data_x/data_y instead of raw pixels.
_norm_fallback = False
if (_x2v is None or _y2v is None) and PLOT_AREA is not None:
    _pfx0, _pfy0, _pfx1, _pfy1 = PLOT_AREA
    if _pfx1 != _pfx0 and _pfy1 != _pfy0:
        if _x2v is None:
            _x2v = (lambda px, a=_pfx0, b=_pfx1: float((px - a) / (b - a)))
        if _y2v is None:
            # bottom edge (_pfy1) -> 0, top edge (_pfy0) -> 1
            _y2v = (lambda py, a=_pfy1, b=_pfy0: float((py - a) / (b - a)))
        _norm_fallback = True
        _coords = {'x_min': 0.0, 'x_max': 1.0, 'y_min': 0.0, 'y_max': 1.0,
                   'x_kind': 'normalized', 'y_kind': 'normalized'}
        print("  Calibration: none -> normalized [0,1] over plot area")

# Build the Excel workbook
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    _wb = openpyxl.Workbook(); _ws = _wb.active; _ws.title = 'data_points'
    _hdr = ['curve', 'data_x', 'data_y']
    _ws.append(_hdr)
    for _c in range(1, len(_hdr)+1):
        _ws.cell(1, _c).font = Font(bold=True, color='FFFFFF')
        _ws.cell(1, _c).fill = PatternFill('solid', start_color='305496')
    _rownum = 1
    for _cname, _pts in all_detections_out.items():
        _label = _CURVE_LABEL.get(_cname, _cname)
        _rgb = _curve_rgb.get(_cname, (0, 0, 0))
        _hexcol = '{:02X}{:02X}{:02X}'.format(
            max(0, min(255, _rgb[0])), max(0, min(255, _rgb[1])), max(0, min(255, _rgb[2])))
        # white text if the swatch is dark, else black, for readability
        _lum = 0.299*_rgb[0] + 0.587*_rgb[1] + 0.114*_rgb[2]
        _txtcol = 'FFFFFF' if _lum < 140 else '000000'
        for _p in _pts:
            _px, _py = _p['x'], _p['y']
            _dx = _x2v(_px) if _x2v else None
            _dy = _y2v(_py) if _y2v else None
            _ws.append([_label,
                        round(_dx, 4) if _dx is not None else None,
                        round(_dy, 6) if _dy is not None else None])
            _rownum += 1
            _cell = _ws.cell(_rownum, 1)
            _cell.fill = PatternFill('solid', start_color=_hexcol)
            _cell.font = Font(color=_txtcol)
    _ws.append([])
    if _coords and _coords.get('x_kind') == 'normalized':
        _ws.append(['calibration', 'normalized [0,1] over plot area (no axis values)'])
    elif _coords:
        _ws.append(['calibration',
                    f"X[{_coords['x_min']:.4g},{_coords['x_max']:.4g}] {_coords['x_kind']}",
                    f"Y[{_coords['y_min']:.4g},{_coords['y_max']:.4g}] {_coords['y_kind']}"])
    else:
        _ws.append(['calibration', 'not available'])
    _ws.column_dimensions['A'].width = 26
    _ws.column_dimensions['B'].width = 14
    _ws.column_dimensions['C'].width = 14
    _xlsx_path = os.path.join(OUT_DIR, 'data_points.xlsx')
    _wb.save(_xlsx_path)
    print(f"Excel: {_xlsx_path}")
except Exception as _e:
    print(f"  [Excel skipped: {_e}]")

# edit_data.json : everything the GUI needs to let a human correct points and
# recompute values -- pixel points per curve, the legend label + colour, and a
# serialisable calibration (two pixel<->value anchors + log flags) so the browser
# can convert any (edited) pixel to a data value without re-running the pipeline.
try:
    _ed = {
        'image': {'width': int(W), 'height': int(H)},
        'plot_area': [int(v) for v in PLOT_AREA] if PLOT_AREA else None,
        'calibration': None,
        'curves': [],
    }
    if PLOT_AREA is not None and _x2v is not None and _y2v is not None:
        _pax0, _pay0, _pax1, _pay1 = PLOT_AREA
        _xk = (_coords or {}).get('x_kind', 'linear')
        _yk = (_coords or {}).get('y_kind', 'linear')
        _ed['calibration'] = {
            'x': {'p0': int(_pax0), 'p1': int(_pax1),
                  'v0': float(_x2v(_pax0)), 'v1': float(_x2v(_pax1)),
                  'log': (_xk == 'log'), 'kind': _xk},
            'y': {'p0': int(_pay1), 'p1': int(_pay0),        # bottom, top
                  'v0': float(_y2v(_pay1)), 'v1': float(_y2v(_pay0)),
                  'log': (_yk == 'log'), 'kind': _yk},
        }
    for _cname, _pts in all_detections_out.items():
        _rgb = _curve_rgb.get(_cname, (0, 0, 0))
        _ed['curves'].append({
            'name': _cname,
            'label': _CURVE_LABEL.get(_cname, _cname),
            'rgb': [int(_rgb[0]), int(_rgb[1]), int(_rgb[2])],
            'points': [{'x': int(_p['x']), 'y': int(_p['y'])} for _p in _pts],
        })
    with open(os.path.join(OUT_DIR, 'edit_data.json'), 'w') as _ef:
        json.dump(_ed, _ef)
    print("  [edit data -> edit_data.json]")
except Exception as _e:
    print(f"  [edit data skipped: {_e}]")

# Data-point overlay image (each point drawn in its curve colour)
try:
    _ov = img.copy()
    for _cname, _pts in all_detections_out.items():
        _rgb = _curve_rgb[_cname]; _bgr = (_rgb[2], _rgb[1], _rgb[0])
        for _p in _pts:
            _x, _y = int(_p['x']), int(_p['y'])
            cv2.circle(_ov, (_x, _y), 5, (0, 0, 0), -1)
            cv2.circle(_ov, (_x, _y), 3, _bgr, -1)
            cv2.circle(_ov, (_x, _y), 6, (0, 0, 0), 1)
    _ov_path = os.path.join(OUT_DIR, 'data_points_overlay.png')
    cv2.imwrite(_ov_path, _ov)
    print(f"Overlay: {_ov_path}")
except Exception as _e:
    print(f"  [Overlay skipped: {_e}]")

# Per-colour mask images: each curve's assigned pixels painted in their ORIGINAL
# colour on white, with that curve's detected data points overlaid. One PNG per
# colour (colormask_00_<name>.png ...) so the GUI can show every mask.
try:
    _asg = getattr(_dig, 'assign', None)
    if _asg is not None:
        # Each mask (assign==k) is, by construction, the pixels nearest palette
        # colour k, and the extraction palette was LOCKED to the legend grid. So
        # the legend swatch colour for mask k IS palette[k] -- use it directly so
        # the chip is always aligned with its mask (the old index-zip against a
        # differently-ordered raw-grid list mismatched them).
        _pal = getattr(_dig, 'palette', None)
        _mask_manifest = []
        for _k, _cname in enumerate(_names):
            if getattr(_dig, 'is_sink', None) is not None and _k < len(_dig.is_sink) and _dig.is_sink[_k]:
                continue
            _m = (_asg == _k)
            _canvas = np.full((H, W, 3), 255, np.uint8)
            _canvas[_m] = img[_m]                     # original pixel colours
            _pts = all_detections_out.get(_cname, [])
            for _p in _pts:                           # draw this colour's data points
                _x, _y = int(_p['x']), int(_p['y'])
                cv2.circle(_canvas, (_x, _y), 5, (255, 255, 255), -1)
                cv2.circle(_canvas, (_x, _y), 5, (0, 0, 0), 2)
            _fname = f'colormask_{_k:02d}_{_cname}.png'
            cv2.imwrite(os.path.join(OUT_DIR, _fname), _canvas)
            _lrgb = ([int(v) for v in _pal[_k]]
                     if (_pal is not None and _k < len(_pal)) else None)
            _mask_manifest.append({'name': _cname, 'file': _fname,
                                   'mask_px': int(_m.sum()), 'points': len(_pts),
                                   'legend_rgb': _lrgb})
        with open(os.path.join(OUT_DIR, 'colormasks.json'), 'w') as _mf:
            json.dump(_mask_manifest, _mf)
        print(f"  [colour masks: {len(_mask_manifest)} -> colormask_*.png]")

        # LEGEND-TABLE overlay drawn on the ORIGINAL plot (like the data-point
        # overlay): legend box, column/row grid lines, filled cells (in their
        # legend colour) and empty cells (grey X).
        try:
            _ug = globals().get('_LAST_LEGEND_GRID') or {}
            _uni = globals().get('_LAST_UNIFIED_GRID')
        except Exception:
            _uni = None
        try:
            _lo = img.copy()
            if _lb is not None:
                cv2.rectangle(_lo, (_lb[0], _lb[1]), (_lb[2], _lb[3]), (255, 120, 0), 2)
            _grid_src = _uni if _uni else _ug
            if _grid_src and _grid_src.get('cols') and _grid_src.get('rows'):
                _cols = _grid_src['cols']; _rows = _grid_src['rows']
                _y0 = _lb[1] if _lb else min(_rows); _y1 = _lb[3] if _lb else max(_rows)
                _x0 = _lb[0] if _lb else min(_cols); _x1 = _lb[2] if _lb else max(_cols)
                for _cx in _cols:
                    cv2.line(_lo, (int(_cx), int(_y0)), (int(_cx), int(_y1)), (0, 180, 0), 1)
                for _ry in _rows:
                    cv2.line(_lo, (int(_x0), int(_ry)), (int(_x1), int(_ry)), (255, 0, 255), 1)
                # cells: unified stores {'cells':{'ci,ri':{type,rgb}}, 'empty':[...]}
                if _uni and isinstance(_uni.get('cells'), dict):
                    for _key, _cell in _uni['cells'].items():
                        _cix, _rix = map(int, _key.split(','))
                        _cx, _cy = int(_cols[_cix]), int(_rows[_rix])
                        _rgb = _cell.get('rgb', [0, 0, 0])
                        _bgr = (int(_rgb[2]), int(_rgb[1]), int(_rgb[0]))
                        cv2.rectangle(_lo, (_cx-7, _cy-7), (_cx+7, _cy+7), _bgr, -1)
                        cv2.rectangle(_lo, (_cx-7, _cy-7), (_cx+7, _cy+7), (0, 0, 0), 1)
                    for _key in _uni.get('empty', []):
                        _cix, _rix = map(int, _key.split(','))
                        _cx, _cy = int(_cols[_cix]), int(_rows[_rix])
                        cv2.drawMarker(_lo, (_cx, _cy), (140, 140, 140), cv2.MARKER_TILTED_CROSS, 12, 1)
            cv2.imwrite(os.path.join(OUT_DIR, 'legend_overlay.png'), _lo)
            print("  [legend overlay -> legend_overlay.png]")
        except Exception as _e:
            print(f"  [legend overlay skipped: {_e}]")

        # LEGEND DIAGNOSTIC image: shows -- inside the legend box -- every swatch
        # CANDIDATE and why it was kept or dropped, so failures are visible on the
        # local machine WITHOUT reading the log. Colour code:
        #   green  = accepted chromatic swatch
        #   cyan   = accepted achromatic swatch (grey/black key)
        #   orange = rejected: near-neutral (chroma <= 24) in the chromatic pass
        #   yellow = rejected: no adjacent text label found
        #   red    = rejected: size/shape out of range (too big/small/long)
        try:
            if _lb is not None:
                _hx0, _hy0, _hx1, _hy1 = _lb
                _hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
                _sc = _hsv[:, :, 1]; _vc = _hsv[:, :, 2]
                _rgbf = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                _keep = np.zeros(img.shape[:2], np.uint8)
                _keep[max(0, _hy0):_hy1+1, max(0, _hx0):_hx1+1] = 1
                _chm = (((_sc > 40) & (_vc > 40) & (_vc < 250)).astype(np.uint8)) & _keep
                _acm = (((_sc <= 40) & (_vc < 220) & (_vc > 25)).astype(np.uint8)) & _keep
                _dbg = img.copy()
                cv2.rectangle(_dbg, (_hx0, _hy0), (_hx1, _hy1), (255, 120, 0), 2)

                def _judge(mask, achro):
                    mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, np.ones((3, 3), np.uint8))
                    n, lbl, st, cen = cv2.connectedComponentsWithStats(mask, 8)
                    for i in range(1, n):
                        a = st[i, cv2.CC_STAT_AREA]; w = st[i, cv2.CC_STAT_WIDTH]; h = st[i, cv2.CC_STAT_HEIGHT]
                        x, y = st[i, cv2.CC_STAT_LEFT], st[i, cv2.CC_STAT_TOP]
                        cx, cy = cen[i]
                        if not (15 <= a <= 2500 and w <= 120 and h <= 60):
                            col = (0, 0, 255)                 # red: size/shape
                        else:
                            tx0, tx1 = int(cx+10), int(min(W, cx+160))
                            ty0, ty1 = int(max(0, cy-12)), int(min(H, cy+12))
                            txt = int(((_vc[ty0:ty1, tx0:tx1] < 120) & (_sc[ty0:ty1, tx0:tx1] < 60)).sum()) if tx1 > tx0 else 0
                            rgb = tuple(int(z) for z in np.median(_rgbf[lbl == i], axis=0))
                            if txt < 100:
                                col = (0, 220, 255)           # yellow: no text label
                            elif (not achro) and (max(rgb)-min(rgb)) <= 24:
                                col = (0, 150, 255)           # orange: near-neutral
                            else:
                                col = (0, 200, 0) if not achro else (255, 200, 0)  # green/cyan accepted
                        cv2.rectangle(_dbg, (x, y), (x+w, y+h), col, 1)
                _judge(_chm, achro=False)
                _judge(_acm, achro=True)
                # overlay the final grid cells/lines if a unified grid was built
                _u = globals().get('_LAST_UNIFIED_GRID')
                if _u and _u.get('cols') and _u.get('rows'):
                    for _cx in _u['cols']:
                        cv2.line(_dbg, (int(_cx), _hy0), (int(_cx), _hy1), (0, 180, 0), 1)
                    for _ry in _u['rows']:
                        cv2.line(_dbg, (_hx0, int(_ry)), (_hx1, int(_ry)), (255, 0, 255), 1)
                # legend key text
                _yk = _hy1 + 16
                for _txt, _c in [("green=chrom OK", (0, 200, 0)), ("cyan=achrom OK", (255, 200, 0)),
                                 ("orange=neutral", (0, 150, 255)), ("yellow=no label", (0, 220, 255)),
                                 ("red=bad size", (0, 0, 255))]:
                    cv2.putText(_dbg, _txt, (_hx0, _yk), cv2.FONT_HERSHEY_SIMPLEX, 0.4, _c, 1)
                    _yk += 14
                cv2.imwrite(os.path.join(OUT_DIR, 'legend_diagnostic.png'), _dbg)
                print("  [legend diagnostic -> legend_diagnostic.png]")
        except Exception as _e:
            print(f"  [legend diagnostic skipped: {_e}]")
except Exception as _e:
    print(f"  [colour masks skipped: {_e}]")

# Combined overlay figure (DEBUG-only: matplotlib text layout for every point's
# fitness label is by far the slowest single step; skip unless DEBUG_DUMPS is set)
if os.environ.get('DEBUG_DUMPS'):
    fig, ax = plt.subplots(figsize=(max(8, W/60), max(5, H/60)))
    ax.imshow(img_rgb)
    ax.set_title('Point detection -- all colours', fontsize=11)
    for cname, dets in all_detections_out.items():
        for det in dets:
            ax.plot(det['x'], det['y'], 'o', color='red', ms=5,
                    markeredgecolor='white', markeredgewidth=0.4, zorder=10)
            ax.text(det['x'], det['y']-7, f"{det.get('fitness',0):.2f}",
                    color='yellow', fontsize=4, ha='center', zorder=11)
    ax.axis('off')
    fig.savefig(os.path.join(OUT_DIR, 'points_combined.jpg'), dpi=150, bbox_inches='tight')
    plt.close(fig)

# Color mask montage (all colors side-by-side with data points overlaid)
_mask_panels = []
for cd in COLORS:
    cname = cd['name']
    if cname not in all_detections_out:
        continue
    mean_rgb = cd.get('mean_rgb', (128, 128, 128))
    raw = cd.get('_raw_mask')
    if raw is None:
        continue
    # Build raw mask panel: white bg, color pixels in their color
    panel_mask = np.full((H, W, 3), 255, dtype=np.uint8)
    panel_mask[raw == 1] = list(mean_rgb)
    panel_mask = cv2.cvtColor(panel_mask, cv2.COLOR_RGB2BGR)
    # Overlay union walk path with gap bridges as vertical lines
    nodes_c = all_walk_nodes.get(cname, [])
    if len(nodes_c) > 1:
        # Draw connected segments; for gaps (x jump > 3), draw vertical bridge
        for i in range(len(nodes_c) - 1):
            x0n, y0n = nodes_c[i]
            x1n, y1n = nodes_c[i + 1]
            if abs(x1n - x0n) <= 3:
                # Continuous segment
                cv2.line(panel_mask, (x0n, y0n), (x1n, y1n), (0, 180, 0), 2)
            else:
                # Gap: draw vertical bridge in orange at the gap boundary
                # Left side: vertical line at x0n
                cv2.line(panel_mask, (x0n, y0n), (x1n, y1n), (0, 140, 255), 1)
    if nodes_c:
        sx, sy = nodes_c[0]
        ex, ey = nodes_c[-1]
        cv2.circle(panel_mask, (sx, sy), 9, (0, 255, 0), -1)    # green = left end
        cv2.circle(panel_mask, (sx, sy), 9, (255, 255, 255), 2)
        cv2.putText(panel_mask, 'L', (sx + 11, sy + 5), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (0, 200, 0), 2)
        cv2.circle(panel_mask, (ex, ey), 9, (0, 220, 255), -1)  # cyan = right end
        cv2.circle(panel_mask, (ex, ey), 9, (255, 255, 255), 2)
        cv2.putText(panel_mask, 'R', (ex + 11, ey + 5), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (0, 200, 220), 2)
    # Overlay detected data points as red circles
    for det in all_detections_out.get(cname, []):
        cv2.circle(panel_mask, (int(det['x']), int(det['y'])), 7, (0, 0, 220), -1)
        cv2.circle(panel_mask, (int(det['x']), int(det['y'])), 7, (255, 255, 255), 1)
    label = f'{cname}  pts={len(all_detections_out.get(cname, []))}  L=green R=cyan'
    cv2.putText(panel_mask, label, (4, 18), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (0,0,0), 3)
    cv2.putText(panel_mask, label, (4, 18), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (255,255,255), 1)
    scale = 380 / W
    panel_mask = cv2.resize(panel_mask, (380, int(H * scale)), interpolation=cv2.INTER_AREA)
    _mask_panels.append(panel_mask)
if _mask_panels:
    _ncols = 3; _nrows = (_ncols + len(_mask_panels) - 1) // _ncols
    _nrows = (len(_mask_panels) + _ncols - 1) // _ncols
    _ph, _pw = _mask_panels[0].shape[:2]
    _grid = np.full((_nrows * _ph + 44, _ncols * _pw, 3), 255, dtype=np.uint8)
    cv2.putText(_grid, 'Raw masks + detected points', (6, 30),
                cv2.FONT_HERSHEY_SIMPLEX, 0.6, (30, 30, 30), 1)
    for _idx, _panel in enumerate(_mask_panels):
        _r, _c = divmod(_idx, _ncols)
        _grid[44 + _r * _ph:44 + _r * _ph + _ph, _c * _pw:_c * _pw + _pw] = _panel
    cv2.imwrite(os.path.join(OUT_DIR, 'color_masks_montage.jpg'), _grid,
                [cv2.IMWRITE_JPEG_QUALITY, 93])

# Montage
generate_montage(OUT_DIR, img, all_detections_out, all_tcaps_out, all_results)

# -- Aligned density montage (raw mask + x-col density + walk density, x-aligned) --
def save_aligned_density_montage(out_dir, colors_list, all_det_out, all_walk_nodes_dict, all_results_dict, all_wm_full_dict=None):
    """
    For each detected colour, produce a 3-row panel:
      Row 1: raw mask image with walk path + detected points overlaid
      Row 2: x-column pixel count (col_px)  -- uses wm_full (same as estimate_data_points)
      Row 3: walk path density (marker_density) + combined signal + peak markers
    All rows share the same x-axis range so they are pixel-aligned.
    All per-colour panels are stacked vertically into one figure.
    """
    import matplotlib.gridspec as gridspec
    from scipy.signal import find_peaks as _fp

    active = [(cd, all_det_out[cd['name']], all_walk_nodes_dict.get(cd['name'], []))
              for cd in colors_list if cd['name'] in all_det_out]
    if not active:
        return

    n = len(active)
    fig_h = n * 4.5  # 4.5 inches per colour
    fig, axes = plt.subplots(n * 3, 1, figsize=(max(10, W / 60), fig_h),
                             gridspec_kw={'height_ratios': [3, 1, 1] * n})
    if n == 1:
        axes = list(axes)

    for ci, (cd, dets, nodes) in enumerate(active):
        cname = cd['name']
        mean_rgb = cd.get('mean_rgb', (128, 128, 128))
        raw = cd.get('_raw_mask')
        ax_img  = axes[ci * 3]
        ax_col  = axes[ci * 3 + 1]
        ax_dens = axes[ci * 3 + 2]

        # -- Row 1: raw mask + walk + detections ------------------------------
        if raw is not None:
            panel_rgb = np.full((H, W, 3), 255, dtype=np.uint8)
            r_c = int(mean_rgb[0]); g_c = int(mean_rgb[1]); b_c = int(mean_rgb[2])
            panel_rgb[raw == 1] = [r_c, g_c, b_c]
            ax_img.imshow(panel_rgb, aspect='auto')
        else:
            ax_img.set_facecolor('#f0f0f0')
        # Walk path
        if len(nodes) > 1:
            xs_n = [p[0] for p in nodes]
            ys_n = [p[1] for p in nodes]
            ax_img.plot(xs_n, ys_n, '-', color='green', lw=1.5, alpha=0.8, zorder=3)
            ax_img.plot(xs_n[0],  ys_n[0],  'o', color='lime',   ms=8, zorder=4)
            ax_img.plot(xs_n[-1], ys_n[-1], 'o', color='cyan',   ms=8, zorder=4)
        # Detected points
        for det in dets:
            ax_img.plot(det['x'], det['y'], 'o', color='red', ms=7,
                        markeredgecolor='white', markeredgewidth=0.8, zorder=5)
        ax_img.set_xlim(0, W)
        ax_img.set_ylim(H, 0)
        ax_img.set_ylabel(f"{cname}\nmask", fontsize=7)
        ax_img.tick_params(labelbottom=False, bottom=False)
        ax_img.set_title(f"{cname}  pts={len(dets)}  L=lime R=cyan", fontsize=8, loc='left')

        # -- Compute signals for rows 2 & 3 -----------------------------------
        # Use wm_full (axis-removed clean mask, same as estimate_data_points) for signal
        # Fall back to raw mask if wm_full not available
        wm = (all_wm_full_dict or {}).get(cname)
        full_m = wm if wm is not None else (raw if raw is not None else np.zeros((H, W), dtype=np.uint8))
        col_px = np.array([full_m[:, x].sum() if 0 <= x < W else 0
                           for x in range(W)], dtype=float)

        marker_density = np.zeros(W, dtype=float)
        if len(nodes) > 1:
            xs_n2 = [p[0] for p in nodes]
            ys_n2 = [p[1] for p in nodes]
            x_min_w = int(min(xs_n2)); x_max_w = int(max(xs_n2))
            walk_y_at2 = {x: float(np.interp(x, xs_n2, ys_n2))
                          for x in range(x_min_w, x_max_w + 1)}
            for x in range(x_min_w, x_max_w + 1):
                wy = walk_y_at2[x]
                ys_col = np.where(full_m[:, x] > 0)[0]
                if len(ys_col) == 0:
                    continue
                if float(np.min(np.abs(ys_col - wy))) > WALK_DENSITY_R:
                    wy = float(np.median(ys_col))
                marker_density[x] = float(np.sum(np.abs(ys_col - wy) <= WALK_DENSITY_R))

        col_norm  = col_px / (col_px.max() + 1e-9)
        mark_norm = marker_density / (marker_density.max() + 1e-9)
        combined  = col_norm * mark_norm
        xs_all = np.arange(W)

        # -- Row 2: x-column pixel count --------------------------------------
        ax_col.fill_between(xs_all, col_px, color='steelblue', alpha=0.6)
        ax_col.set_xlim(0, W)
        ax_col.set_ylabel('col px', fontsize=7)
        ax_col.tick_params(labelbottom=False, bottom=False, labelsize=6)
        ax_col.yaxis.set_tick_params(labelsize=6)
        # Mark detected x positions
        for det in dets:
            ax_col.axvline(det['x'], color='red', lw=0.8, alpha=0.7)

        # -- Row 3: marker density + combined signal + peaks ------------------
        ax_dens.fill_between(xs_all, mark_norm, color='brown', alpha=0.5, label='walk density')
        ax_dens.plot(xs_all, combined, color='darkgreen', lw=1.0, alpha=0.9, label='combined')
        # Mark peaks (detected data points)
        for det in dets:
            ax_dens.plot(det['x'], combined[det['x']] if det['x'] < W else 0,
                         '*', color='red', ms=8, zorder=5)
        ax_dens.set_xlim(0, W)
        ax_dens.set_ylabel('density', fontsize=7)
        ax_dens.set_xlabel('x (px)', fontsize=7)
        ax_dens.tick_params(labelsize=6)
        ax_dens.yaxis.set_tick_params(labelsize=6)
        ax_dens.legend(fontsize=5, loc='upper right')

        # Align x-ticks across all 3 rows
        for ax in [ax_img, ax_col, ax_dens]:
            ax.set_xlim(0, W)

    fig.suptitle('Raw mask + x-col density + walk density  (x-aligned)', fontsize=9, y=1.002)
    fig.tight_layout(h_pad=0.3)
    out_path = os.path.join(out_dir, 'aligned_density_montage.jpg')
    fig.savefig(out_path, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f"  Aligned density montage: {out_path}")

if os.environ.get('DEBUG_DUMPS'):
    save_aligned_density_montage(OUT_DIR, COLORS, all_detections_out, all_walk_nodes, all_results, all_wm_full)

total = sum(len(v) for v in all_detections_out.values())
print(f"\nTotal: {total} data points  ->  {OUT_DIR}")
print("Done.")
